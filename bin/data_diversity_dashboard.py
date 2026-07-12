#!/usr/bin/env python3
"""
G160 · bin/data_diversity_dashboard.py

Cluster E: per-cohort diversity dashboard for buyer pre-purchase verification.
Generates route_type / biome / time-of-day / action-entropy histograms.
"""

import argparse
import csv
import datetime
import json
import logging
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
log = logging.getLogger("data_diversity_dashboard")

try:
    import numpy as np

    HAS_NUMPY = True
except ImportError:
    np = None
    HAS_NUMPY = False

try:
    from PIL import Image, ImageDraw, ImageFont

    HAS_PIL = True
except ImportError:
    Image = ImageDraw = ImageFont = None
    HAS_PIL = False  # type: ignore[misc]

try:
    import yaml

    HAS_YAML = True
except ImportError:
    yaml = None
    HAS_YAML = False

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    HAS_OPENPYXL = True
except ImportError:
    openpyxl = Font = PatternFill = Alignment = None
    HAS_OPENPYXL = False  # type: ignore[misc]

_DIMENSIONS = ("route_type", "biome", "time_of_day", "action_entropy")
_PALETTE = ["#4C72B0", "#55A868", "#C44E52", "#8172B2", "#CCB974", "#64B5CD"]


def load_data(path: Path) -> List[Dict[str, Any]]:
    """Load records from JSON, YAML, or CSV."""
    suffix = path.suffix.lower()
    if suffix == ".json":
        with open(path, "r", encoding="utf-8") as fh:
            raw = json.load(fh)
        if isinstance(raw, list):
            return raw
        if isinstance(raw, dict):
            for k in ("records", "data", "rows", "items"):
                if k in raw and isinstance(raw[k], list):
                    return raw[k]
            return [raw]
        raise ValueError(f"Unexpected JSON in {path}")
    if suffix in (".yaml", ".yml"):
        if not HAS_YAML:
            raise ImportError("PyYAML required for YAML")
        with open(path, "r", encoding="utf-8") as fh:
            raw = yaml.safe_load(fh)
        if isinstance(raw, list):
            return raw
        if isinstance(raw, dict):
            for k in ("records", "data", "rows", "items"):
                if k in raw and isinstance(raw[k], list):
                    return raw[k]
            return [raw]
        raise ValueError(f"Unexpected YAML in {path}")
    if suffix == ".csv":
        rows: List[Dict[str, str]] = []
        with open(path, "r", newline="", encoding="utf-8") as fh:
            for row in csv.DictReader(fh):
                rows.append(dict(row))
        return rows
    raise ValueError(f"Unsupported format: {suffix}")


def _extract_cohorts(records: List[Dict], field: str) -> Dict[str, List[Dict]]:
    groups: Dict[str, List[Dict]] = defaultdict(list)
    for r in records:
        groups[str(r.get(field, "unknown"))].append(r)
    return dict(groups)


def _histogram(values: List[Any], bins: Optional[List[str]] = None) -> Tuple[List[str], List[int]]:
    counts: Dict[str, int] = defaultdict(int)
    for v in values:
        counts[str(v)] += 1
    labels = bins if bins else sorted(counts.keys())
    return labels, [counts.get(label, 0) for label in labels]


def _tod_bucket(ts: Any) -> str:
    hour: Optional[int] = None
    if isinstance(ts, (int, float)):
        try:
            hour = datetime.datetime.fromtimestamp(ts).hour
        except (ValueError, OSError) as e:
            log.warning("Failed to parse timestamp %r: %s", ts, e)
    elif isinstance(ts, str):
        for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                hour = datetime.datetime.strptime(ts[:19], fmt).hour
                break
            except ValueError:
                continue
    if hour is None:
        return "unknown"
    if 6 <= hour < 12:
        return "morning"
    if 12 <= hour < 18:
        return "afternoon"
    if 18 <= hour < 22:
        return "evening"
    return "night"


def _entropy_bucket(entropy: float) -> str:
    if entropy is None:
        return "none"
    if entropy < 0.5:
        return "low"
    if entropy < 1.5:
        return "medium"
    return "high"


def _route_type(rec: Dict) -> str:
    return rec.get("route_type", rec.get("route", "unknown"))


def _biome(rec: Dict) -> str:
    return rec.get("biome", rec.get("environment", "unknown"))


def _cohort_key(rec: Dict) -> str:
    return rec.get("cohort", rec.get("id", "unknown")[:8])


def _render_bar(
    labels: List[str], values: List[int], title: str, color: str, height: int = 20
) -> Image.Image:
    if not HAS_PIL:
        raise ImportError("PIL required for rendering")
    width = max(800, len(labels) * 60)
    img = Image.new("RGB", (width, height + 40), "white")
    draw = ImageDraw.Draw(img)
    max_val = max(values) if values else 1
    bar_width = (width - 40) // max(len(labels), 1)
    for i, (lab, val) in enumerate(zip(labels, values)):
        bar_h = int((val / max_val) * height)
        x0, y0 = 20 + i * bar_width, height - bar_h
        x1, y1 = 20 + (i + 1) * bar_width - 2, height
        draw.rectangle([x0, y0, x1, y1], fill=color)
        if len(lab) < 12:
            draw.text((x0 + 2, height + 2), lab, fill="black", font=ImageFont.load_default())
    draw.text((10, 10), title, fill="black", font=ImageFont.load_default())
    return img


def _render_table(headers: List[str], rows: List[List[str]], title: str) -> Image.Image:
    if not HAS_PIL:
        raise ImportError("PIL required for rendering")
    cell_w, cell_h = 120, 24
    width = cell_w * (len(headers) + 1)
    height = cell_h * (len(rows) + 2)
    img = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(img)
    for i, h in enumerate(["Cohort"] + headers):
        x, y = i * cell_w, 0
        draw.rectangle([x, y, x + cell_w, cell_h], outline="black", fill="#DDDDDD")
        draw.text((x + 4, 4), h, fill="black", font=ImageFont.load_default())
    for r_idx, row in enumerate(rows):
        y = (r_idx + 1) * cell_h
        for c_idx, cell in enumerate([f"#{r_idx + 1}"] + row):
            x = c_idx * cell_w
            draw.rectangle([x, y, x + cell_w, y + cell_h], outline="black")
            txt = str(cell)[:18]
            draw.text((x + 4, y + 4), txt, fill="black", font=ImageFont.load_default())
    draw.text((4, 4), title, fill="black", font=ImageFont.load_default())
    return img


def dashboard(records: List[Dict], output: Path, cohort_field: str = "cohort") -> None:
    """Generate dashboard for each cohort."""
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl required for Excel output")
    cohorts = _extract_cohorts(records, cohort_field)
    summary: List[Dict[str, Any]] = []
    for cohort, recs in cohorts.items():
        route_vals = [_route_type(r) for r in recs]
        biome_vals = [_biome(r) for r in recs]
        tod_vals = [_tod_bucket(r.get("timestamp")) for r in recs]
        ent_vals = [_entropy_bucket(r.get("action_entropy")) for r in recs]
        route_lbl, route_cnt = _histogram(route_vals)
        biome_lbl, biome_cnt = _histogram(biome_vals)
        tod_lbl, tod_cnt = _histogram(tod_vals)
        ent_lbl, ent_cnt = _histogram(ent_vals)
        summary.append({
            "cohort": cohort,
            "n": len(recs),
            "routes": route_lbl,
            "biomes": biome_lbl,
            "times": tod_lbl,
            "entropies": ent_lbl,
        })
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Summary"
        ws.append(["Cohort", "N", "Route Types", "Biomes", "Times of Day", "Entropies"])
        for _i, (lbl, cnt) in enumerate(zip(route_lbl, route_cnt)):
            ws.append([f"Route: {lbl}", "", cnt, "", "", ""])
        for _i, (lbl, cnt) in enumerate(zip(biome_lbl, biome_cnt)):
            ws.append([f"Biome: {lbl}", "", "", cnt, "", ""])
        for _i, (lbl, cnt) in enumerate(zip(tod_lbl, tod_cnt)):
            ws.append([f"Time: {lbl}", "", "", "", cnt, ""])
        for _i, (lbl, cnt) in enumerate(zip(ent_lbl, ent_cnt)):
            ws.append([f"Entropy: {lbl}", "", "", "", "", cnt])
        wb.save(output / f"{cohort}_dashboard.xlsx")
    wb_summary = openpyxl.Workbook()
    ws_s = wb_summary.active
    ws_s.title = "Cohorts"
    ws_s.append(["Cohort", "N", "Route Types", "Biomes", "Times of Day", "Entropies"])
    for s in summary:
        ws_s.append([
            s["cohort"],
            s["n"],
            ", ".join(s["routes"]),
            ", ".join(s["biomes"]),
            ", ".join(s["times"]),
            ", ".join(s["entropies"]),
        ])
    wb_summary.save(output / "cohort_summary.xlsx")


def main(argv: List[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Generate per-cohort diversity dashboards.")
    parser.add_argument("input", type=Path, help="Input file (JSON/YAML/CSV)")
    parser.add_argument(
        "-o", "--output", type=Path, default=Path("dashboard_out"), help="Output directory"
    )
    parser.add_argument(
        "--cohort-field",
        default="cohort",
        help="Field to group by (default: cohort)",
    )
    args = parser.parse_args(argv)
    args.output.mkdir(parents=True, exist_ok=True)
    records = load_data(args.input)
    dashboard(records, args.output, args.cohort_field)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
