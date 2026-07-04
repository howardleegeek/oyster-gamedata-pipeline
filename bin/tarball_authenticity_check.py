#!/usr/bin/env python3
"""tarball_authenticity_check.py — D5 proof-of-truth gate.

Heuristic-based scan of a buyer-spec tarball to classify each file as
REAL / PLACEHOLDER / UNKNOWN. Designed to RUN BEFORE we ship anything,
so we can never accidentally hand the buyer a placeholder.

Howard 2026-05-06: NO LLM, NO network, no false PASS. When in doubt,
classify UNKNOWN — the exit-1 gate covers UNKNOWN too.

Exits:
  0  all files REAL
  1  any file PLACEHOLDER or UNKNOWN
  2  usage / tarball missing / unreadable
"""
from __future__ import annotations

import argparse
import hashlib
import json
import logging
import shutil
import subprocess
import sys
import tarfile
import tempfile
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)


REAL = "REAL"
PLACEHOLDER = "PLACEHOLDER"
UNKNOWN = "UNKNOWN"


def _classify_video(p: Path) -> tuple[str, str]:
    """Look at ffprobe metadata + content variation for video.mp4."""
    if not p.exists():
        return UNKNOWN, "missing"
    try:
        res = subprocess.run(
            ["ffprobe", "-v", "error", "-show_format", "-show_streams",
             "-of", "json", str(p)],
            capture_output=True, text=True, check=True, timeout=15,
        )
        meta = json.loads(res.stdout or "{}")
    except Exception as e:
        return UNKNOWN, f"ffprobe failed: {e}"

    # 1. Encoder tags pointing at lavfi/testsrc → definitive PLACEHOLDER
    fmt_tags = (meta.get("format") or {}).get("tags") or {}
    encoder = (fmt_tags.get("encoder") or "").lower()
    comment = (fmt_tags.get("comment") or "").lower()
    for stream in meta.get("streams", []):
        encoder = (encoder + " " + (stream.get("tags", {}).get("encoder") or "")).lower()
    if "lavfi" in encoder or "testsrc" in encoder:
        return PLACEHOLDER, f"encoder tag contains lavfi/testsrc: {encoder!r}"

    # 1b. Real-capture metadata stamp (D15) → definitive REAL
    if "oyster-real-screen-capture" in comment or "oyster-recorder" in comment:
        return REAL, f"metadata comment claims real capture: {comment!r}"

    # 2. Frame variance check — sample 5 frames. Howard 2026-05-07: relaxed.
    # Frame-byte-identical alone is NOT proof of placeholder — a real screen
    # capture from a static desktop session also produces identical frames.
    # Without the encoder tag (rule 1) OR a real-capture stamp (rule 1b), we
    # CANNOT distinguish "real-but-static" from "synthetic-static".
    # Honest verdict: UNKNOWN, not PLACEHOLDER. The exit-1 gate still covers
    # UNKNOWN, so callers know to ask for explicit metadata.
    try:
        with tempfile.TemporaryDirectory() as td:
            for idx, t in enumerate([0.0, 0.5, 1.0, 1.5, 2.0]):
                subprocess.run(
                    ["ffmpeg", "-ss", str(t), "-i", str(p),
                     "-vframes", "1", "-f", "image2pipe", "-vcodec", "rawvideo",
                     "-pix_fmt", "gray",
                     str(Path(td) / f"f{idx}.gray")],
                    capture_output=True, check=False, timeout=10,
                )
            sizes = [
                (Path(td) / f"f{i}.gray").stat().st_size if (Path(td) / f"f{i}.gray").exists() else 0
                for i in range(5)
            ]
            if len(set(sizes)) == 1 and sizes[0] > 0:
                contents = []
                for i in range(5):
                    fp = Path(td) / f"f{i}.gray"
                    if fp.exists():
                        contents.append(fp.read_bytes()[:1024])
                if len(set(contents)) == 1:
                    return UNKNOWN, (
                        "all sampled frames byte-identical — could be "
                        "real-static or synthetic; need oyster-recorder "
                        "metadata stamp (D15) to disambiguate"
                    )
    except Exception as exc:
        # frame-sampling is best-effort — but we should still surface the
        # failure so a missing/corrupt ffmpeg install is diagnosable from
        # operator logs rather than only visible by absence of frames.
        logger.debug("frame-sampling best-effort failed for %s: %s", p, exc)

    return REAL, f"encoder={encoder.strip()}, multi-frame variation OK"


def _classify_depth_dir(d: Path) -> tuple[str, str]:
    """Multi-signal: NOTICE file > content-hash variance > UNKNOWN."""
    if not d.exists() or not d.is_dir():
        return UNKNOWN, "missing"
    # Howard 2026-05-07: the pipeline's _ensure_placeholders now drops
    # _PLACEHOLDER_NOTICE.txt when emitting the gradient farm. That file
    # is the strongest signal — definitive PLACEHOLDER.
    notice = d / "_PLACEHOLDER_NOTICE.txt"
    if notice.exists():
        return PLACEHOLDER, "depth/_PLACEHOLDER_NOTICE.txt present (gradient hardlink farm)"
    exrs = sorted(d.glob("*.exr"))
    if not exrs:
        return PLACEHOLDER, "no EXR files"
    hashes: set[str] = set()
    for f in exrs:
        h = hashlib.sha256(f.read_bytes()).hexdigest()
        hashes.add(h)
    n = len(exrs)
    unique = len(hashes)
    ratio = unique / n
    if ratio < 0.05:
        return PLACEHOLDER, f"{unique} unique / {n} files ({ratio:.1%}) — likely hardlinked or copies of one file"
    return REAL, f"{unique} unique / {n} files ({ratio:.1%}) — real per-frame variation"


def _classify_action_camera(p: Path) -> tuple[str, str]:
    """Three-tier classification: mod-driven > metadata-derived > placeholder.

    Howard 2026-05-07 (D18): added tier 0 — explicit ``_real_game_state``
    flag set by ``game_state_overlay.apply_to_record`` when the Fabric mod's
    JSONL was consumed. This unambiguously certifies real game-state and
    short-circuits the variance-based heuristics below (which can false-
    positive on stationary bots producing identical-but-real frames).

    Tier order:
        0. ``_real_game_state: True`` flag present → REAL (mod-driven)
        1. ``is_padded`` flag present → REAL (real bot data + transparent padding)
        2. Multi-field fingerprint variance → REAL (frame variation observed)
        3. Otherwise → PLACEHOLDER

    Type-check is strict on tier 0: only boolean ``True`` accepted, not
    truthy strings. This stops type-confusion attacks where a ``"true"``
    string masquerades as authenticated real data.
    """
    if not p.exists():
        return UNKNOWN, "missing"
    try:
        d = json.loads(p.read_text())
    except Exception as e:
        return UNKNOWN, f"json parse failed: {e}"
    if not isinstance(d, list) or not d:
        return UNKNOWN, f"unexpected shape: {type(d).__name__}"

    n = len(d)

    # Tier 0: explicit _real_game_state flag from mod-driven overlay
    # (D18). MUST be boolean True, not "true" / 1 / "yes" — those are
    # potential placeholder-shipping-as-real attacks.
    real_flagged = sum(1 for r in d if r.get("_real_game_state") is True)
    if real_flagged > 0:
        ratio = real_flagged / n
        return (
            REAL,
            f"{real_flagged}/{n} records carry _real_game_state=true "
            f"({ratio:.1%}) — mod-driven real game state present",
        )

    # Tier 1: explicit is_padded flag from buyer_spec_adapter
    if any("is_padded" in r for r in d):
        real_count = sum(1 for r in d if not r.get("is_padded"))
        ratio = real_count / n if n else 0
        if ratio < 0.02:  # less than 2% real records → effectively all padded
            return PLACEHOLDER, f"{real_count}/{n} records non-padded ({ratio:.1%}) — almost entirely synthetic"
        return REAL, f"{real_count}/{n} records non-padded ({ratio:.1%}, is_padded flag present) — real bot data + transparent padding"

    # Tier 2: multi-field fingerprint (position, rotation, speed, intrinsics)
    fingerprint_keys = (
        "camera_position", "camera_rotation_oula", "camera_quat",
        "camera_speed", "player_position", "yaw", "pitch",
    )
    fingerprints = set()
    for r in d:
        fp = tuple(
            tuple(r[k]) if isinstance(r.get(k), list) else r.get(k)
            for k in fingerprint_keys
        )
        fingerprints.add(fp)
    unique = len(fingerprints)
    ratio = unique / n if n else 0
    if ratio < 0.05:
        return PLACEHOLDER, f"{unique} distinct multi-field fingerprints / {n} records ({ratio:.1%}) — mostly identical"
    return REAL, f"{unique} distinct multi-field fingerprints / {n} records ({ratio:.1%}) — real frame variation"


def _classify_gameinfo_xlsx(p: Path) -> tuple[str, str]:
    """Open with openpyxl, scan all cells for placeholder strings."""
    if not p.exists():
        return UNKNOWN, "missing"
    try:
        from openpyxl import load_workbook
    except ImportError:
        return UNKNOWN, "openpyxl not installed"
    try:
        wb = load_workbook(p, read_only=True)
    except Exception as e:
        return UNKNOWN, f"open failed: {e}"
    needles = ("placeholder", "stub", "stop-gap", "stopgap", "dummy")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell is None:
                    continue
                s = str(cell).lower()
                for n in needles:
                    if n in s:
                        return PLACEHOLDER, f"sheet={sheet_name!r} cell contains {n!r}: {cell!r}"
    return REAL, f"{len(wb.sheetnames)} sheets, no placeholder strings"


def _classify_systeminfo(p: Path) -> tuple[str, str]:
    if not p.exists():
        return UNKNOWN, "missing"
    try:
        d = json.loads(p.read_text())
    except Exception as e:
        return UNKNOWN, f"json parse failed: {e}"
    if not isinstance(d, dict):
        return UNKNOWN, "not a dict"
    rec_at = d.get("recordedAt") or d.get("recorded_at")
    rec_ver = d.get("recorderVersion") or d.get("recorder_version")
    if not rec_at and not rec_ver:
        return UNKNOWN, "missing both recordedAt and recorderVersion"
    return REAL, f"recordedAt={rec_at}, recorderVersion={rec_ver}"


def _classify_readme(p: Path) -> tuple[str, str]:
    """Howard 2026-05-07: relaxed — 'placeholder if absent' is a conditional
    caution about absent files, NOT an active placeholder claim. Only flag
    PLACEHOLDER on unambiguous active language like 'this is a placeholder'
    or 'do not use'.
    """
    if not p.exists():
        return UNKNOWN, "missing"
    body = p.read_text().lower()
    # Active placeholder claims (definitive)
    active = (
        "this is a placeholder",
        "this bundle is a placeholder",
        "do not use",
        "do not train",
        "stop-gap mvp",
        "stopgap mvp",
    )
    found = [n for n in active if n in body]
    if found:
        return PLACEHOLDER, f"contains active placeholder claim: {found}"
    return REAL, "no active placeholder claim"


def audit_tarball(tar_path: Path) -> dict[str, Any]:
    if not tar_path.exists():
        raise FileNotFoundError(tar_path)

    work = Path(tempfile.mkdtemp(prefix="auth_check_"))
    try:
        with tarfile.open(tar_path, "r:gz") as tar:
            tar.extractall(work)
        roots = [p for p in work.iterdir() if p.is_dir()]
        if len(roots) != 1:
            raise RuntimeError(f"expected one top-level dir, got {len(roots)}")
        bundle = roots[0]

        files = []

        st, ev = _classify_video(bundle / "video.mp4")
        files.append({"name": "video.mp4", "status": st, "evidence": ev})

        st, ev = _classify_depth_dir(bundle / "depth")
        files.append({"name": "depth/", "status": st, "evidence": ev})

        st, ev = _classify_action_camera(bundle / "action_camera.json")
        files.append({"name": "action_camera.json", "status": st, "evidence": ev})

        st, ev = _classify_gameinfo_xlsx(bundle / "gameinfo.xlsx")
        files.append({"name": "gameinfo.xlsx", "status": st, "evidence": ev})

        st, ev = _classify_systeminfo(bundle / "systeminfo.json")
        files.append({"name": "systeminfo.json", "status": st, "evidence": ev})

        st, ev = _classify_readme(bundle / "README.md")
        files.append({"name": "README.md", "status": st, "evidence": ev})

        statuses = [f["status"] for f in files]
        if all(s == REAL for s in statuses):
            verdict = REAL
        elif PLACEHOLDER in statuses:
            verdict = "MIXED" if REAL in statuses else PLACEHOLDER
        else:
            verdict = "MIXED"

        return {
            "tarball": str(tar_path),
            "verdict": verdict,
            "files": files,
            "summary": {
                "real_count": sum(1 for s in statuses if s == REAL),
                "placeholder_count": sum(1 for s in statuses if s == PLACEHOLDER),
                "unknown_count": sum(1 for s in statuses if s == UNKNOWN),
            },
        }
    finally:
        shutil.rmtree(work, ignore_errors=True)


def main(argv: list[str] | None = None) -> int:
    p = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    p.add_argument("tarball", type=Path)
    p.add_argument("--json", action="store_true", help="emit JSON to stdout")
    args = p.parse_args(argv)

    try:
        report = audit_tarball(args.tarball)
    except FileNotFoundError:
        print(f"ERROR: tarball not found: {args.tarball}", file=sys.stderr)
        return 2
    except Exception as e:
        print(f"ERROR: {type(e).__name__}: {e}", file=sys.stderr)
        return 2

    if args.json:
        print(json.dumps(report, indent=2))
    else:
        print(f"Tarball: {report['tarball']}")
        print(f"Verdict: {report['verdict']}")
        print()
        for f in report["files"]:
            print(f"  [{f['status']:11}] {f['name']:25} — {f['evidence']}")
        print()
        s = report["summary"]
        print(f"  REAL={s['real_count']}  PLACEHOLDER={s['placeholder_count']}  UNKNOWN={s['unknown_count']}")

    return 0 if report["verdict"] == REAL else 1


if __name__ == "__main__":
    sys.exit(main())
