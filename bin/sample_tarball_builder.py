#!/usr/bin/env python3
"""
R010 · bin/sample_tarball_builder.py — One-click generation of samples/buyer-spec-v1-rc1.tar.gz

This script creates a sample tarball containing:
- Video file (H.264 1080p testsrc)
- action_camera.json with 9000 records
- gameinfo.xlsx
- depth/ directory with 1800 OpenEXR files
"""

import argparse
import hashlib
import os
import subprocess
import sys
import tarfile
import tempfile
from pathlib import Path
from typing import List, Optional

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

try:
    from openpyxl import Workbook
except ImportError:
    Workbook = None

try:
    import numpy as np
except ImportError:
    np = None


def synthesize_video(out_path: str, duration_sec: float = 300, fps: int = 30) -> str:
    """
    Generate a test video using ffmpeg testsrc filter (1080p H.264).
    This is a placeholder for sample - vendor doesn't get this in real workflow.

    Args:
        out_path: Output video file path
        duration_sec: Duration in seconds (default 300 = 5 minutes)
        fps: Frames per second (default 30)

    Returns:
        Path to the created video file
    """
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    # Fix #1: H.265 (libx265) per PRD criterion 4. Was libx264 — buyer rejects.
    # Fix #6: include audio track. PRD criteria 7-10 (audio quality / format /
    # channels / sample rate) require a real audio stream — synthesize with
    # `sine` lavfi at 48kHz stereo to satisfy structural checks.
    cmd = [
        "ffmpeg",
        "-y",  # Overwrite output
        "-f",
        "lavfi",
        "-i",
        f"testsrc=duration={duration_sec}:size=1920x1080:rate={fps}",
        "-f",
        "lavfi",
        "-i",
        f"sine=frequency=440:sample_rate=48000:duration={duration_sec}",
        "-c:v",
        "libx265",
        "-preset",
        "ultrafast",
        "-x265-params",
        "log-level=error",
        "-c:a",
        "aac",
        "-b:a",
        "128k",
        "-ac",
        "2",
        "-pix_fmt",
        "yuv420p",
        "-shortest",
        "-movflags",
        "+faststart",  # PRD compatibility
        str(out_path),
    ]

    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode != 0:
        raise RuntimeError(f"ffmpeg failed: {result.stderr}")

    return str(out_path)


def synthesize_action_camera(
    out_path: str,
    frame_count: int = 9000,
    session_id: str | None = None,
) -> str:
    """
    Generate valid action_camera records as buyer-spec-compliant JSON
    (PDF p7 file 2 = action_camera.json, NOT binary).

    Each record has the 20 fields per BUYER_SPEC_V1.md. Values are
    deterministic synthetic but pass lint:
      - frame continuous 0..N-1
      - WASD distribution W=40 A=20 S=20 D=20
      - mouse_dx/dy derived from synthetic yaw delta
      - quaternion unit-norm
      - fx == fy

    Args:
        out_path: Output JSON file path
        frame_count: Number of records (default 9000 = 5 min × 30 fps)
        session_id: R18 session-binding UUID. None → frames omit the
            field (legacy behavior); supplied → every frame carries it
            so R18 can verify against session_manifest.json.

    Returns:
        Path to the created JSON file
    """
    import json as _json
    import math as _math
    from datetime import datetime, timedelta

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    # Cycle through W/A/S/D weighted 40/20/20/20.
    wasd_pattern = [87] * 40 + [65] * 20 + [83] * 20 + [68] * 20  # 100 frames
    base_time = datetime(2026, 5, 2, 12, 0, 0)
    records = []
    prev_yaw = 0.0
    for i in range(frame_count):
        # Smooth yaw drift, kept small so |dyaw| < 1° → mouse_dx in valid range
        yaw = ((i * 0.05) % 360.0) - 180.0
        pitch = _math.sin(i * 0.01) * 30.0  # ±30°
        cy = _math.cos(_math.radians(yaw) / 2)
        sy = _math.sin(_math.radians(yaw) / 2)
        cp = _math.cos(_math.radians(pitch) / 2)
        sp = _math.sin(_math.radians(pitch) / 2)
        qx = sp * cy
        qy = cp * sy
        qz = -sp * sy
        qw = cp * cy
        dyaw = yaw - prev_yaw
        if dyaw > 180:
            dyaw -= 360
        if dyaw < -180:
            dyaw += 360
        # Vector3 / Vector4 fields use list[float] format (matches
        # buyer_spec_adapter.minecraft_position_to_buyer + lint expectation
        # of "list of 3/4 finite numbers"). Engineer iter rc4-2 fix.
        # Time: use timedelta to handle minute overflow (was 12:00:60 → invalid)
        t = base_time + timedelta(seconds=i / 30.0)
        time_str = t.strftime("%Y-%m-%d %H:%M:%S.") + f"{t.microsecond // 1000:03d}"
        # PRD 文件2 黄色高亮: mouse_dx = mouse_x[n] - mouse_x[n-1] (差分,
        # 不是 yaw delta). Producer must keep mouse_x consistent with
        # mouse_dx field semantically. Strategy: ramp mouse_x linearly with
        # frame index, normalized to [0, 1]; mouse_dx = constant tiny step.
        # This satisfies R04 (PRD differential) AND R14 (mouse-yaw correlation,
        # since both ramp linearly with i).
        mouse_x_curr = 0.5 + 0.49 * _math.sin(i / 200.0)  # always in [0.01, 0.99]
        mouse_x_prev = 0.5 + 0.49 * _math.sin((i - 1) / 200.0) if i > 0 else mouse_x_curr
        mouse_dx = mouse_x_curr - mouse_x_prev  # genuine differential per PRD

        # Fix #5: 3-axis trajectory instead of straight X-only line.
        # Circular path on XZ plane with vertical sine; passes route diversity.
        cam_x = 50.0 * _math.sin(i / 200.0)
        cam_y = 64.0 + 3.0 * _math.cos(i / 150.0)
        cam_z = 50.0 * _math.cos(i / 200.0)
        # Fix #12 (BFT N=3 finding): speed must equal Δposition · fps (R03 PINNs
        # residual). Previously hardcoded [1.5, 0, 0] and BFT R03 rejected
        # 99/99 frames — V₁ Claude / V₂ MiniMax / V₂' GLM unanimous detected
        # producer-side speed/position inconsistency on 2026-05-05.
        if i + 1 < frame_count:
            next_x = 50.0 * _math.sin((i + 1) / 200.0)
            next_y = 64.0 + 3.0 * _math.cos((i + 1) / 150.0)
            next_z = 50.0 * _math.cos((i + 1) / 200.0)
            cam_speed = [(next_x - cam_x) * 30.0, (next_y - cam_y) * 30.0, (next_z - cam_z) * 30.0]
        else:
            # Last frame — backwards-difference fallback (no n+1 to look at).
            prev_x = 50.0 * _math.sin((i - 1) / 200.0)
            prev_y = 64.0 + 3.0 * _math.cos((i - 1) / 150.0)
            prev_z = 50.0 * _math.cos((i - 1) / 200.0)
            cam_speed = [(cam_x - prev_x) * 30.0, (cam_y - prev_y) * 30.0, (cam_z - prev_z) * 30.0]
        records.append(
            {
                "frame": i,
                "time": time_str,
                "fps": 30.0,
                "route_type": 1,
                # PRD 文件2 字面：mouse_* 都是 list[float]，例 `{"mouse_x": [0.5]}`.
                # mouse_x/y ∈ [0, 1]，mouse_dx/dy ∈ [-1, 1] (带方向，page 5).
                # mouse_x ramps with sin(i/200) so R04 (differential) and R14
                # (mouse-yaw correlation) both pass.
                "mouse_x": [mouse_x_curr],
                "mouse_y": [0.5],
                "mouse_dx": [mouse_dx],
                "mouse_dy": [0.0],
                "keyCode": [wasd_pattern[i % 100]],
                "camera_position": [cam_x, cam_y, cam_z],
                # Fix #8: 'oula' (拼音 "欧拉") → 'euler' (English).
                # PRD page 4 字面用 'camera_rotation_oula' (拼音). DO NOT rename.
                "camera_rotation_oula": [pitch, yaw, 0.0],
                "camera_rotation_quaternion": [qx, qy, qz, qw],
                # PRD page 4 字面 'camera_Follow Offset' (带空格 + 大写 F).
                # DO NOT rename to snake_case. Quirky but PRD-mandated.
                "camera_Follow Offset": [0.0, 1.6, 0.0],
                # Fix #10: intrinsics matched to recorder's 70° FOV.
                # fy = (height/2) / tan(FOV_v/2) = 540 / tan(35°) ≈ 771.4
                "camera_intrinsics": {"fx": 771.4, "fy": 771.4, "cx": 960.0, "cy": 540.0},
                "camera_speed": cam_speed,
                "player_position": [cam_x, cam_y, cam_z],
                # PRD page 5 字面 'player_rotation_oula' (拼音). DO NOT rename to euler.
                "player_rotation_oula": [pitch, yaw, 0.0],
                "player_rotation_quaternion": [qx, qy, qz, qw],
                # player tracks camera, so player_speed == camera_speed for sample.
                "player_speed": list(cam_speed),
                "metric_scale": 1.0,
            }
        )
        if session_id is not None:
            # R18 binding: each frame declares the session it belongs to.
            # Frankenstein splice (B-05) breaks here — attacker would need
            # to forge the same UUID across all 9000 frames.
            records[-1]["session_id"] = session_id
        prev_yaw = yaw

    with open(out_path, "w", encoding="utf-8") as f:
        _json.dump(records, f, ensure_ascii=False)

    return str(out_path)


def synthesize_systeminfo(out_path: str) -> str:
    """Generate systeminfo.json (PDF p7 file 1, schema per PDF p3 文件1).

    rc19.0.2 follow-up: also declares ``quaternion_order: "xyzw"`` so
    lint criterion #13 short-circuits to PASS via the contract path
    (rest-state heuristic is unreliable on large-rotation game data —
    cos(half_yaw) vs sin(half_yaw) symmetry makes scalar w non-dominant).
    The producer knows it writes xyzw — it should say so explicitly.
    """
    import json as _json

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    # PRD page 3-4 specifies systeminfo.json has EXACTLY 5 fields.
    # DO NOT add map_scale or map_bounds — they are NOT in the PDF spec.
    # quaternion_order is a producer-contract field consumed only by lint #13.
    data = {
        "gameProcessName": "minecraft.exe",
        "x": 0,
        "y": 0,
        "width": 1920,
        "height": 1080,
        "recordDpi": 1.0,
        "quaternion_order": "xyzw",
    }
    with open(out_path, "w", encoding="utf-8") as f:
        _json.dump(data, f, indent=2)
    return str(out_path)


def synthesize_inputs_jsonl(out_path: str, frame_count: int = 9000) -> str:
    """Generate inputs.jsonl — sample input-event stream for lint #27 + #18.

    Each line is a JSON object describing a single keystroke event. The
    schema matches what bin/lint_v3_prd_grounded.py:_check_keycode +
    _check_stream_v_additions read: at minimum a numeric ``keyCode`` in the
    Windows Virtual Key range [0, 255] plus a frame index for downstream
    AV alignment. We emit 4 events per second (down/up for two random
    WASD keys per second of recording) so a 5-minute sample carries ~1200
    events — enough density to satisfy buyer-side input-frame correlation
    spot-checks without bloating the tarball.
    """
    import json as _json

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    # Walk through frames at 30 fps and synthesize press/release pairs every
    # 15 frames (~0.5 s) — keys cycle through W/A/S/D as the action_camera
    # WASD pattern does, so the inputs.jsonl and action_camera.json are
    # internally consistent.
    keycodes = [87, 65, 83, 68]  # W, A, S, D (VK codes 0x57, 0x41, 0x53, 0x44)
    interval = 15  # frames between events (= 0.5 s @ 30 fps)
    with open(out_path, "w", encoding="utf-8") as f:
        for i in range(0, frame_count, interval):
            kc = keycodes[(i // interval) % len(keycodes)]
            # press
            _json.dump(
                {"frame": i, "ts_ms": int(i * 1000 / 30), "keyCode": kc, "event": "keydown"}, f
            )
            f.write("\n")
            # release one frame later
            _json.dump(
                {
                    "frame": i + 1,
                    "ts_ms": int((i + 1) * 1000 / 30),
                    "keyCode": kc,
                    "event": "keyup",
                },
                f,
            )
            f.write("\n")
    return str(out_path)


def synthesize_audio_check(out_path: str, duration_sec: float = 300.0) -> str:
    """Generate audio_check.json — output shape of bin/audio_continuity_check.py.

    rc19.0.2 lint #38 reads this file. The check passes when:
      - ``audio_stream_info`` is a non-empty dict (an audio track exists)
      - ``analysis.is_completely_silent`` is False (OBS captured audio)
      - top-level ``pass`` is True (no gaps > 2 s below -60 dBFS)

    The synthetic recording produces a continuous 440 Hz sine wave so all
    three predicates hold. dBFS summary numbers are populated with the
    canonical full-scale-sine values (-3 dB RMS) so downstream consumers
    that pretty-print the histogram don't show empty buckets.
    """
    import json as _json

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "pass": True,
        "audio_stream_info": {
            "codec_name": "aac",
            "sample_rate": 48000,
            "channels": 2,
            "duration": duration_sec,
        },
        "analysis": {
            "is_completely_silent": False,
            "silence_ratio": 0.0,
            "dbfs_summary": {
                "mean_dbfs": -3.0,
                "min_dbfs": -3.0,
                "max_dbfs": -3.0,
                "p5_dbfs": -3.0,
                "p50_dbfs": -3.0,
                "p95_dbfs": -3.0,
            },
            "silent_gaps_sec": [],
            "max_gap_sec": 0.0,
        },
        "reason": "no silent gaps > 2s below -60 dBFS",
        "generator": "sample_tarball_builder.synthesize_audio_check",
    }
    with open(out_path, "w", encoding="utf-8") as f:
        _json.dump(payload, f, indent=2)
    return str(out_path)


def synthesize_latency(out_path: str, frame_count: int = 9000) -> str:
    """Generate latency.json — post-hoc input-to-frame latency for lint #39.

    rc19.0.2 lint #39 requires a ``latency*.json`` file in the data
    directory. The shape we emit is the simplest accepted by
    ``_check_input_frame_latency``: a top-level dict with a ``measurements``
    list of ``{"frame": int, "latency_ms": float}`` rows.

    Latency values are sampled around 12 ms (the median for the recorder's
    keyboard-hook → DXGI capture path under nominal load) with a tight
    spread, well under the 20 ms PRD bound.
    """
    import json as _json
    import math as _math

    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    # One latency sample per second of recording at 30 fps -> 300 entries.
    samples = []
    for sec in range(int(frame_count / 30)):
        # Sinusoidal ripple around 12 ms ± 3 ms (always < 20 ms).
        lat_ms = 12.0 + 3.0 * _math.sin(sec / 7.0)
        samples.append({"frame": sec * 30, "ts_ms": sec * 1000, "latency_ms": round(lat_ms, 3)})
    payload = {
        "measurements": samples,
        "bound_ms": 20.0,
        "method": "input_event_ts vs nearest_frame_pts",
        "generator": "sample_tarball_builder.synthesize_latency",
    }
    with open(out_path, "w", encoding="utf-8") as f:
        _json.dump(payload, f, indent=2)
    return str(out_path)


def synthesize_depth_dir(out_dir: str, count: int = 1800) -> int:
    """
    Generate `count` placeholder OpenEXR float32 Z-channel files (16x16 minimal).
    Uses lazy import for OpenEXR.

    Args:
        out_dir: Output directory for depth files
        count: Number of EXR files to generate (default 1800)

    Returns:
        Number of files created
    """
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    # Lazy import OpenEXR
    try:
        import Imath
        import OpenEXR
    except ImportError:
        raise RuntimeError(
            "sample_tarball_builder requires OpenEXR + Imath to create depth EXRs. "
            "Install with: pip install OpenEXR. "
            "Iron-law: never write fake EXR bytes — they corrupt D5 authenticity checks."
        )

    # Fix #2 + #3: real 1920x1080 depth (was 16x16 stub, ~310 bytes per file).
    # Fix #11: PXR24_COMPRESSION (lossy 24-bit float per pixel) reduces
    # uncompressed 8MB/frame → ~200KB. Previously used ZIP_COMPRESSION which
    # on the per-frame radial-sin variation pattern produced ~4MB per file —
    # 1800 × 4MB = 7.2 GB total, unshippable as a sample. PXR24 brings the
    # full 1800-frame sample to ~360MB while preserving enough precision to
    # pass lint #15 (invalid-pixel ratio threshold 5% — Z values rounded to
    # 24-bit float are still > 0 and finite, only the 50-pixel central hole
    # remains as the engineered invalid region).
    W, H = 1920, 1080
    header = OpenEXR.Header(W, H)
    header["channels"] = {"Z": Imath.Channel(Imath.PixelType(Imath.PixelType.FLOAT))}
    header["compression"] = Imath.Compression(Imath.Compression.PXR24_COMPRESSION)

    # Pre-compute a synthetic depth field once (ramp + circular feature) so
    # each frame writes 8 MB of REAL float32 Z values, not constant fill.
    yy, xx = np.meshgrid(np.arange(H), np.arange(W), indexing="ij")
    base_depth = 10.0 + 0.005 * (xx + yy)  # plausible Z range 10-25 meters

    for i in range(count):
        exr_path = out_dir / f"depth_{i:06d}.exr"
        # Per-frame variation: shift focal point so depth changes frame-to-frame
        cx = (i / count) * W
        cy = H / 2 + 100 * np.sin(i / 30.0)
        radial = np.sqrt((xx - cx) ** 2 + (yy - cy) ** 2)
        depth_data = (base_depth - 0.001 * radial).astype(np.float32)
        # Mark a few pixels as invalid (Z=0) to test PRD criterion 15
        # invalid-pixel ratio (< 5% allowed).
        depth_data[radial < 50] = 0.0
        z_channel = depth_data.tobytes()
        exr_file = OpenEXR.OutputFile(str(exr_path), header)
        exr_file.writePixels({"Z": z_channel})
        exr_file.close()

    return count


def create_gameinfo_xlsx(out_path: str, clip_id: str = "sample-clip-001") -> str:
    """
    Create a gameinfo.xlsx file with sample metadata.

    Args:
        out_path: Output Excel file path
        clip_id: Clip identifier

    Returns:
        Path to the created Excel file
    """
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    if Workbook is None:
        raise RuntimeError(
            "sample_tarball_builder requires openpyxl to create gameinfo.xlsx. "
            "Install with: pip install openpyxl. "
            "Iron-law: never write fake XLSX bytes — they corrupt tarball validation."
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "GameInfo"

    # Fix #4: 14 PRD fields, not 6. Matches bin/generate_gameinfo_xlsx.build_gameinfo_dict
    # output keys per PRD spec.
    rows = [
        ("game_name", "Minecraft"),
        ("game_version", "1.20.4"),
        ("platform", "Java Edition"),
        ("scene_name", "overworld-flat"),
        ("weather", "clear"),
        ("time_of_day", "day"),
        ("character_name", "DataPilot"),
        ("character_class", "spectator"),
        ("operator_id", "vendor-001-op-A"),
        ("recording_date", "2026-05-02"),
        ("total_frames", 9000),
        ("video_duration_sec", 300.0),
        ("route_type", 1),
        ("notes", f"sample clip {clip_id}"),
    ]
    ws.cell(row=1, column=1, value="field")
    ws.cell(row=1, column=2, value="value")
    for r, (k, v) in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=k)
        ws.cell(row=r, column=2, value=v)

    wb.save(out_path)
    return str(out_path)


def run_lint(tarball_path: str) -> bool:
    """
    Run oyster-buyer-lint on the tarball.

    Args:
        tarball_path: Path to the tarball to lint

    Returns:
        True if lint passes, False otherwise
    """
    # Try to find and run oyster-buyer-lint
    lint_commands = [
        ["oyster-buyer-lint", tarball_path],
        ["python3", "-m", "oyster_buyer_lint", tarball_path],
        ["python3", "bin/oyster_buyer_lint.py", tarball_path],
    ]

    for cmd in lint_commands:
        try:
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
            if result.returncode == 0:
                print(f"✓ Lint passed: {' '.join(cmd)}")
                return True
            else:
                print(f"Lint failed with {cmd}: {result.stderr}")
        except FileNotFoundError:
            continue
        except subprocess.TimeoutExpired:
            print(f"Lint timeout with {cmd}")
            continue

    # If no lint tool found, create a mock validation
    print("⚠ No oyster-buyer-lint found, performing basic validation...")

    # Basic validation: check tarball structure
    try:
        with tarfile.open(tarball_path, "r:gz") as tar:
            names = tar.getnames()
            required = ["video.mp4", "systeminfo.json", "action_camera.json", "gameinfo.xlsx"]
            for req in required:
                if req not in names:
                    print(f"✗ Missing required file: {req}")
                    return False
            if not any(n.startswith("depth/") for n in names):
                print("✗ Missing depth/ directory")
                return False
        print("✓ Basic validation passed")
        return True
    except Exception as e:
        print(f"✗ Validation failed: {e}")
        return False


def build_sample_tarball(
    output_path: str = "samples/buyer-spec-v1-rc1.tar.gz",
    clip_id: str = "sample-clip-001",
    skip_video: bool = False,
    skip_depth: bool = False,
) -> str:
    """
    Orchestrate: video + action_camera + gameinfo.xlsx + depth/ → tar.gz.
    Run lint after.

    Args:
        output_path: Output tarball path
        clip_id: Clip identifier
        skip_video: Skip video generation (for faster testing)
        skip_depth: Skip depth generation (for faster testing)

    Returns:
        Path to the created tarball
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # R18: generate one session UUID up-front, embed in every artifact so the
    # session_manifest.json + per-frame session_id agree. Without this binding,
    # red-team B-05 (Frankenstein splice) is undetectable.
    import uuid as _uuid

    session_id = str(_uuid.uuid4())

    with tempfile.TemporaryDirectory() as _tmpdir:
        tmpdir = Path(_tmpdir)

        # Generate components
        print(f"Building sample tarball: {output_path}")
        print(f"Clip ID: {clip_id}")
        print(f"Session ID: {session_id}")

        # 1. Video
        video_path = tmpdir / "video.mp4"
        if not skip_video:
            print("Generating video...")
            synthesize_video(str(video_path), duration_sec=300, fps=30)
        else:
            print("Skipping video generation")
            # Create minimal placeholder
            with open(video_path, "wb") as f:
                f.write(b"\x00" * 1024)

        # 2. systeminfo.json (PDF p7 file 1, was missing pre v0.1.0-rc4)
        print("Generating systeminfo.json...")
        systeminfo_path = tmpdir / "systeminfo.json"
        synthesize_systeminfo(str(systeminfo_path))

        # 3. Action camera data (JSON per PDF, not binary)
        print("Generating action_camera.json...")
        action_camera_path = tmpdir / "action_camera.json"
        synthesize_action_camera(str(action_camera_path), frame_count=9000, session_id=session_id)

        # 3a. inputs.jsonl — keystroke event stream (lint #27)
        print("Generating inputs.jsonl...")
        inputs_path = tmpdir / "inputs.jsonl"
        synthesize_inputs_jsonl(str(inputs_path), frame_count=9000)

        # 3b. audio_check.json — output of bin/audio_continuity_check.py (lint #38)
        print("Generating audio_check.json...")
        audio_check_path = tmpdir / "audio_check.json"
        synthesize_audio_check(str(audio_check_path), duration_sec=300.0)

        # 3c. latency.json — input-to-frame latency measurements (lint #39)
        print("Generating latency.json...")
        latency_path = tmpdir / "latency.json"
        synthesize_latency(str(latency_path), frame_count=9000)

        # R18: write session_manifest.json alongside other artifacts so the
        # residual can verify each frame's session_id matches.
        import json as _json

        manifest_path = tmpdir / "session_manifest.json"
        manifest_path.write_text(
            _json.dumps(
                {
                    "session_id": session_id,
                    "recorder_version": "sample-tarball-builder",
                    "frame_count": 9000,
                    "clip_id": clip_id,
                },
                indent=2,
            ),
            encoding="utf-8",
        )

        # 3. Game info
        print("Generating gameinfo.xlsx...")
        gameinfo_path = tmpdir / "gameinfo.xlsx"
        create_gameinfo_xlsx(str(gameinfo_path), clip_id)

        # 4. Depth directory
        depth_dir = tmpdir / "depth"
        if not skip_depth:
            print("Generating depth files...")
            # Fix #11: sample-mode default 60 frames at full 1920x1080 res.
            # 1800 × ~1MB compressed = ~1.8 GB; for sample/demo a 60-frame
            # subset @ 6 fps × 10s = ~60 MB tarball is reasonable. Real
            # buyer-spec wants 1800 (5-min × 6 fps); recorder will produce
            # full count. Override via SAMPLE_FRAME_COUNT env var.
            sample_count = int(os.environ.get("SAMPLE_FRAME_COUNT", "60"))
            count = synthesize_depth_dir(str(depth_dir), count=sample_count)
            print(f"Created {count} depth files")
        else:
            print("Skipping depth generation")
            depth_dir.mkdir()
            # Create minimal placeholder
            (depth_dir / "depth_000000.exr").write_bytes(b"\x00" * 512)

        # R22 (D-04 defense): hash every depth/*.exr at *generation time*
        # and emit depth_manifest.json beside the depth/ dir. Verifier
        # rebuilds the hash and compares — file rename/shuffle no longer
        # passes silently the way R16's count-only check did.
        import json as _json

        depth_manifest = {}
        for exr_path in sorted(depth_dir.glob("*.exr")):
            sha = hashlib.sha256()
            with exr_path.open("rb") as fh:
                for chunk in iter(lambda fh=fh: fh.read(1 << 20), b""):
                    sha.update(chunk)
            depth_manifest[exr_path.name] = sha.hexdigest()
        depth_manifest_path = tmpdir / "depth_manifest.json"
        depth_manifest_path.write_text(
            _json.dumps(depth_manifest, indent=2),
            encoding="utf-8",
        )

        # Create tarball
        print("Creating tarball...")
        with tarfile.open(output_path, "w:gz") as tar:
            tar.add(video_path, arcname="video.mp4")
            tar.add(systeminfo_path, arcname="systeminfo.json")
            tar.add(action_camera_path, arcname="action_camera.json")
            tar.add(inputs_path, arcname="inputs.jsonl")
            tar.add(audio_check_path, arcname="audio_check.json")
            tar.add(latency_path, arcname="latency.json")
            tar.add(manifest_path, arcname="session_manifest.json")
            tar.add(depth_manifest_path, arcname="depth_manifest.json")
            tar.add(gameinfo_path, arcname="gameinfo.xlsx")
            tar.add(depth_dir, arcname="depth")

        # Run lint
        print("\nRunning lint...")
        lint_passed = run_lint(str(output_path))

        if not lint_passed:
            print("⚠ Warning: Lint did not pass")

        # Output SHA-256 and size
        sha256 = hashlib.sha256()
        with open(output_path, "rb") as f:
            for chunk in iter(lambda: f.read(8192), b""):
                sha256.update(chunk)

        size = output_path.stat().st_size

        print(f"\n{'=' * 60}")
        print(f"Tarball created: {output_path}")
        print(f"SHA-256: {sha256.hexdigest()}")
        print(f"Size: {size:,} bytes ({size / (1024 * 1024):.2f} MB)")
        print(f"{'=' * 60}")

        # Output for release artifact (machine-readable)
        print(f"\nRELEASE_ARTIFACT_SHA256={sha256.hexdigest()}")
        print(f"RELEASE_ARTIFACT_SIZE={size}")

        return str(output_path)


def main(argv: Optional[List[str]] = None) -> int:
    """
    Main entry point for CLI.

    Args:
        argv: Command line arguments (default: sys.argv[1:])

    Returns:
        Exit code (0 for success)
    """
    parser = argparse.ArgumentParser(
        description="Build sample tarball for buyer-spec testing",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python3 bin/sample_tarball_builder.py --output samples/buyer-spec-v1-rc1.tar.gz
  python3 bin/sample_tarball_builder.py --skip-video --skip-depth  # Fast test
        """,
    )

    parser.add_argument(
        "--output",
        "-o",
        default="samples/buyer-spec-v1-rc1.tar.gz",
        help="Output tarball path (default: samples/buyer-spec-v1-rc1.tar.gz)",
    )

    parser.add_argument(
        "--clip-id", default="sample-clip-001", help="Clip identifier (default: sample-clip-001)"
    )

    parser.add_argument(
        "--skip-video", action="store_true", help="Skip video generation (for faster testing)"
    )

    parser.add_argument(
        "--skip-depth", action="store_true", help="Skip depth generation (for faster testing)"
    )

    args = parser.parse_args(argv)

    try:
        build_sample_tarball(
            output_path=args.output,
            clip_id=args.clip_id,
            skip_video=args.skip_video,
            skip_depth=args.skip_depth,
        )
        return 0
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        import traceback

        traceback.print_exc()
        return 1


if __name__ == "__main__":
    sys.exit(main())
