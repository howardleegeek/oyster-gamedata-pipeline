#!/usr/bin/env python3
"""
R009 · bin/generate_gameinfo_xlsx.py
14-field gameinfo.xlsx generator

Generates an Excel file with game information metadata for video recordings.
"""

import argparse
import json
import logging
import os
import re
import sys
import zipfile
from datetime import date
from pathlib import Path
from typing import Any, Dict, Iterator, List, Optional

logger = logging.getLogger(__name__)


# Field names in order (PRD §3.3)
FIELD_NAMES = [
    "game_name",
    "game_version",
    "platform",
    "scene_name",
    "weather",
    "time_of_day",
    "character_name",
    "character_class",
    "operator_id",
    "recording_date",
    "total_frames",
    "video_duration_sec",
    "route_type",
    "notes",
]

MC_VERSION_IN_TITLE_RE = re.compile(
    r"\bMinecraft\s+(?P<version>\d+\.\d+(?:\.\d+)?(?:[-+._A-Za-z0-9]*)?)\b",
    re.IGNORECASE,
)
MC_VERSION_TOKEN_RE = re.compile(
    r"(?<![\d.])(?P<version>\d+\.\d+(?:\.\d+)?(?:[-+._A-Za-z0-9]*)?)(?![\d.])"
)


def _clean_version(value: Any) -> Optional[str]:
    """Return a plausible Minecraft version string from ``value``."""
    if not isinstance(value, str):
        return None
    value = value.strip()
    if not value:
        return None
    title_match = MC_VERSION_IN_TITLE_RE.search(value)
    if title_match:
        return title_match.group("version")
    token_match = MC_VERSION_TOKEN_RE.fullmatch(value)
    if token_match:
        return token_match.group("version")
    return value


def parse_game_version_from_window_title(title: str) -> Optional[str]:
    """Extract a Minecraft version from a window title like ``Minecraft 1.21.4``."""
    match = MC_VERSION_IN_TITLE_RE.search(title)
    return match.group("version") if match else None


def _load_json_object(path: Path) -> Optional[Any]:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return None


def detect_game_version_from_metadata(session_dir: os.PathLike[str] | str) -> Optional[str]:
    """Read ``metadata.json`` in priority order: ``mc_version``, then ``game_version``."""
    metadata = _load_json_object(Path(session_dir) / "metadata.json")
    if not isinstance(metadata, dict):
        return None
    for key in ("mc_version", "game_version"):
        version = _clean_version(metadata.get(key))
        if version:
            return version
    return None


def _iter_json_strings(value: Any) -> Iterator[str]:
    if isinstance(value, str):
        yield value
    elif isinstance(value, dict):
        for child in value.values():
            yield from _iter_json_strings(child)
    elif isinstance(value, list):
        for child in value:
            yield from _iter_json_strings(child)


def detect_game_version_from_fps_log(session_dir: os.PathLike[str] | str) -> Optional[str]:
    """Parse Minecraft version from any window-title string in ``fps_log.json``."""
    fps_log = _load_json_object(Path(session_dir) / "fps_log.json")
    if fps_log is None:
        return None
    for value in _iter_json_strings(fps_log):
        version = parse_game_version_from_window_title(value)
        if version:
            return version
    return None


def _version_from_manifest_json(path: Path) -> Optional[str]:
    data = _load_json_object(path)
    if not isinstance(data, dict):
        return None

    # Fabric profile JSONs carry the real MC version in inheritsFrom.
    for key in ("inheritsFrom", "minecraft_version", "version_id"):
        version = _clean_version(data.get(key))
        if version:
            return version

    pin = data.get("pin")
    if isinstance(pin, dict):
        for key in ("minecraft_version", "version_id"):
            version = _clean_version(pin.get(key))
            if version:
                return version

    version = _clean_version(data.get("id"))
    if version and not version.startswith("fabric-loader-"):
        return version
    return None


def _version_from_path_name(name: str) -> Optional[str]:
    matches = [m.group("version") for m in MC_VERSION_TOKEN_RE.finditer(name)]
    if not matches:
        return None
    # Fabric profile names contain both loader and MC versions; the MC version is last.
    return matches[-1]


def _candidate_install_roots(install_root: Optional[os.PathLike[str] | str]) -> List[Path]:
    candidates: List[Path] = []
    if install_root:
        candidates.append(Path(install_root))
    for env_name in ("OYSTER_INSTALL_ROOT", "GAMEDATA_INSTALL_ROOT"):
        value = os.environ.get(env_name)
        if value:
            candidates.append(Path(value))
    localappdata = os.environ.get("LOCALAPPDATA")
    if localappdata:
        candidates.extend(
            [
                Path(localappdata) / "OysterRecorder",
                Path(localappdata) / "GameDataRecorder",
            ]
        )

    unique: List[Path] = []
    for candidate in candidates:
        if candidate not in unique:
            unique.append(candidate)
    return unique


def detect_game_version_from_launcher_manifest(
    install_root: Optional[os.PathLike[str] | str] = None,
) -> Optional[str]:
    """Detect the bundled Minecraft version from ``mc-instance/versions`` manifests."""
    for root in _candidate_install_roots(install_root):
        instance = root / "mc-instance"
        for manifest_name in ("manifest-fabric.json", "manifest-mc.json"):
            version = _version_from_manifest_json(instance / manifest_name)
            if version:
                return version

        versions_dir = instance / "versions"
        if not versions_dir.is_dir():
            continue

        versions: List[str] = []
        for child in sorted(versions_dir.iterdir()):
            if not child.is_dir():
                continue
            manifest = child / f"{child.name}.json"
            version = _version_from_manifest_json(manifest) or _version_from_path_name(child.name)
            if version and version not in versions:
                versions.append(version)
        if len(versions) == 1:
            return versions[0]
        if len(versions) > 1:
            logger.warning(
                "Multiple Minecraft versions found under %s: %s; leaving game_version unset",
                versions_dir,
                ", ".join(versions),
            )
            return None
    return None


def detect_game_version(
    session_dir: Optional[os.PathLike[str] | str] = None,
    install_root: Optional[os.PathLike[str] | str] = None,
) -> Optional[str]:
    """Resolve game version from session metadata, fps log, then bundled install manifest."""
    if session_dir is not None:
        version = detect_game_version_from_metadata(session_dir)
        if version:
            return version
        version = detect_game_version_from_fps_log(session_dir)
        if version:
            return version
    return detect_game_version_from_launcher_manifest(install_root)


def build_gameinfo_dict(
    game_name: str = "Minecraft",
    game_version: Optional[str] = None,
    platform: str = "Java Edition",
    scene_name: str = "flat-overworld",
    weather: str = "clear",
    time_of_day: str = "day",
    character_name: str = "DataPilot",
    character_class: str = "spectator",
    operator_id: str = "vendor-001-op-A",
    recording_date: Optional[str] = None,
    total_frames: int = 9000,
    video_duration_sec: float = 300.0,
    route_type: int = 1,
    notes: str = "",
) -> Dict[str, Any]:
    """
    Build a dictionary with all 14 gameinfo fields.

    Args:
        game_name: Name of the game
        game_version: Version of the game
        platform: Platform/edition (e.g., "Java Edition")
        scene_name: Name of the scene or world
        weather: Weather condition
        time_of_day: Time of day (day/night/etc)
        character_name: Name of the character/player
        character_class: Class of the character
        operator_id: ID of the operator
        recording_date: Date of recording (None = today ISO date)
        total_frames: Total number of frames in the video
        video_duration_sec: Video duration in seconds
        route_type: Route type (1, 2, or 3)
        notes: Additional notes

    Returns:
        Dictionary with all 14 fields
    """
    if recording_date is None:
        recording_date = date.today().isoformat()

    return {
        "game_name": game_name,
        "game_version": game_version,
        "platform": platform,
        "scene_name": scene_name,
        "weather": weather,
        "time_of_day": time_of_day,
        "character_name": character_name,
        "character_class": character_class,
        "operator_id": operator_id,
        "recording_date": recording_date,
        "total_frames": total_frames,
        "video_duration_sec": video_duration_sec,
        "route_type": route_type,
        "notes": notes,
    }


def _escape_xml(s: str) -> str:
    """Escape special XML characters."""
    return (
        s.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&apos;")
    )


def _write_xlsx_fallback(data: Dict[str, Any], out_path: str) -> None:
    """
    Write XLSX file without openpyxl, using stdlib zipfile + XML.
    This is a minimal XLSX writer that creates a valid .xlsx file.
    """
    # Ensure field order matches FIELD_NAMES
    values = ["" if data.get(field) is None else str(data.get(field, "")) for field in FIELD_NAMES]

    # Build worksheet XML
    cells_xml = ""
    for col_idx, value in enumerate(values):
        col_letter = chr(65 + col_idx)  # A, B, C, ...
        cell_ref = f"{col_letter}2"
        # Determine cell type
        if value.isdigit() or (value.replace(".", "", 1).isdigit()):
            cell_type = 't="n"'
            cell_value = value
        else:
            cell_type = 't="inlineStr"'
            cell_value = f"<is><t>{_escape_xml(value)}</t></is>"
        cells_xml += f'<cell r="{cell_ref}" {cell_type}>{cell_value}</cell>\n'

    # Build sheet XML
    sheet_xml = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
    <sheetViews>
        <sheetView tabSelected="1" workbookViewId="0"/>
    </sheetViews>
    <cols>
        <col min="1" max="14" width="15" customWidth="1"/>
    </cols>
    <sheetData>
        <row r="1">
"""
    for col_idx, field in enumerate(FIELD_NAMES):
        col_letter = chr(65 + col_idx)
        cell_ref = f"{col_letter}1"
        sheet_xml += f'            <cell r="{cell_ref}" t="inlineStr"><is><t>{_escape_xml(field)}</t></is></cell>\n'

    sheet_xml += """        </row>
        <row r="2">
"""
    sheet_xml += cells_xml
    sheet_xml += """        </row>
    </sheetData>
</worksheet>"""

    # Build workbook.xml
    workbook_xml = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
    <sheets>
        <sheet name="gameinfo" sheetId="1" r:id="rId1"/>
    </sheets>
</workbook>"""

    # Build [Content_Types].xml
    content_types = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
    <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
    <Default Extension="xml" ContentType="application/xml"/>
    <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
    <Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
</Types>"""

    # Build _rels/.rels
    rels = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
    <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
</Relationships>"""

    # Build xl/_rels/workbook.xml.rels
    workbook_rels = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
    <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
</Relationships>"""

    # Write the XLSX file
    with zipfile.ZipFile(out_path, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", content_types)
        zf.writestr("_rels/.rels", rels)
        zf.writestr("xl/workbook.xml", workbook_xml)
        zf.writestr("xl/_rels/workbook.xml.rels", workbook_rels)
        zf.writestr("xl/worksheets/sheet1.xml", sheet_xml)


def write_xlsx(data: Dict[str, Any], out_path: str) -> None:
    """
    Write data to an XLSX file.

    Uses openpyxl if available, otherwise falls back to stdlib zipfile.

    Args:
        data: Dictionary with field names as keys
        out_path: Path to output XLSX file
    """
    try:
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "gameinfo"

        # Write header row (field names)
        for col_idx, field in enumerate(FIELD_NAMES, start=1):
            ws.cell(row=1, column=col_idx, value=field)

        # Write data row
        for col_idx, field in enumerate(FIELD_NAMES, start=1):
            value = data.get(field, "")
            ws.cell(row=2, column=col_idx, value=value)

        wb.save(out_path)

    except ImportError:
        # Fallback: use stdlib zipfile to create XLSX
        _write_xlsx_fallback(data, out_path)


def read_xlsx(path: str) -> Dict[str, Any]:
    """
    Read an XLSX file and return a dictionary.

    Reads the first sheet: row 1 as keys, row 2 as values.

    Args:
        path: Path to the XLSX file

    Returns:
        Dictionary with field names as keys
    """
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, data_only=True)
        ws = wb.active

        # Read header row
        keys = []
        col = 1
        while True:
            cell_value = ws.cell(row=1, column=col).value
            if cell_value is None:
                break
            keys.append(str(cell_value))
            col += 1

        # Read data row. Preserve blank cells in the middle of the schema.
        values = [ws.cell(row=2, column=col).value for col in range(1, len(keys) + 1)]

        # Combine keys and values
        result = {}
        for key, value in zip(keys, values):
            result[key] = value

        return result

    except ImportError:
        # Fallback: parse XML manually
        with zipfile.ZipFile(path, "r") as zf, zf.open("xl/worksheets/sheet1.xml") as f:
            content = f.read().decode("utf-8")

        # Extract cell values using regex
        # Find row 1 (header)
        keys = []
        values = []

        # Simple XML parsing with regex
        cell_pattern = re.compile(r'<cell r="([A-Z]+)(\d+)"[^>]*>(.*?)</cell>')
        for match in cell_pattern.finditer(content):
            row_num = int(match.group(2))
            cell_content = match.group(3)

            # Extract text from <is><t>...</t></is> or direct value
            text_match = re.search(r"<t>(.*?)</t>", cell_content, re.DOTALL)
            if text_match:
                cell_value = text_match.group(1)
            else:
                # Numeric value
                cell_value = cell_content.strip()

            if row_num == 1:
                keys.append(cell_value)
            elif row_num == 2:
                values.append(cell_value)

        result = {}
        for key, value in zip(keys, values):
            result[key] = value

        return result


def validate_route_type(route_type: int) -> bool:
    """Validate that route_type is one of 1, 2, or 3."""
    return route_type in (1, 2, 3)


def main(argv: Optional[List[str]] = None) -> int:
    """
    CLI entry point.

    Args:
        argv: Command line arguments (defaults to sys.argv)

    Returns:
        Exit code (0 for success, non-zero for error)
    """
    parser = argparse.ArgumentParser(description="Generate a gameinfo.xlsx file with 14 fields")

    parser.add_argument("--output", "-o", required=True, help="Output XLSX file path")

    parser.add_argument("--game-name", default="Minecraft")
    parser.add_argument(
        "--game-version",
        default=None,
        help=(
            "Explicit game version override. If omitted, detect from session "
            "metadata.json, fps_log.json, then bundled mc-instance manifests."
        ),
    )
    parser.add_argument(
        "--session-dir",
        default=None,
        help="Session directory containing metadata.json/fps_log.json (defaults to output parent).",
    )
    parser.add_argument(
        "--install-root",
        default=None,
        help="Recorder install root containing mc-instance/versions for version fallback.",
    )
    parser.add_argument("--platform", default="Java Edition")
    parser.add_argument("--scene-name", default="flat-overworld")
    parser.add_argument("--weather", default="clear")
    parser.add_argument("--time-of-day", default="day")
    parser.add_argument("--character-name", default="DataPilot")
    parser.add_argument("--character-class", default="spectator")
    parser.add_argument("--operator-id", default="vendor-001-op-A")
    parser.add_argument("--recording-date", default=None)
    parser.add_argument("--total-frames", type=int, default=9000)
    parser.add_argument("--video-duration-sec", type=float, default=300.0)
    parser.add_argument("--route-type", type=int, default=1)
    parser.add_argument("--notes", default="")

    args = parser.parse_args(argv)

    # Validate route_type
    if not validate_route_type(args.route_type):
        print(f"Error: route_type must be 1, 2, or 3, got {args.route_type}", file=sys.stderr)
        return 1

    output_path = Path(args.output)
    session_dir = Path(args.session_dir) if args.session_dir else output_path.parent
    game_version = (
        args.game_version.strip() if isinstance(args.game_version, str) else args.game_version
    )
    if not game_version:
        game_version = detect_game_version(
            session_dir=session_dir,
            install_root=args.install_root,
        )
        if game_version is None:
            logger.warning(
                "Could not detect Minecraft version for %s; game_version will be blank",
                session_dir,
            )

    # Build the data dictionary
    data = build_gameinfo_dict(
        game_name=args.game_name,
        game_version=game_version,
        platform=args.platform,
        scene_name=args.scene_name,
        weather=args.weather,
        time_of_day=args.time_of_day,
        character_name=args.character_name,
        character_class=args.character_class,
        operator_id=args.operator_id,
        recording_date=args.recording_date,
        total_frames=args.total_frames,
        video_duration_sec=args.video_duration_sec,
        route_type=args.route_type,
        notes=args.notes,
    )

    # Write the XLSX file
    try:
        write_xlsx(data, args.output)
        print(f"Successfully wrote {args.output}")
        return 0
    except Exception as e:
        print(f"Error writing XLSX file: {e}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    sys.exit(main())
