#!/usr/bin/env python3
"""
G165 · bin/batch_tracker.py — Stream rc18-batchtracker

Batch route_type distribution tracker per PRD §4.1.

PRD §4.1 requires per batch:
  - 50% route_type=1 (normal)
  - 25% route_type=2 (special)
  - 25% route_type=3 (loop)

This tool reads ALL session_dirs under the recordings directory,
extracts route_type from each session's gameinfo.xlsx (or action_camera.json),
and computes the cumulative distribution. Alerts if any band deviates
from PRD spec by > 10 percentage points.

CLI: python batch_tracker.py [--recordings-root <path>] [--verbose]
"""
from __future__ import annotations

import argparse
import json
import logging
import os
import sys
from pathlib import Path
from typing import Dict, List, Optional

logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)


# PRD §4.1 target distribution
TARGET_DISTRIBUTION = {
    1: 0.50,  # normal: 50%
    2: 0.25,  # special: 25%
    3: 0.25,  # loop: 25%
}

TOLERANCE_PERCENT = 10  # Alert if deviation > 10 percentage points


def get_default_recordings_root() -> Path:
    """Get default recordings root from LOCALAPPDATA or fallback."""
    localappdata = os.environ.get("LOCALAPPDATA")
    if localappdata:
        return Path(localappdata) / "GameData Recorder" / "recordings"
    # Fallback for non-Windows or missing env
    return Path.home() / "GameData Recorder" / "recordings"


def find_session_dirs(recordings_root: Path) -> List[Path]:
    """Find all session directories under recordings root."""
    if not recordings_root.exists():
        logger.warning(f"Recordings root not found: {recordings_root}")
        return []
    
    session_dirs = []
    for item in recordings_root.iterdir():
        if item.is_dir():
            # Check if it looks like a session (has gameinfo.xlsx or action_camera.json)
            if (item / "gameinfo.xlsx").exists() or (item / "action_camera.json").exists():
                session_dirs.append(item)
            # Also check for nested sessions
            for subitem in item.iterdir():
                if subitem.is_dir():
                    if (subitem / "gameinfo.xlsx").exists() or (subitem / "action_camera.json").exists():
                        session_dirs.append(subitem)
    
    return sorted(session_dirs)


def read_route_type_from_gameinfo(session_dir: Path) -> Optional[int]:
    """Read route_type from gameinfo.xlsx (PRD §3.3 field)."""
    gameinfo_path = session_dir / "gameinfo.xlsx"
    if not gameinfo_path.exists():
        return None
    
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(gameinfo_path), read_only=True)
        ws = wb.active
        if ws is None:
            wb.close()
            return None
        
        # Read headers
        headers = []
        for cell in ws[1]:
            val = cell.value
            if val is not None:
                headers.append(str(val).strip())
        
        if "route_type" not in headers:
            return None
        
        col_idx = headers.index("route_type") + 1
        
        # Read first data row (row 2)
        route_type_val = None
        for row in ws.iter_rows(min_row=2, max_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    try:
                        route_type_val = int(cell.value)
                    except (ValueError, TypeError):
                        try:
                            route_type_val = int(float(cell.value))
                        except (ValueError, TypeError):
                            pass
        
        wb.close()
        return route_type_val
        
    except ImportError:
        logger.debug("openpyxl not available, falling back to action_camera.json")
        return None
    except Exception as e:
        logger.debug(f"Failed to read gameinfo.xlsx: {e}")
        return None


def read_route_type_from_action_camera(session_dir: Path) -> Optional[int]:
    """Read route_type from action_camera.json (fallback)."""
    action_camera_path = session_dir / "action_camera.json"
    if not action_camera_path.exists():
        return None
    
    try:
        with open(action_camera_path, "r") as f:
            data = json.load(f)
        
        # action_camera.json is a list of frames, each may have route_type
        if isinstance(data, list) and len(data) > 0:
            # Get route_type from first frame
            first_frame = data[0]
            if isinstance(first_frame, dict) and "route_type" in first_frame:
                return int(first_frame["route_type"])
        
        # Or it might be a dict with route_type at top level
        if isinstance(data, dict) and "route_type" in data:
            return int(data["route_type"])
        
        return None
        
    except Exception as e:
        logger.debug(f"Failed to read action_camera.json: {e}")
        return None


def get_route_type(session_dir: Path) -> Optional[int]:
    """Get route_type from session directory (gameinfo.xlsx preferred)."""
    # Try gameinfo.xlsx first (PRD §3.3)
    route_type = read_route_type_from_gameinfo(session_dir)
    if route_type is not None:
        return route_type
    
    # Fallback to action_camera.json
    return read_route_type_from_action_camera(session_dir)


def compute_distribution(session_dirs: List[Path], verbose: bool = False) -> Dict[int, int]:
    """Compute route_type distribution across all sessions."""
    counts: Dict[int, int] = {1: 0, 2: 0, 3: 0}
    unknown = 0
    errors: List[tuple] = []
    
    for session_dir in session_dirs:
        route_type = get_route_type(session_dir)
        
        if route_type is None:
            unknown += 1
            if verbose:
                errors.append((session_dir.name, "route_type not found"))
        elif route_type in counts:
            counts[route_type] += 1
        else:
            unknown += 1
            if verbose:
                errors.append((session_dir.name, f"invalid route_type: {route_type}"))
    
    if verbose and errors:
        logger.debug(f"Errors reading route_type: {errors}")
    if verbose and unknown:
        logger.debug(f"Sessions with unknown route_type: {unknown}")
    
    return counts


def check_distribution(counts: Dict[int, int], total: int) -> List[Dict]:
    """Check if distribution deviates from target by more than tolerance."""
    alerts = []
    
    if total == 0:
        alerts.append({
            "type": "error",
            "message": "No sessions with valid route_type found",
        })
        return alerts
    
    for route_type, target_pct in TARGET_DISTRIBUTION.items():
        actual_count = counts.get(route_type, 0)
        actual_pct = actual_count / total * 100
        target_pct_val = target_pct * 100
        
        deviation = actual_pct - target_pct_val
        
        if abs(deviation) > TOLERANCE_PERCENT:
            alerts.append({
                "type": "alert",
                "route_type": route_type,
                "actual_pct": round(actual_pct, 1),
                "target_pct": target_pct_val,
                "deviation": round(deviation, 1),
                "count": actual_count,
                "total": total,
                "message": (
                    f"route_type={route_type}: actual {actual_pct:.1f}% "
                    f"vs target {target_pct_val:.0f}% "
                    f"(deviation {deviation:+.1f}% exceeds {TOLERANCE_PERCENT}% tolerance)"
                ),
            })
    
    return alerts


def print_report(counts: Dict[int, int], total: int, alerts: List[Dict], verbose: bool = False) -> None:
    """Print distribution report."""
    print("\n" + "=" * 60)
    print("Batch Route Type Distribution Report")
    print("=" * 60)
    
    print(f"\nTotal sessions analyzed: {total}")
    
    print("\nDistribution:")
    print("-" * 40)
    for route_type in [1, 2, 3]:
        count = counts.get(route_type, 0)
        pct = count / total * 100 if total > 0 else 0
        target = TARGET_DISTRIBUTION[route_type] * 100
        status = "OK" if abs(pct - target) <= TOLERANCE_PERCENT else "ALERT"
        print(f"  route_type={route_type}: {count} sessions ({pct:.1f}%) "
              f"[target: {target:.0f}%] {status}")
    
    print("-" * 40)
    
    if alerts:
        print("\nALERTS:")
        for alert in alerts:
            if alert.get("type") == "error":
                print(f"  ERROR: {alert['message']}")
            else:
                print(f"  {alert['message']}")
    else:
        print("\nAll route types within tolerance")
    
    print("\n" + "=" * 60)


def main(argv: Optional[List[str]] = None) -> int:
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description="Batch route_type distribution tracker (PRD section 4.1)")
    parser.add_argument(
        "--recordings-root", "-r", type=Path, default=None,
        help="Path to recordings root directory")
    parser.add_argument(
        "--verbose", "-v", action="store_true",
        help="Enable verbose output")
    args = parser.parse_args(argv)
    
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    
    recordings_root = args.recordings_root or get_default_recordings_root()
    logger.info(f"Recordings root: {recordings_root}")
    
    session_dirs = find_session_dirs(recordings_root)
    logger.info(f"Found {len(session_dirs)} session directories")
    
    if not session_dirs:
        logger.warning("No session directories found")
        print("\nNo sessions found to analyze.")
        return 0
    
    counts = compute_distribution(session_dirs, verbose=args.verbose)
    total = sum(counts.values())
    
    alerts = check_distribution(counts, total)
    print_report(counts, total, alerts, verbose=args.verbose)
    
    # Return non-zero if there are alerts
    return 1 if alerts else 0


if __name__ == "__main__":
    sys.exit(main())