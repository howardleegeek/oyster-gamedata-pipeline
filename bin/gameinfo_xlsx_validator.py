#!/usr/bin/env python3
"""Validate gameinfo.xlsx against buyer-spec: 3 sheets + required fields."""

import argparse
import sys

REQUIRED_FIELDS = {
    "metadata": ["game_id", "game_name", "version", "developer", "publisher"],
    "scene_table": ["scene_id", "scene_name", "scene_type", "background_asset"],
    "asset_ramp": ["asset_id", "asset_name", "asset_type", "priority", "load_order"],
}

EXPECTED_SHEETS = set(REQUIRED_FIELDS.keys())


def _load_openpyxl():
    """Lazy-import openpyxl to avoid startup cost when not needed."""
    try:
        import openpyxl

        return openpyxl
    except ImportError:
        print("ERROR: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
        sys.exit(1)


def validate_xlsx(xlsx_path: str) -> dict:
    """Read *xlsx_path* and validate sheets + required fields.

    Returns a dict with keys:
        ok          – bool, True when every check passes
        sheets_found – list of sheet names present in the workbook
        missing_sheets – list of expected sheet names that are absent
        field_errors – dict mapping sheet_name -> list of missing field names
    """
    openpyxl = _load_openpyxl()
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet_names = set(wb.sheetnames)

    missing_sheets = sorted(EXPECTED_SHEETS - sheet_names)
    sheets_found = sorted(EXPECTED_SHEETS & sheet_names)

    field_errors: dict[str, list[str]] = {}
    for sname in sheets_found:
        ws = wb[sname]
        rows = ws.iter_rows(min_row=1, max_row=1, values_only=True)
        headers = [str(c).strip() for c in next(rows, ()) if c is not None]
        missing = [f for f in REQUIRED_FIELDS[sname] if f not in headers]
        if missing:
            field_errors[sname] = missing

    wb.close()

    ok = (len(missing_sheets) == 0) and (len(field_errors) == 0)
    return {
        "ok": ok,
        "sheets_found": sheets_found,
        "missing_sheets": missing_sheets,
        "field_errors": field_errors,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Validate gameinfo.xlsx per buyer-spec")
    parser.add_argument("--xlsx", required=True, help="Path to gameinfo.xlsx")
    args = parser.parse_args()

    result = validate_xlsx(args.xlsx)

    if result["ok"]:
        print("PASS: all sheets and required fields present.")
    else:
        if result["missing_sheets"]:
            print(f"FAIL: missing sheets -> {result['missing_sheets']}")
        for sname, missing in result["field_errors"].items():
            print(f"FAIL: sheet '{sname}' missing fields -> {missing}")
        sys.exit(1)

    print(result)


if __name__ == "__main__":
    main()
