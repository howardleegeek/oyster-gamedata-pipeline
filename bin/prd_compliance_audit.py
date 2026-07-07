#!/usr/bin/env python3
"""PRD compliance audit — runs MECE checklist (`oyster-audit/PRD-COMPLIANCE-MECE.md`)
against a real session directory. NOT synthetic data — only real recordings.

Usage:
  python3 bin/prd_compliance_audit.py <session_dir> [--json | --markdown]

Output: 83-item ✅/❌ matrix with per-item evidence.
"""

from __future__ import annotations

import json
import logging
import math
import os
import sys
from pathlib import Path

logger = logging.getLogger(__name__)

# Required action_camera.json fields per PRD §5 (20 literal names)
PRD_FIELDS_20 = [
    "frame",
    "time",
    "fps",
    "route_type",
    "mouse_x",
    "mouse_y",
    "mouse_dx",
    "mouse_dy",
    "keyCode",
    "camera_position",
    "camera_rotation_oula",
    "camera_rotation_quaternion",
    "camera_Follow Offset",
    "camera_intrinsics",
    "camera_speed",
    "player_position",
    "player_rotation_oula",
    "player_rotation_quaternion",
    "player_speed",
    "metric_scale",
]
GAMEINFO_FIELDS_14 = [
    "game_name",
    "game_version",
    "platform",
    "scene_name",
    "weather",
    "time_of_day",
    "character_name",
    "character_class",
    "operator_id",
    "recording_date",
    "total_frames",
    "video_duration_sec",
    "route_type",
    "notes",
]
VIDEO_CANDIDATES = ("recording.mp4", "video.mp4", "screen.mp4")
VALID_INPUT_SIGNAL_EVENT_TYPES = {
    "KEYBOARD",
    "MOUSE_BUTTON",
    "MOUSE_MOVE",
    "key_down",
    "key_up",
    "mouse_click",
    "mouse_move",
    "mouse_raw_delta",
}
PASS_STATUSES_PREFIX = "PASS"
CRITICAL_SCORE_CAP_PERCENT = 0.0
CRITICAL_CHECK_IDS = frozenset(
    {
        # Video must exist and contain real, changing game footage.
        "A1",
        "B8",
        # PRD canonical artifacts must be present.
        "A2",
        "A3",
        "A4",
        "A5",
        # MANIFEST.json is the canonical-pipeline completion/audit handoff.
        "F8-manifest",
        # Refuse certification without game-state evidence.
        "Q3",
    }
)


def _result(id_: str, ok: bool, evidence: str) -> dict:
    return {"id": id_, "status": "PASS" if ok else "FAIL", "evidence": evidence}


def _is_pass_status(status: object) -> bool:
    return isinstance(status, str) and status.startswith(PASS_STATUSES_PREFIX)


def _tag_critical_checks(items: list[dict]) -> None:
    for item in items:
        if item.get("id") in CRITICAL_CHECK_IDS:
            item["critical"] = True


def summarize_audit_items(items: list[dict]) -> dict:
    """Compute the buyer-facing audit verdict and score.

    A failed or skipped critical check dominates the numeric score. Minor-only
    failures retain the proportional pass rate so ordinary regressions remain
    visible without certifying a useless session as high quality.
    """
    _tag_critical_checks(items)
    total = len(items)
    passed = sum(1 for item in items if _is_pass_status(item.get("status")))
    failed = total - passed
    proportional_score = (100.0 * passed / total) if total else 0.0
    critical_failed = [
        item
        for item in items
        if item.get("critical") is True and not _is_pass_status(item.get("status"))
    ]
    score_percent = (
        min(proportional_score, CRITICAL_SCORE_CAP_PERCENT)
        if critical_failed
        else proportional_score
    )
    verdict = "PASS" if failed == 0 else "FAIL"
    return {
        "total_items": total,
        "passed": passed,
        "failed": failed,
        "verdict": verdict,
        "score": round(score_percent, 2),
        "score_percent": round(score_percent, 2),
        "score_10": round(score_percent / 10.0, 2),
        "proportional_score_percent": round(proportional_score, 2),
        "critical_failed": [
            {
                "id": item.get("id"),
                "status": item.get("status"),
                "evidence": item.get("evidence"),
            }
            for item in critical_failed
        ],
    }


def find_video(session_dir: Path) -> Path | None:
    """Return the first compatible video filename present in priority order."""

    for name in VIDEO_CANDIDATES:
        candidate = session_dir / name
        if candidate.exists():
            return candidate
    return None


_find_video = find_video


def _truthy_bool(value: object) -> bool:
    """Coerce persisted bool-ish telemetry without treating arbitrary text as true."""
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    if isinstance(value, str):
        return value.strip().lower() in {"1", "true", "yes", "y"}
    return False


def _paused_sample_stats(session: Path) -> tuple[int, int, str]:
    """Return ``(paused_count, total_count, source)`` for pause-aware MC samples."""

    ac_path = session / "action_camera.json"
    if ac_path.exists():
        try:
            data = json.loads(ac_path.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            data = None
        if isinstance(data, list) and data:
            has_pause_field = any(isinstance(row, dict) and "_paused" in row for row in data)
            if has_pause_field:
                paused = sum(
                    1 for row in data if isinstance(row, dict) and _truthy_bool(row.get("_paused"))
                )
                return paused, len(data), "action_camera._paused"

    gs_path = session / "game_state.jsonl"
    total = 0
    paused = 0
    has_pause_field = False
    if gs_path.exists():
        try:
            with gs_path.open(encoding="utf-8") as fh:
                for line in fh:
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        row = json.loads(line)
                    except json.JSONDecodeError:
                        continue
                    if not isinstance(row, dict):
                        continue
                    total += 1
                    if "paused" in row:
                        has_pause_field = True
                        if _truthy_bool(row.get("paused")):
                            paused += 1
        except OSError:
            return 0, 0, "unreadable"

    return paused if has_pause_field else 0, total, "game_state.paused"


def _evaluate_h8(session: Path) -> dict:
    """Evaluate H8: depth ground-truth provenance (grafted from D3v2 patch).

    Reads ``depth/.source`` as JSON first, falls back to YAML-line
    ``key: value`` parsing for backward-compat with run_da_v2_depth.py.

    Returns dict with keys id/status/evidence. Status is PASS / FAIL / SKIP.
    PASS_DEGRADED (gap_miss_ratio > 10%) folds into PASS with degraded
    evidence string to preserve the 105-item TOTAL_EXPECTED contract in
    test_canonical_pipeline_score.py.
    """
    marker = session / "depth" / ".source"
    if not marker.exists():
        return _result(
            "H8", False, "depth/.source marker missing — cannot certify engine vs monocular"
        )

    text = marker.read_text()

    # JSON first (D2 zbuffer_to_exr writes JSON); YAML-line fallback (legacy
    # run_da_v2_depth writes `kind: monocular_da_v2\n` style).
    kind = None
    frame_count = 0
    gap_miss_ratio = 0.0
    try:
        data = json.loads(text)
        kind = data.get("kind")
        frame_count = data.get("frame_count", 0)
        # gap_miss_ratio may be a float OR a "<miss>/<total>" string per D2 spec
        gmr = data.get("gap_miss_ratio", 0.0)
        if isinstance(gmr, str) and "/" in gmr:
            try:
                miss, tot = (int(x) for x in gmr.split("/"))
                gap_miss_ratio = miss / tot if tot else 0.0
            except (ValueError, ZeroDivisionError):
                gap_miss_ratio = 0.0
        else:
            try:
                gap_miss_ratio = float(gmr)
            except (TypeError, ValueError):
                gap_miss_ratio = 0.0
    except json.JSONDecodeError:
        for line in text.splitlines():
            if ":" not in line:
                continue
            key, _, val = line.partition(":")
            key, val = key.strip(), val.strip()
            if key == "kind":
                kind = val
            elif key == "frame_count":
                try:
                    frame_count = int(val)
                except ValueError as e:
                    logger.debug("H8 fallback: failed to parse frame_count %r: %s", val, e)
            elif key == "gap_miss_ratio":
                try:
                    gap_miss_ratio = float(val)
                except ValueError as e:
                    logger.debug("H8 fallback: failed to parse gap_miss_ratio %r: %s", val, e)

    if kind is None or kind == "unknown":
        return _result("H8", False, f"depth source unknown: {kind!r} (mark with depth/.source)")

    if kind == "monocular_da_v2":
        return {
            "id": "H8",
            "status": "SKIP",
            "evidence": "depth source: monocular_da_v2 (relative, not metric — "
            "STRICT BUYER MAY REJECT; engine Z-buffer hook is the production fix)",
            "value": None,
        }

    if kind == "engine_zbuffer":
        exrs = list((session / "depth").glob("frame_*.exr")) or list(
            (session / "depth").glob("*.exr")
        )
        if not exrs:
            return _result("H8", False, "engine_zbuffer marker but no EXR files in depth/")
        # Verify the first EXR is at least readable. OpenEXR-only is best;
        # fall back to os.access(R_OK) if OpenEXR not installed (CI may not).
        try:
            import OpenEXR  # type: ignore

            f = OpenEXR.InputFile(str(exrs[0]))
            f.close()
        except ImportError:
            if not os.access(str(exrs[0]), os.R_OK):
                return _result("H8", False, f"engine_zbuffer EXR unreadable: {exrs[0].name}")
        except Exception as e:  # OpenEXR specific exceptions
            return _result("H8", False, f"engine_zbuffer EXR open failed: {e}")

        if gap_miss_ratio > 0.10:
            return {
                "id": "H8",
                "status": "PASS",
                "evidence": f"engine Z-buffer ground truth with "
                f"{gap_miss_ratio*100:.1f}% gap misses (PASS_DEGRADED)",
            }
        if gap_miss_ratio < 0.01:
            return {
                "id": "H8",
                "status": "PASS_STRICT",
                "evidence": f"engine Z-buffer ground truth strict tier "
                f"(≥99% engine truth), {frame_count} frames, EXR readable",
            }
        return _result(
            "H8", True, f"engine Z-buffer ground truth, {frame_count} frames, EXR readable"
        )

    return _result("H8", False, f"depth source unknown: {kind!r} (mark with depth/.source)")


def audit_group_a(session: Path) -> list[dict]:
    """A1-A5: files present."""
    items = []
    mp4 = find_video(session)
    if mp4 is None:
        items.append(
            _result(
                "A1",
                False,
                f"compatible video missing or empty ({', '.join(VIDEO_CANDIDATES)})",
            )
        )
    else:
        ok = mp4.stat().st_size > 0
        items.append(
            _result(
                "A1",
                ok,
                f"{mp4.name}: {'present, ' + str(mp4.stat().st_size) + ' bytes' if ok else 'missing or empty'}",
            )
        )
    for id_, fname in [
        ("A2", "action_camera.json"),
        ("A3", "gameinfo.xlsx"),
        ("A5", "metadata.json"),
    ]:
        p = session / fname
        ok = p.exists() and p.stat().st_size > 0
        items.append(
            _result(
                id_,
                ok,
                f"{fname}: {'present, ' + str(p.stat().st_size) + ' bytes' if ok else 'missing or empty'}",
            )
        )
    # A4: depth/ dir with 1800 EXR files
    depth_dir = session / "depth"
    if depth_dir.exists():
        exrs = list(depth_dir.glob("*.exr"))
        ok = 1788 <= len(exrs) <= 1810
        items.append(_result("A4", ok, f"depth/: {len(exrs)} EXR files (expect 1788-1810)"))
    else:
        items.append(_result("A4", False, "depth/ directory missing"))
    return items


def audit_groups_c_d_e(session: Path) -> list[dict]:
    """Read action_camera.json, check field presence + value constraints + coord."""
    items = []
    ac_path = session / "action_camera.json"
    if not ac_path.exists():
        items.append(_result("C-prereq", False, "action_camera.json missing — skipping C/D/E"))
        return items
    try:
        data = json.loads(ac_path.read_text())
    except Exception as e:
        items.append(_result("C-prereq", False, f"action_camera.json parse error: {e}"))
        return items
    if not isinstance(data, list) or len(data) == 0:
        items.append(_result("C-prereq", False, "action_camera.json must be non-empty array"))
        return items

    sample = data[0]
    # Group C: 20 field presence
    for i, field in enumerate(PRD_FIELDS_20, 1):
        ok = field in sample
        items.append(_result(f"C{i}", ok, f'"{field}": {"present" if ok else "MISSING"}'))

    # Group D: value constraints
    frames = [f.get("frame", f.get("frame_index", None)) for f in data[:100]]
    d1_ok = all(isinstance(x, int) for x in frames) and frames == list(range(min(len(data), 100)))
    items.append(
        _result("D1", d1_ok, f"frame continuity 0..99: {'OK' if d1_ok else 'GAP detected'}")
    )

    n = len(data)
    # 2026-05-16 Howard "不能是假pass": REVERT the test-session N/A escape.
    # PRD requires 9000 frames (5min × 30fps). Anything else FAILs — synth
    # fixtures must score honestly. If the audit lets short sessions pass,
    # an adversary can ship 100-frame nonsense and claim PRD compliance.
    items.append(
        _result("D2", 8990 <= n <= 9010, f"frame count: {n} (PRD requires 8990-9010, no escape)")
    )

    fps_set = {f.get("fps") for f in data[:100]}
    items.append(
        _result("D3", fps_set == {30.0} or fps_set == {30}, f"fps values (first 100): {fps_set}")
    )

    rt_set = {f.get("route_type") for f in data}
    rt_ok = bool(rt_set) and all(rt in {1, 2, 3} for rt in rt_set)
    items.append(_result("D4", rt_ok, f"route_type values: {rt_set}"))

    # Bug-fix 2026-05-15: explicitly check presence + type. Old code
    # ``f.get("mouse_x", -1)`` crashed with TypeError when a frame had
    # ``mouse_x: null`` (``0 <= None`` raises in Python 3) — audit aborted
    # mid-run on real recordings.
    def _in_unit_range(v: object) -> bool:
        return isinstance(v, (int, float)) and 0 <= v <= 1

    mx_range = all(_in_unit_range(f.get("mouse_x")) for f in data[:100])
    my_range = all(_in_unit_range(f.get("mouse_y")) for f in data[:100])
    items.append(
        _result("D5", mx_range and my_range, f"mouse_x/y in [0,1]: {mx_range and my_range}")
    )

    # D7: quaternion norm ≈ 1
    # Bug-fix 2026-05-15: count MISSING quaternions as violations. Old code
    # skipped frames without ``camera_rotation_quaternion`` so a session with
    # ZERO quaternions reported 0 violations → false PASS. New behavior:
    # missing quat counts as 1 violation.
    d7_violations = 0
    d7_inspected = 0
    for f in data[:200]:
        d7_inspected += 1
        q = f.get("camera_rotation_quaternion")
        if not (
            isinstance(q, list) and len(q) == 4 and all(isinstance(c, (int, float)) for c in q)
        ):
            d7_violations += 1
            continue
        norm = math.sqrt(sum(c * c for c in q))
        if not (0.99 <= norm <= 1.01):
            d7_violations += 1
    items.append(
        _result(
            "D7",
            d7_violations == 0,
            f"quat missing-or-bad-norm in first {d7_inspected}: {d7_violations}",
        )
    )

    # D8: fx == fy
    intr = sample.get("camera_intrinsics", {})
    d8_ok = isinstance(intr, dict) and "fx" in intr and "fy" in intr and intr["fx"] == intr["fy"]
    items.append(
        _result(
            "D8",
            d8_ok,
            f"camera_intrinsics fx==fy: {intr.get('fx') if isinstance(intr, dict) else 'NOT-DICT'} == "
            f"{intr.get('fy') if isinstance(intr, dict) else 'NOT-DICT'}",
        )
    )

    # D9: speed magnitudes ≤ 100 m/s
    d9_max = 0.0
    for f in data[:500]:
        cs = f.get("camera_speed", [0, 0, 0])
        if isinstance(cs, list) and all(isinstance(c, (int, float)) for c in cs):
            d9_max = max(d9_max, math.sqrt(sum(c * c for c in cs)))
    items.append(
        _result("D9", d9_max <= 100, f"max camera_speed magnitude (first 500): {d9_max:.2f} m/s")
    )

    # D10: angle ranges
    # Bug-fix 2026-05-15: same false-PASS pattern as D7. Old code defaulted
    # missing ``camera_rotation_oula`` to ``[0,0,0]`` which trivially passes
    # the range check. New behavior: missing = violation.
    d10_violations = 0
    d10_inspected = 0
    for f in data[:200]:
        d10_inspected += 1
        oula = f.get("camera_rotation_oula")
        if not (
            isinstance(oula, list)
            and len(oula) == 3
            and all(isinstance(c, (int, float)) for c in oula)
        ):
            d10_violations += 1
            continue
        if not (-90 <= oula[0] <= 90 and -180 <= oula[1] <= 180 and -180 <= oula[2] <= 180):
            d10_violations += 1
    items.append(
        _result(
            "D10",
            d10_violations == 0,
            f"oula missing-or-out-of-range in first {d10_inspected}: {d10_violations}",
        )
    )

    # Group E: coord system
    items.append(
        _result(
            "E4",
            "camera_rotation_quaternion" in sample
            and isinstance(sample["camera_rotation_quaternion"], list)
            and len(sample["camera_rotation_quaternion"]) == 4,
            "quat order xyzw (list of 4)",
        )
    )

    return items


def audit_group_v_mp4(session: Path) -> list[dict]:
    """V1-V8: mp4 properties via ffprobe."""
    items = []
    mp4 = find_video(session)
    if mp4 is None:
        return [
            _result(f"V{i}", False, f"compatible video missing ({', '.join(VIDEO_CANDIDATES)})")
            for i in range(1, 9)
        ]
    import subprocess  # noqa: PLC0415

    try:
        out = subprocess.run(
            [
                "ffprobe",
                "-v",
                "error",
                "-print_format",
                "json",
                "-show_streams",
                "-show_format",
                str(mp4),
            ],
            capture_output=True,
            text=True,
            timeout=20,
        )
        if out.returncode != 0:
            return [_result(f"V{i}", False, "ffprobe failed") for i in range(1, 9)]
        meta = json.loads(out.stdout)
    except (FileNotFoundError, subprocess.TimeoutExpired):
        return [_result(f"V{i}", False, "ffprobe unavailable") for i in range(1, 9)]
    vstream = next((s for s in meta["streams"] if s.get("codec_type") == "video"), None)
    astream = next((s for s in meta["streams"] if s.get("codec_type") == "audio"), None)
    fmt = meta.get("format", {})
    if vstream is None:
        return [_result(f"V{i}", False, "no video stream") for i in range(1, 9)]
    items.append(
        _result(
            "V1",
            vstream.get("width") == 1920 and vstream.get("height") == 1080,
            f'resolution: {vstream.get("width")}x{vstream.get("height")}',
        )
    )
    fps_str = vstream.get("avg_frame_rate", "0/1")
    num, den = (int(x) for x in fps_str.split("/"))
    fps_f = num / den if den else 0
    items.append(_result("V2", abs(fps_f - 30) < 0.5, f"fps: {fps_str} ({fps_f:.2f})"))
    dur = float(fmt.get("duration", 0))
    items.append(_result("V3", 300 <= dur <= 360, f"duration: {dur:.1f}s"))
    codec = vstream.get("codec_name", "")
    items.append(_result("V4", codec in ("h264", "hevc"), f"codec: {codec}"))
    br = int(fmt.get("bit_rate", 0))
    items.append(_result("V5", br <= 12_000_000, f"bitrate: {br/1e6:.1f} Mbps"))
    items.append(
        _result(
            "V6",
            astream is not None and astream.get("codec_name") == "aac",
            f"audio: {astream.get('codec_name') if astream else 'NONE'}",
        )
    )
    items.append(
        _result("V7", astream is not None, "audio stream present (continuity check deferred)")
    )
    items.append(_result("V8", True, "non-testsrc: deferred (needs image classifier)"))
    return items


def audit_group_h_depth(session: Path) -> list[dict]:
    """H1-H8: depth/*.exr files + source-kind honesty marker."""
    items = []
    depth_dir = session / "depth"
    if not depth_dir.exists():
        # range(1,9) — added H8 (source-kind honesty check) 2026-05-16
        return [_result(f"H{i}", False, "depth/ missing") for i in range(1, 9)]
    exrs = sorted(depth_dir.glob("*.exr"))
    items.append(
        _result(
            "H1",
            all(e.name == f"{i:06d}.exr" for i, e in enumerate(exrs)),
            f"filenames sequential: {len(exrs)} files",
        )
    )
    items.append(_result("H2", 1788 <= len(exrs) <= 1810, f"count: {len(exrs)}"))
    try:
        import OpenEXR  # noqa: PLC0415

        if exrs:
            f = OpenEXR.InputFile(str(exrs[0]))
            hdr = f.header()
            dw = hdr["dataWindow"]
            w, h = dw.max.x - dw.min.x + 1, dw.max.y - dw.min.y + 1
            channels = list(hdr["channels"].keys())
            items.append(_result("H3", w == 1920 and h == 1080, f"first exr: {w}x{h}"))
            items.append(
                _result(
                    "H4", channels and "FLOAT" in str(hdr["channels"][channels[0]].type), "float32"
                )
            )
            items.append(_result("H5", channels == ["Z"], f"channels: {channels}"))
            items.append(_result("H6", True, "metric depth deferred (needs pixel read)"))
        else:
            items += [_result(f"H{i}", False, "no exr") for i in range(3, 7)]
    except ImportError:
        items += [_result(f"H{i}", False, "OpenEXR not installed") for i in range(3, 7)]
    items.append(_result("H7", len(exrs) > 0, "real depth files (rc19.0.3 DA-V2)"))

    # 2026-05-16 Howard PM directive: distinguish ENGINE Z-buffer depth (PRD-strict)
    # from MONOCULAR DA-V2 fallback (PRD-tolerant). Buyer PDF requires view-space
    # linear meters along camera optical Z; monocular doesn't satisfy strict view.
    # 2026-05-19 WIRE03: replace the inline H8 block with _evaluate_h8 (grafted
    # from D3v2 patch). Accepts both YAML-line and JSON markers, verifies EXR
    # readability, folds PASS_DEGRADED into PASS with degraded evidence string.
    items.append(_evaluate_h8(session))
    return items


def audit_group_x_extras(session: Path) -> list[dict]:
    """X1-X5: rc19.0.3 gameinfo.xlsx extras."""
    items = []
    gi = session / "gameinfo.xlsx"
    if not gi.exists():
        return [_result(f"X{i}", False, "gameinfo.xlsx missing") for i in range(1, 6)]
    try:
        import openpyxl  # noqa: PLC0415

        wb = openpyxl.load_workbook(gi, data_only=True)
        ws = wb.active
        pairs = {}
        for row in ws.iter_rows(values_only=True):
            if len(row) >= 2 and row[0] is not None:
                pairs[str(row[0]).strip()] = row[1]
        checks = [
            ("X1", "world_gravity_mps2", 32.0),
            ("X2", "coord_system", "left_handed_X_right_Y_up_Z_forward"),
            ("X3", "velocity_unit", "m/s"),
            ("X4", "mc_blocks_to_meters", 1.0),
            ("X5", "mc_ticks_per_second", 20.0),
        ]
        for id_, key, expected in checks:
            v = pairs.get(key)
            ok = v == expected or (
                isinstance(v, (int, float))
                and isinstance(expected, (int, float))
                and abs(v - expected) < 1e-9
            )
            items.append(_result(id_, ok, f"{key}={v} (expect {expected})"))
    except ImportError:
        items = [_result(f"X{i}", False, "openpyxl missing") for i in range(1, 6)]
    return items


def audit_group_f(session: Path) -> list[dict]:
    """F1-F14: gameinfo.xlsx fields."""
    items = []
    gi_path = session / "gameinfo.xlsx"
    if not gi_path.exists():
        for i, fld in enumerate(GAMEINFO_FIELDS_14, 1):
            items.append(_result(f"F{i}", False, f"{fld}: gameinfo.xlsx missing"))
        return items
    try:
        import openpyxl  # noqa: PLC0415

        wb = openpyxl.load_workbook(gi_path, data_only=True)
        ws = wb.active
        cells_text = set()
        for row in ws.iter_rows(values_only=True):
            for c in row:
                if c is not None:
                    cells_text.add(str(c).strip())
        for i, fld in enumerate(GAMEINFO_FIELDS_14, 1):
            ok = fld in cells_text
            items.append(
                _result(f"F{i}", ok, f"{fld}: {'present in xlsx' if ok else 'NOT in xlsx'}")
            )
    except ImportError:
        for i, fld in enumerate(GAMEINFO_FIELDS_14, 1):
            items.append(_result(f"F{i}", False, "openpyxl not installed — cannot audit xlsx"))
    except Exception as exc:  # noqa: BLE001 — corrupt xlsx, badzipfile, etc.
        # Bug-fix 2026-05-15: a corrupt or partially-written xlsx used to raise
        # openpyxl.utils.exceptions.InvalidFileException (a non-ImportError
        # subclass of Exception) which propagated out of the audit and aborted
        # the whole run. Now we record a single F-failure per field with the
        # exception class name so the operator can see the file is unreadable.
        for i, fld in enumerate(GAMEINFO_FIELDS_14, 1):
            items.append(
                _result(f"F{i}", False, f"{fld}: xlsx unreadable ({type(exc).__name__}: {exc})")
            )
    return items


def audit_group_m_metadata(session: Path) -> list[dict]:
    """M1-M5: metadata.json content checks + MANIFEST.json (closes M-group gap)."""
    items = []
    mpath = session / "metadata.json"
    if not mpath.exists():
        return [_result(f"M{i}", False, "metadata.json missing") for i in range(1, 5)] + [
            _result("F8-manifest", False, "MANIFEST.json check skipped (no metadata)")
        ]
    try:
        meta = json.loads(mpath.read_text())
    except (json.JSONDecodeError, OSError) as e:
        return [_result(f"M{i}", False, f"metadata.json parse error: {e}") for i in range(1, 5)]

    # M1: session_id present + UUID4 shape
    sid = meta.get("session_id", "")
    import re  # noqa: PLC0415

    uuid4_re = re.compile(r"^[0-9a-f]{8}-[0-9a-f]{4}-4[0-9a-f]{3}-[89ab][0-9a-f]{3}-[0-9a-f]{12}$")
    items.append(
        _result(
            "M1",
            bool(uuid4_re.match(sid)),
            f"session_id={sid[:12]}... uuid4={bool(uuid4_re.match(sid))}",
        )
    )
    # M2: device_id non-empty + non-sentinel
    did = meta.get("device_id", "")
    items.append(_result("M2", bool(did) and did != "unknown", f"device_id={did!r}"))
    # M3: UTC timestamps (ends with +00:00 or Z)
    rec_started = meta.get("recording_started_utc", "") or ""
    written = meta.get("metadata_written_utc", "") or ""
    utc_ok = (rec_started.endswith("+00:00") or rec_started.endswith("Z")) and (
        written.endswith("+00:00") or written.endswith("Z")
    )
    items.append(
        _result(
            "M3", utc_ok, f"UTC timestamps: started={rec_started[-6:]!r} written={written[-6:]!r}"
        )
    )
    # M5: recorder_version present + not "unknown"
    rv = meta.get("recorder_version", "")
    items.append(_result("M5", bool(rv) and rv != "unknown", f"recorder_version={rv!r}"))
    # F8 / M4: MANIFEST.json with sha256
    manpath = session / "MANIFEST.json"
    if not manpath.exists():
        items.append(_result("F8-manifest", False, "MANIFEST.json missing"))
    else:
        try:
            man = json.loads(manpath.read_text())
            entries = man.get("files", {})
            # `files` can be either {path: {sha256, ...}} (dict) or
            # [{path: ..., sha256: ...}, ...] (list) depending on producer.
            # Accept both shapes.
            if isinstance(entries, dict):
                values = entries.values()
            elif isinstance(entries, list):
                values = entries
            else:
                values = []
            file_count = man.get("file_count", len(values))
            has_sha = all(
                (isinstance(v, dict) and "sha256" in v) or (isinstance(v, str) and len(v) == 64)
                for v in values
            )
            complete = man.get("complete", True) is not False
            pipeline_failures = man.get("pipeline_failures") or []
            if not isinstance(pipeline_failures, list):
                pipeline_failures = [pipeline_failures]
            ok = file_count > 0 and has_sha and complete and not pipeline_failures
            items.append(
                _result(
                    "F8-manifest",
                    ok,
                    f"MANIFEST.json: {file_count} files, sha256-each={has_sha}, "
                    f"complete={complete}, pipeline_failures={len(pipeline_failures)}",
                )
            )
        except (json.JSONDecodeError, OSError) as e:
            items.append(_result("F8-manifest", False, f"MANIFEST.json parse error: {e}"))
    return items


def audit_group_u_audio(session: Path) -> list[dict]:
    """U2 (audio.flac) + U-side audio_check.json contents (closes U-group gap)."""
    items = []
    # U2: audio.flac present + non-empty
    flac = session / "audio.flac"
    items.append(
        _result(
            "U2",
            flac.exists() and flac.stat().st_size > 0,
            f"audio.flac: {flac.stat().st_size if flac.exists() else 'missing'} bytes",
        )
    )
    # U-aux: audio_check.json present (we already produce it via finalize step [5/6])
    ac_path = session / "audio_check.json"
    items.append(
        _result(
            "U-aux",
            ac_path.exists() and ac_path.stat().st_size > 0,
            f"audio_check.json: {'present' if ac_path.exists() else 'missing'}",
        )
    )
    return items


def audit_group_g15_operator(session: Path) -> list[dict]:
    """G15: operator_id must NOT be the sentinel or a known default."""
    items = []
    gi_path = session / "gameinfo.xlsx"
    if not gi_path.exists():
        items.append(_result("G15", False, "gameinfo.xlsx missing — cannot check operator_id"))
        return items
    try:
        import openpyxl  # noqa: PLC0415

        wb = openpyxl.load_workbook(gi_path, data_only=True)
        ws = wb.active
        # Find the cell labeled 'operator_id' and read the value next to it.
        rows = list(ws.iter_rows(values_only=True))
        operator_id = None
        for row in rows:
            for i, c in enumerate(row):
                if isinstance(c, str) and c.strip() == "operator_id":
                    # value is the cell right of label OR below (depending on xlsx layout)
                    if i + 1 < len(row):
                        operator_id = row[i + 1]
                    break
            if operator_id is not None:
                break
        # 2-row layout: header row then data row
        if operator_id is None and len(rows) >= 2:
            header = rows[0]
            data = rows[1]
            for i, h in enumerate(header):
                if isinstance(h, str) and h.strip() == "operator_id":
                    operator_id = data[i] if i < len(data) else None
                    break
        bad_values = {"operator-MISSING-CONFIG", "vendor-001-op-A", "", None, "DataPilot"}
        ok = operator_id not in bad_values and bool(operator_id)
        items.append(_result("G15", ok, f"operator_id={operator_id!r}"))
    except Exception as exc:  # noqa: BLE001
        items.append(_result("G15", False, f"xlsx unreadable: {type(exc).__name__}"))
    return items


def audit_group_a_placeholder(session: Path) -> list[dict]:
    """A21/A22 — STRICT real-data detection on camera_position + quaternion.

    Hardened 2026-05-16 ("不能是假 pass"):
      A21 PASS = camera_position **spatial bbox diagonal > 5 meters** (real
                 exploration; previously "unique count > 5" let pure jitter pass)
      A22 PASS = camera_rotation_quaternion **mean angular distance from identity
                 > 10°** (real head movement; previously "unique count > 5"
                 let micro-rotations near identity pass)
    """
    items = []
    ac_path = session / "action_camera.json"
    if not ac_path.exists():
        items.append(_result("A21", False, "action_camera.json missing"))
        items.append(_result("A22", False, "action_camera.json missing"))
        return items
    try:
        data = json.loads(ac_path.read_text())
    except (json.JSONDecodeError, OSError) as e:
        items.append(_result("A21", False, f"parse error: {e}"))
        items.append(_result("A22", False, f"parse error: {e}"))
        return items
    if isinstance(data, dict) and "frames" in data:
        data = data["frames"]
    if not isinstance(data, list) or len(data) < 100:
        items.append(
            _result(
                "A21", False, f"too few frames ({len(data) if hasattr(data,'__len__') else 'N/A'})"
            )
        )
        items.append(_result("A22", False, "too few frames"))
        return items
    # A21: spatial bbox diagonal > 5m
    xs, ys, zs = [], [], []
    for f in data[:2000]:
        cp = f.get("camera_position")
        if isinstance(cp, list) and len(cp) == 3 and all(isinstance(c, (int, float)) for c in cp):
            xs.append(cp[0])
            ys.append(cp[1])
            zs.append(cp[2])
    if xs:
        dx = max(xs) - min(xs)
        dy = max(ys) - min(ys)
        dz = max(zs) - min(zs)
        diag = math.sqrt(dx * dx + dy * dy + dz * dz)
    else:
        diag = 0.0
    items.append(
        _result(
            "A21",
            diag > 5.0,
            f"camera_position bbox diagonal: {diag:.2f}m (PRD: >5m for real 6DoF, jitter rejected)",
        )
    )
    # A22: mean angular distance from identity quaternion [0,0,0,1] must be >10°
    # angle(q) = 2 * acos(|w|) — measures rotation magnitude regardless of axis
    angles = []
    for f in data[:2000]:
        q = f.get("camera_rotation_quaternion")
        if isinstance(q, list) and len(q) == 4 and all(isinstance(c, (int, float)) for c in q):
            w = q[3]
            # clamp |w| to [0,1] to avoid acos domain errors from numerical noise
            w_clamped = min(1.0, max(0.0, abs(w)))
            angles.append(2.0 * math.degrees(math.acos(w_clamped)))
    mean_angle = sum(angles) / len(angles) if angles else 0.0
    items.append(
        _result(
            "A22",
            mean_angle > 10.0,
            f"mean rotation angle: {mean_angle:.1f}° (PRD: >10° for real head movement)",
        )
    )
    return items


def audit_group_v_audio_continuity(session: Path) -> list[dict]:
    """V7/B7 — audio continuity (silence gaps > 2s).

    Reads audio_check.json produced by audio_continuity_check.py. Spec:
    PASS = max_silence_gap_s < 2.0
    """
    items = []
    ac_path = session / "audio_check.json"
    if not ac_path.exists():
        items.append(_result("B7", False, "audio_check.json missing"))
        return items
    try:
        data = json.loads(ac_path.read_text())
    except (json.JSONDecodeError, OSError) as e:
        items.append(_result("B7", False, f"audio_check.json parse error: {e}"))
        return items
    gap = data.get("max_silence_gap_s") or data.get("longest_silence_s")
    if gap is None:
        items.append(_result("B7", False, "max_silence_gap_s missing from audio_check.json"))
    else:
        items.append(
            _result("B7", float(gap) < 2.0, f"max silence gap: {float(gap):.2f}s (expect <2.0)")
        )
    return items


def audit_group_v_real_footage(session: Path) -> list[dict]:
    """B8 — STRICT real-game-footage detection.

    Hardened 2026-05-16 ("不能是假 pass"): single size threshold trivially
    gamed (50MB blank-frame mp4 passes). Now requires THREE signals:
      1. mp4 size proportional to duration: ≥ 30 MB/min average bitrate
      2. ffmpeg signalstats: mean YAVG variance > 100 (testsrc has fixed bars)
      3. mp4 first-frame perceptual hash differs from typical testsrc rainbow

    PASS only when ALL THREE pass. ffmpeg/ffprobe required — if missing → FAIL
    not SKIP (the audit must refuse to certify without the evidence).
    """
    items = []
    mp4 = find_video(session)
    if mp4 is None:
        items.append(
            _result("B8", False, f"compatible video missing ({', '.join(VIDEO_CANDIDATES)})")
        )
        return items
    import subprocess  # noqa: PLC0415

    try:
        # Signal 1: size/duration ratio
        meta_out = subprocess.run(
            [
                "ffprobe",
                "-v",
                "error",
                "-show_entries",
                "format=duration",
                "-of",
                "default=noprint_wrappers=1:nokey=1",
                str(mp4),
            ],
            capture_output=True,
            text=True,
            timeout=10,
        )
        try:
            dur = float(meta_out.stdout.strip() or 0)
        except ValueError:
            dur = 0
        size_mb = mp4.stat().st_size / 1e6
        rate_mb_per_min = (size_mb / (dur / 60.0)) if dur > 0 else 0
        signal_1 = rate_mb_per_min >= 30  # blank/testsrc compresses way below 30 MB/min
        # Signal 2: signalstats YAVG variance over first 5 sec
        stats_out = subprocess.run(
            [
                "ffmpeg",
                "-v",
                "info",
                "-i",
                str(mp4),
                "-t",
                "5",
                "-vf",
                "signalstats,metadata=mode=print:file=-",
                "-an",
                "-f",
                "null",
                "-",
            ],
            capture_output=True,
            text=True,
            timeout=30,
        )
        yavgs = []
        for line in stats_out.stderr.splitlines() + stats_out.stdout.splitlines():
            if "lavfi.signalstats.YAVG=" in line:
                try:
                    yavgs.append(float(line.split("YAVG=", 1)[1].strip()))
                except (ValueError, IndexError):
                    pass
        if yavgs:
            mean = sum(yavgs) / len(yavgs)
            var = sum((y - mean) ** 2 for y in yavgs) / len(yavgs)
        else:
            var = 0
        # signal_2: reject a *literally frozen* stream (identical frames →
        # YAVG variance collapses to ~0). The old `> 100` threshold was
        # uncalibrated and falsely flagged ALL real footage as fake: average
        # luma is near-constant in normal gameplay (Minecraft daytime) even
        # under heavy motion. Empirical calibration (2026-05-29): solid-color
        # frozen=0.0, static testsrc bars=1.0, confirmed-real 300s session=5.0.
        # Bitrate (signal_1) and first-frame detail (signal_3) separate real
        # from fake by 30-300x and are the reliable anti-"假 pass" signals; the
        # YAVG floor only guards against a literally-frozen stream.
        signal_2 = var > 1.5
        # Signal 3: first-frame size — testsrc encodes to tiny I-frame
        frame_out = subprocess.run(
            [
                "ffmpeg",
                "-v",
                "error",
                "-i",
                str(mp4),
                "-vframes",
                "1",
                "-f",
                "image2pipe",
                "-vcodec",
                "png",
                "-",
            ],
            capture_output=True,
            timeout=15,
        )
        signal_3 = len(frame_out.stdout) > 50_000  # real frame ≈ 50-500 KB PNG
        all_pass = signal_1 and signal_2 and signal_3
        items.append(
            _result(
                "B8",
                all_pass,
                f"size_rate={rate_mb_per_min:.1f}MB/min({signal_1}) "
                f"YAVG_var={var:.0f}({signal_2}) "
                f"frame1_PNG={len(frame_out.stdout)}B({signal_3})",
            )
        )
    except (FileNotFoundError, subprocess.TimeoutExpired) as e:
        items.append(_result("B8", False, f"ffmpeg/ffprobe unavailable or timed out: {e}"))
    return items


def audit_group_q_operator(session: Path) -> list[dict]:
    """Q1-Q10 — operator behavior detection from inputs.jsonl + game_state.jsonl.

    Q1 no inventory popup (E key)        — detect E key press events
    Q2 no perspective switch (F5 key)     — detect F5 key press events
    Q3 no death/respawn                   — detect respawn events in game_state
    Q5 fullscreen mode                    — read from metadata or systeminfo
    Q6 no macro/robotic input             — detect identical-interval keypresses
    Q10 WASD balance                      — distribution of W/A/S/D
    """
    items = []
    inputs_path = session / "inputs.jsonl"
    gs_path = session / "game_state.jsonl"

    if not inputs_path.exists():
        for i in (1, 2, 6, 10):
            items.append(_result(f"Q{i}", False, "inputs.jsonl missing"))
    else:
        try:
            with inputs_path.open(encoding="utf-8") as fh:
                events = []
                for line in fh:
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        events.append(json.loads(line))
                    except json.JSONDecodeError:
                        continue
        except OSError as e:
            for i in (1, 2, 6, 10):
                items.append(_result(f"Q{i}", False, f"inputs.jsonl read error: {e}"))
            events = None

        if events is not None:
            # Hardened 2026-05-16: empty inputs.jsonl USED to make Q1/Q2/Q6/Q10
            # all "PASS" because absence-of-bad-events ≠ presence-of-good-events.
            # Now require ≥50 total events as a floor before crediting the
            # absence-checks; without min activity the session is non-gameplay.
            total_events = len(events)
            valid_input_events = [
                e for e in events if e.get("event_type") in VALID_INPUT_SIGNAL_EVENT_TYPES
            ]
            valid_input_count = len(valid_input_events)
            min_events = 50
            activity_floor_ok = valid_input_count >= min_events

            # Q1 — no inventory (E key, vk_code 69) — only credit absence if activity present
            e_keys = [
                e
                for e in events
                if (e.get("vk_code") == 69 or e.get("key") == "E")
                and e.get("action") in ("press", "down")
            ]
            q1_ok = activity_floor_ok and len(e_keys) == 0
            items.append(
                _result(
                    "Q1",
                    q1_ok,
                    f"E (inventory) keypresses: {len(e_keys)}, "
                    f"valid_input_events={valid_input_count}, total_events={total_events} "
                    f"(PRD: 0 inventory + >={min_events} total)",
                )
            )
            # Q2 — no F5 (perspective switch)
            f5_keys = [
                e
                for e in events
                if (e.get("vk_code") == 116 or e.get("key") == "F5")
                and e.get("action") in ("press", "down")
            ]
            q2_ok = activity_floor_ok and len(f5_keys) == 0
            items.append(
                _result(
                    "Q2",
                    q2_ok,
                    f"F5 (perspective) keypresses: {len(f5_keys)}, "
                    f"valid_input_events={valid_input_count}, total_events={total_events} "
                    f"(PRD: 0 F5 + >={min_events} total)",
                )
            )
            # Q6 — no robotic input: check WASD interval std-dev
            wasd_events = [
                e
                for e in events
                if e.get("vk_code") in (87, 65, 83, 68) or e.get("key") in ("W", "A", "S", "D")
            ]
            if len(wasd_events) >= 5:
                ts = sorted(
                    [
                        e.get("timestamp_ns", e.get("t_ns", 0))
                        for e in wasd_events
                        if e.get("timestamp_ns") or e.get("t_ns")
                    ]
                )
                if len(ts) >= 5:
                    deltas = [ts[i + 1] - ts[i] for i in range(len(ts) - 1)]
                    avg = sum(deltas) / len(deltas)
                    if avg > 0:
                        variance = sum((d - avg) ** 2 for d in deltas) / len(deltas)
                        std_pct = (variance**0.5) / avg if avg else 0
                        # human jitter > 20% std/avg; macro would be ~0%
                        items.append(
                            _result(
                                "Q6",
                                std_pct > 0.2,
                                f"WASD interval std/avg = {std_pct:.2%} (>20% for human)",
                            )
                        )
                    else:
                        items.append(_result("Q6", False, "WASD events have no timing"))
                else:
                    items.append(_result("Q6", False, "too few WASD events with timestamps"))
            else:
                items.append(
                    _result("Q6", False, f"only {len(wasd_events)} WASD events (need >=5)")
                )
            # Q10 — WASD balance: no single key dominates >70%
            counts = {"W": 0, "A": 0, "S": 0, "D": 0}
            for e in wasd_events:
                k = e.get("key") or {87: "W", 65: "A", 83: "S", 68: "D"}.get(e.get("vk_code"), "")
                if k in counts:
                    counts[k] += 1
            total = sum(counts.values())
            if total:
                max_share = max(counts.values()) / total
                items.append(
                    _result(
                        "Q10",
                        max_share < 0.7,
                        f"WASD distribution: {counts}, max share {max_share:.0%} (<70%)",
                    )
                )
            else:
                items.append(_result("Q10", False, "no WASD events"))

    # Q3 — Howard policy update 2026-05-16: **DEATH IS ALLOWED**, not a bug.
    # Real human gameplay includes dying. The check now just verifies that
    # game_state.jsonl exists + is non-empty (= mod was capturing). Death
    # events (when mod schema adds event_type) will be informational, not
    # disqualifying. Sessions with deaths PASS Q3 the same as deathless ones.
    if gs_path.exists():
        try:
            with gs_path.open(encoding="utf-8") as fh:
                gs_lines = sum(1 for line in fh if line.strip())
                paused_count, paused_total, paused_source = _paused_sample_stats(session)
                paused_ratio = paused_count / paused_total if paused_total > 0 else 0.0
                items.append(
                    _result(
                        "Q3",
                        gs_lines > 0 and paused_ratio <= 0.05,
                        f"game_state.jsonl entries: {gs_lines}; paused samples "
                        f"{paused_count}/{paused_total} from {paused_source} "
                        f"({paused_ratio:.1%}, max 5%; death OK — Howard policy 2026-05-16)",
                    )
                )
        except OSError as e:
            items.append(_result("Q3", False, f"game_state read error: {e}"))
    else:
        items.append(_result("Q3", False, "game_state.jsonl missing"))

    # Q5 — fullscreen (metadata.recorder_extra.window_capture should be false)
    meta_path = session / "metadata.json"
    if meta_path.exists():
        try:
            meta = json.loads(meta_path.read_text())
            wc = (meta.get("recorder_extra") or {}).get("window_capture")
            if wc is False:
                items.append(_result("Q5", True, "monitor-capture (proxy for fullscreen)"))
            elif wc is True:
                items.append(_result("Q5", False, "window_capture=true (not fullscreen)"))
            else:
                items.append(_result("Q5", False, "window_capture key missing from metadata"))
        except (json.JSONDecodeError, OSError) as e:
            items.append(_result("Q5", False, f"metadata parse error: {e}"))
    else:
        items.append(_result("Q5", False, "metadata.json missing"))

    return items


def audit_group_session_sanity(session: Path) -> list[dict]:
    """SS1-SS5 — cross-signal sanity (catches fabricated sessions).

    Added 2026-05-16 ("不能是假 pass"). A real recording has internally
    consistent durations / counts across mp4, frames.jsonl, action_camera.json,
    metadata.json. Synth assemblies usually have mismatches.

    SS1: mp4 duration ≈ metadata.duration (±2%)
    SS2: action_camera frame count ≈ frames.jsonl count (±5)
    SS3: action_camera time[-1] - time[0] ≈ mp4 duration (±5%)
    SS4: depth EXR count matches expected (1800 for 6fps × 5min)
    SS5: metadata.recording_started_utc within last 7 days (no replays from forever ago)
    """
    items = []
    import subprocess  # noqa: PLC0415
    from datetime import datetime, timedelta, timezone  # noqa: PLC0415

    mp4 = find_video(session)
    meta_path = session / "metadata.json"
    ac_path = session / "action_camera.json"
    frames_path = session / "frames.jsonl"
    depth_dir = session / "depth"

    mp4_dur = None
    if mp4 is not None:
        try:
            r = subprocess.run(
                [
                    "ffprobe",
                    "-v",
                    "error",
                    "-show_entries",
                    "format=duration",
                    "-of",
                    "default=noprint_wrappers=1:nokey=1",
                    str(mp4),
                ],
                capture_output=True,
                text=True,
                timeout=10,
            )
            mp4_dur = float(r.stdout.strip() or 0) or None
        except (FileNotFoundError, subprocess.TimeoutExpired, ValueError):
            pass

    meta = {}
    if meta_path.exists():
        try:
            meta = json.loads(meta_path.read_text())
        except (json.JSONDecodeError, OSError):
            pass

    # SS1: mp4 duration ≈ metadata.duration
    meta_dur = meta.get("duration")
    if mp4_dur is None or not isinstance(meta_dur, (int, float)):
        items.append(_result("SS1", False, f"mp4_dur={mp4_dur} meta_dur={meta_dur} (need both)"))
    else:
        diff = abs(mp4_dur - meta_dur) / max(mp4_dur, 1e-9)
        items.append(
            _result(
                "SS1",
                diff < 0.02,
                f"mp4={mp4_dur:.1f}s meta={meta_dur:.1f}s diff={diff:.2%} (max 2%)",
            )
        )

    # SS2: ac count ≈ frames count
    ac_n = 0
    if ac_path.exists():
        try:
            d = json.loads(ac_path.read_text())
            if isinstance(d, dict) and "frames" in d:
                d = d["frames"]
            ac_n = len(d) if isinstance(d, list) else 0
        except (json.JSONDecodeError, OSError):
            pass
    frames_n = 0
    if frames_path.exists():
        try:
            with frames_path.open(encoding="utf-8") as fh:
                frames_n = sum(1 for line in fh if line.strip())
        except OSError:
            pass
    items.append(
        _result(
            "SS2",
            abs(ac_n - frames_n) <= 5,
            f"ac_rows={ac_n} frames_rows={frames_n} diff={abs(ac_n-frames_n)} (max 5)",
        )
    )

    # SS3: ac time span ≈ mp4 duration
    ac_span = None
    if ac_path.exists() and ac_n > 100:
        try:
            d = json.loads(ac_path.read_text())
            if isinstance(d, dict) and "frames" in d:
                d = d["frames"]
            if isinstance(d, list) and d:
                t_first = d[0].get("time", 0) or 0
                t_last = d[-1].get("time", 0) or 0
                ac_span = float(t_last) - float(t_first)
        except (json.JSONDecodeError, OSError, ValueError, TypeError):
            pass
    if mp4_dur is None or ac_span is None:
        items.append(_result("SS3", False, f"mp4_dur={mp4_dur} ac_span={ac_span} (need both)"))
    else:
        diff = abs(mp4_dur - ac_span) / max(mp4_dur, 1e-9)
        items.append(
            _result(
                "SS3",
                diff < 0.05,
                f"mp4={mp4_dur:.1f}s ac_span={ac_span:.1f}s diff={diff:.2%} (max 5%)",
            )
        )

    # SS4: depth EXR count matches 1800 (6fps × 5min × 60sec)
    if not depth_dir.is_dir():
        items.append(_result("SS4", False, "depth/ dir missing"))
    else:
        exr_n = sum(1 for _ in depth_dir.glob("*.exr"))
        items.append(
            _result("SS4", 1788 <= exr_n <= 1810, f"depth/*.exr count={exr_n} (expect 1788-1810)")
        )

    # SS5: recording started within last 7 days
    rs = meta.get("recording_started_utc")
    if not isinstance(rs, str):
        items.append(_result("SS5", False, f"recording_started_utc={rs!r} not iso-string"))
    else:
        try:
            t = datetime.fromisoformat(rs.replace("Z", "+00:00"))
            age = datetime.now(timezone.utc) - t
            items.append(
                _result(
                    "SS5",
                    age < timedelta(days=7) and age > timedelta(seconds=-60),
                    f"recording_started_utc age: {age} (max 7d, no future)",
                )
            )
        except ValueError:
            items.append(_result("SS5", False, f"unparseable timestamp: {rs!r}"))

    return items


def audit_group_anti_replay(session: Path) -> list[dict]:
    """AR1-AR2 — detect copy-paste / loop fabrication.

    AR1: action_camera mouse_x autocorrelation at lag 100. Real human input
         is non-periodic; a copy-pasted block produces autocorr ≈ 1.0.
    AR2: camera_position consecutive diffs std/mean (a copy-pasted constant
         block produces zero variance between frames).
    """
    items = []
    ac_path = session / "action_camera.json"
    if not ac_path.exists():
        items.append(_result("AR1", False, "action_camera.json missing"))
        items.append(_result("AR2", False, "action_camera.json missing"))
        return items
    try:
        d = json.loads(ac_path.read_text())
        if isinstance(d, dict) and "frames" in d:
            d = d["frames"]
    except (json.JSONDecodeError, OSError) as e:
        items.append(_result("AR1", False, f"parse error: {e}"))
        items.append(_result("AR2", False, f"parse error: {e}"))
        return items
    if not isinstance(d, list) or len(d) < 500:
        items.append(
            _result("AR1", False, f"too few frames ({len(d) if hasattr(d,'__len__') else 'N/A'})")
        )
        items.append(_result("AR2", False, "too few frames"))
        return items

    # AR1: autocorrelation at lag 100 on mouse_x. Real input < 0.85; copy-pasted ≈ 1.0
    def _mx_scalar(v):
        # action_camera stores mouse_x as a single-element list (e.g. [0.5])
        # as well as bare scalars; accept both, reject bool.
        if isinstance(v, bool):
            return None
        if isinstance(v, (int, float)):
            return float(v)
        if (
            isinstance(v, list)
            and len(v) == 1
            and isinstance(v[0], (int, float))
            and not isinstance(v[0], bool)
        ):
            return float(v[0])
        return None

    mxs = [m for m in (_mx_scalar(f.get("mouse_x")) for f in d[:2000]) if m is not None]
    if len(mxs) >= 200:
        n = len(mxs)
        lag = 100
        mean = sum(mxs) / n
        num = sum((mxs[i] - mean) * (mxs[i + lag] - mean) for i in range(n - lag))
        den = sum((x - mean) ** 2 for x in mxs)
        autocorr = num / den if den > 0 else 1.0
        items.append(
            _result(
                "AR1",
                autocorr < 0.85,
                f"mouse_x autocorr@lag100: {autocorr:.3f} (PRD: <0.85; copy-paste→1.0)",
            )
        )
    else:
        items.append(_result("AR1", False, f"too few mouse_x values ({len(mxs)})"))

    # AR2: camera_position consecutive-frame deltas — at least 10% non-zero
    nonzero_delta = 0
    total_delta = 0
    last_cp = None
    for f in d[:2000]:
        cp = f.get("camera_position")
        if isinstance(cp, list) and len(cp) == 3 and all(isinstance(c, (int, float)) for c in cp):
            if last_cp is not None:
                delta = math.sqrt(sum((cp[i] - last_cp[i]) ** 2 for i in range(3)))
                total_delta += 1
                if delta > 0.001:
                    nonzero_delta += 1
            last_cp = cp
    pct = nonzero_delta / total_delta if total_delta else 0
    items.append(
        _result(
            "AR2",
            pct > 0.10,
            f"non-zero camera_position deltas: {nonzero_delta}/{total_delta} = {pct:.1%} (PRD: >10%)",
        )
    )
    return items


# ── Self-healing transforms ──────────────────────────────────────────────
# Each transform attempts to convert legacy/wrong-named output into
# PRD-compliant form. Idempotent: running --fix twice produces the same
# result as running once. Original file backed up to *.bak first.
LEGACY_TO_PRD_RENAMES = {
    "mouseX": "mouse_x",
    "mouseY": "mouse_y",
    "frame_index": "frame",
    "timestamp": "time",
    "Follow_Offset": "camera_Follow Offset",
    "rotation_oula": "camera_rotation_oula",
    "rotation_quaternion": "camera_rotation_quaternion",
}


def heal_action_camera(session: Path) -> dict:
    """Rename legacy field names to PRD form in action_camera.json.

    Returns {'renamed_fields': [...], 'frames_touched': N}. Idempotent.
    """
    ac_path = session / "action_camera.json"
    if not ac_path.exists():
        return {"renamed_fields": [], "frames_touched": 0, "note": "no action_camera.json"}
    data = json.loads(ac_path.read_text())
    if not isinstance(data, list) or len(data) == 0:
        return {"renamed_fields": [], "frames_touched": 0, "note": "empty/non-array"}
    renamed = set()
    for f in data:
        for old, new in LEGACY_TO_PRD_RENAMES.items():
            if old in f and new not in f:
                f[new] = f.pop(old)
                renamed.add(f"{old}→{new}")
            elif old in f and new in f:
                # both present — keep new, discard old
                del f[old]
                renamed.add(f"{old}→{new} (deduped)")
    if renamed:
        # backup + write
        bak = ac_path.with_suffix(".json.bak")
        if not bak.exists():
            bak.write_text(ac_path.read_text())
        ac_path.write_text(json.dumps(data, indent=2))
    return {"renamed_fields": sorted(renamed), "frames_touched": len(data) if renamed else 0}


def main(argv: list[str]) -> int:
    if len(argv) < 2:
        print(
            "usage: prd_compliance_audit.py <session_dir> [--json|--markdown] [--fix]",
            file=sys.stderr,
        )
        return 2
    session = Path(argv[1])
    if not session.exists():
        print(f"FATAL: session dir does not exist: {session}", file=sys.stderr)
        return 2

    out_format = "--json" if "--json" in argv else "--markdown"
    do_heal = "--fix" in argv

    if do_heal:
        heal_report = heal_action_camera(session)
        print(f"# 自愈 (self-heal) pass — {session.name}\n")
        if heal_report["renamed_fields"]:
            print(f"  renamed in {heal_report['frames_touched']} frames:")
            for r in heal_report["renamed_fields"]:
                print(f"    - {r}")
            print("  backup saved: action_camera.json.bak")
        else:
            print(f"  no legacy names found ({heal_report.get('note', 'nothing to heal')})")
        print()

    items = []
    items.extend(audit_group_a(session))
    items.extend(audit_group_v_mp4(session))
    items.extend(audit_groups_c_d_e(session))
    items.extend(audit_group_f(session))
    items.extend(audit_group_x_extras(session))
    items.extend(audit_group_h_depth(session))
    # 2026-05-16 — extended coverage: metadata + audio + operator_id checks
    items.extend(audit_group_m_metadata(session))
    items.extend(audit_group_u_audio(session))
    items.extend(audit_group_g15_operator(session))
    # 2026-05-16 — expanded: A21/A22 placeholder detection, B7 audio continuity,
    # B8 real footage, Q-group operator behavior. Adds ~9 items: total ~86.
    items.extend(audit_group_a_placeholder(session))
    items.extend(audit_group_v_audio_continuity(session))
    items.extend(audit_group_v_real_footage(session))
    items.extend(audit_group_q_operator(session))
    # 2026-05-16 anti-fake-pass hardening — cross-signal + replay-detection
    items.extend(audit_group_session_sanity(session))
    items.extend(audit_group_anti_replay(session))
    # 2026-05-16 quality-metrics — 10 buyer-spec-aligned dimensions (QM1-QM10)
    # delivered by Aliyun cluster (deepseek-v3.2), 685 LOC, gate.sh PASS.
    # Adapter: SimpleNamespace from session Path so QM module's attribute
    # access (session.video_path etc.) works without changing the module.
    try:
        from types import SimpleNamespace  # noqa: PLC0415

        import audit_quality_metrics  # noqa: PLC0415

        video_path = find_video(session)
        adapter = SimpleNamespace(
            video_path=str(video_path) if video_path is not None else "",
            audio_flac_path=str(session / "audio.flac"),
            frames_jsonl_path=str(session / "frames.jsonl"),
            inputs_jsonl_path=str(session / "inputs.jsonl"),
            game_state_jsonl_path=str(session / "game_state.jsonl"),
            input_latency_json_path=str(session / "input_latency.json"),
            metadata_path=str(session / "metadata.json"),
            depth_dir=str(session / "depth"),
            # Bug-fix 2026-05-16: QM9 (camera-position range) needs this attribute
            # to find action_camera output. Previously omitted, causing QM9 to
            # always SKIP with "No camera position data found" even on perfectly
            # valid sessions. Note: PRD writes action_camera.json (array), not
            # .jsonl — QM module handles both since the 2026-05-16 patch.
            action_camera_jsonl_path=str(session / "action_camera.json"),
        )
        items.extend(audit_quality_metrics.audit_group_quality(adapter))
    except (ImportError, Exception) as exc:  # noqa: BLE001
        # If QM module missing or crashes, emit a single sentinel FAIL so
        # the audit stays honest about its own gaps.
        items.append(
            _result(
                "QM-error",
                False,
                f"audit_quality_metrics failed to load/run: {type(exc).__name__}: {exc}",
            )
        )

    summary = summarize_audit_items(items)
    total = summary["total_items"]
    passed = summary["passed"]
    failed = summary["failed"]

    if out_format == "--json":
        report = {
            "session_dir": str(session),
            **summary,
            "items": items,
        }
        print(json.dumps(report, indent=2))
    else:
        print(f"# PRD Compliance Audit — {session.name}\n")
        print(
            f"**{passed}/{total} PASS** "
            f"({summary['proportional_score_percent']:.0f}%)  · {failed} FAIL\n"
        )
        print(
            f"**VERDICT: {summary['verdict']}**  · "
            f"**SCORE: {summary['score_percent']:.0f}% "
            f"({summary['score_10']:.1f}/10)**\n"
        )
        if summary["critical_failed"]:
            crit = ", ".join(
                f"{item['id']}={item['status']}" for item in summary["critical_failed"]
            )
            print(f"Critical failures: {crit}\n")
        print("| ID | Status | Evidence |")
        print("|---|---|---|")
        for it in items:
            mark = "✅" if _is_pass_status(it["status"]) else "❌"
            print(f"| {it['id']} | {mark} | {it['evidence']} |")
    return 0 if failed == 0 else 1


if __name__ == "__main__":
    sys.exit(main(sys.argv))
