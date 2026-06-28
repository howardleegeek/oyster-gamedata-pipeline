#!/usr/bin/env python3
"""Vendor batch manifest generator.

Scans a vendor batch directory for tarballs and generates manifest.yaml
according to SUBMISSION_FORMAT.md §4.1 schema.
"""

import hashlib
import json
import os
import re
import sys
import tarfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any


def compute_sha256(path: str, chunk_size: int = 65536) -> str:
    """SHA-256 hex digest of file."""
    sha256_hash = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(chunk_size), b""):
            sha256_hash.update(chunk)
    return sha256_hash.hexdigest()


def extract_clip_metadata(tarball_path: str) -> dict:
    """Open tarball read-only, extract action_camera.json + gameinfo.xlsx + count depth EXRs.
    
    Return {'clip_id': str, 'duration_sec': float, 'frame_count': int, 'depth_count': int,
            'route_type': int, 'scene': str, 'operator_id': str}.
    """
    result = {
        'clip_id': '',
        'duration_sec': 0.0,
        'frame_count': 0,
        'depth_count': 0,
        'route_type': 0,
        'scene': '',
        'operator_id': ''
    }
    
    clip_id = parse_clip_id(os.path.basename(tarball_path))
    result['clip_id'] = clip_id
    
    with tarfile.open(tarball_path, 'r:gz') as tar:
        members = tar.getmembers()
        
        # Count depth EXRs
        depth_count = sum(1 for m in members if m.name.endswith('.exr') and 'depth' in m.name.lower())
        result['depth_count'] = depth_count
        
        # Extract action_camera.json
        action_camera_member = None
        gameinfo_member = None
        
        for m in members:
            if m.name.endswith('action_camera.json'):
                action_camera_member = m
            elif m.name.endswith('gameinfo.xlsx'):
                gameinfo_member = m
        
        if action_camera_member:
            try:
                f = tar.extractfile(action_camera_member)
                if f:
                    data = json.load(f)
                    result['duration_sec'] = float(data.get('duration_sec', 0))
                    result['frame_count'] = int(data.get('frame_count', 0))
                    result['route_type'] = int(data.get('route_type', 0))
            except (json.JSONDecodeError, ValueError, KeyError):
                pass
        
        # Extract gameinfo.xlsx for scene and operator_id
        if gameinfo_member:
            try:
                import openpyxl
                f = tar.extractfile(gameinfo_member)
                if f:
                    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
                    ws = wb.active
                    if ws:
                        # Try to find scene and operator_id in common locations
                        for row in ws.iter_rows(max_row=20, max_col=10):
                            for cell in row:
                                if cell.value and isinstance(cell.value, str):
                                    cell_lower = cell.value.lower().strip()
                                    if cell_lower == 'scene':
                                        # Next cell might be the value
                                        next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                                        if next_cell.value:
                                            result['scene'] = str(next_cell.value)
                                    elif cell_lower == 'operator_id' or cell_lower == 'operator':
                                        next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                                        if next_cell.value:
                                            result['operator_id'] = str(next_cell.value)
                    wb.close()
            except ImportError:
                # openpyxl not available, return empty scene/operator_id
                pass
            except Exception:
                pass
    
    return result


def list_batch_tarballs(batch_dir: str) -> list[str]:
    """Find *.tar.gz matching pattern vendor-*_batch-*_clip-*_v*.tar.gz, sorted by clip_id."""
    pattern = re.compile(r'^vendor-[\w-]+_batch-[\w-]+_clip-\d+_v\d+\.tar\.gz$')
    tarballs = []
    
    batch_path = Path(batch_dir)
    if not batch_path.exists():
        return []
    
    for f in batch_path.iterdir():
        if f.is_file() and pattern.match(f.name):
            tarballs.append(str(f))
    
    # Sort by clip_id
    tarballs.sort(key=lambda x: parse_clip_id(os.path.basename(x)))
    return tarballs


def parse_clip_id(filename: str) -> str:
    """Extract clip-NNNNN from full filename. Return e.g. 'clip-00042'."""
    match = re.search(r'clip-(\d+)', filename)
    if match:
        return f"clip-{match.group(1)}"
    return ''


def build_manifest(
    batch_dir: str,
    vendor_id: str,
    batch_id: str,
    spec_version: str = "v1",
    notes: str = "",
) -> dict:
    """Build manifest dict matching SUBMISSION_FORMAT.md §4.1 schema."""
    tarballs = list_batch_tarballs(batch_dir)
    
    clips = []
    total_size_bytes = 0
    operator_stats = {}  # operator_id -> {clip_count, routes: {normal, special, loop}}
    
    for i, tarball in enumerate(tarballs):
        if (i + 1) % 10 == 0:
            print(f"Processed {i + 1}/{len(tarballs)} tarballs...")
        
        # Get file size and sha256
        size_bytes = os.path.getsize(tarball)
        total_size_bytes += size_bytes
        sha256 = compute_sha256(tarball)
        
        # Extract metadata
        metadata = extract_clip_metadata(tarball)
        
        clip_entry = {
            'clip_id': metadata['clip_id'],
            'operator': metadata['operator_id'],
            'duration_sec': metadata['duration_sec'],
            'sha256': sha256,
            'size_bytes': size_bytes,
            'route_type': metadata['route_type'],
            'scene': metadata['scene']
        }
        clips.append(clip_entry)
        
        # Aggregate operator stats
        op_id = metadata['operator_id'] or 'unknown'
        if op_id not in operator_stats:
            operator_stats[op_id] = {'clip_count': 0, 'routes': {'normal': 0, 'special': 0, 'loop': 0}}
        
        operator_stats[op_id]['clip_count'] += 1
        
        # Route type: 0=normal, 1=special, 2=loop (assumed mapping)
        route_type = metadata['route_type']
        if route_type == 0:
            operator_stats[op_id]['routes']['normal'] += 1
        elif route_type == 1:
            operator_stats[op_id]['routes']['special'] += 1
        elif route_type == 2:
            operator_stats[op_id]['routes']['loop'] += 1
        else:
            operator_stats[op_id]['routes']['normal'] += 1
    
    # Build operators list
    operators = []
    for op_id, stats in sorted(operator_stats.items()):
        operators.append({
            'operator_id': op_id,
            'clip_count': stats['clip_count'],
            'routes': stats['routes']
        })
    
    total_size_gb = round(total_size_bytes / (1024 ** 3), 2)
    
    manifest = {
        'batch_id': batch_id,
        'vendor_id': vendor_id,
        'spec_version': spec_version,
        'upload_date': datetime.now(timezone.utc).isoformat(),
        'total_clips': len(clips),
        'total_size_gb': total_size_gb,
        'operators': operators,
        'clips': clips,
        'notes': notes,
        'manifest_sha256': ''  # Will be computed after
    }
    
    # Compute manifest_sha256 (excluding the field itself)
    manifest_for_hash = {k: v for k, v in manifest.items() if k != 'manifest_sha256'}
    manifest_json = json.dumps(manifest_for_hash, sort_keys=True, default=str)
    manifest['manifest_sha256'] = hashlib.sha256(manifest_json.encode()).hexdigest()
    
    return manifest


def write_manifest_yaml(manifest: dict, out_path: str) -> None:
    """Lazy import yaml (PyYAML); fallback to manual YAML write if missing."""
    try:
        import yaml
        with open(out_path, 'w') as f:
            yaml.dump(manifest, f, default_flow_style=False, sort_keys=False, allow_unicode=True)
    except ImportError:
        # Manual YAML write
        lines = []
        
        def yaml_value(v: Any) -> str:
            """Convert a Python value to a YAML-compatible string representation.

            Args:
                v: A Python value (str, bool, int, float, None, or other).

            Returns:
                A string representation suitable for YAML output.
            """
            if isinstance(v, str):
                # Quote strings that might need it
                if any(c in v for c in [':', '#', '{', '}', '[', ']', ',', '&', '*', '?', '|', '-', '<', '>', '=', '!', '%', '@', '`']):
                    return f'"{v}"'
                return v
            elif isinstance(v, bool):
                return 'true' if v else 'false'
            elif isinstance(v, (int, float)):
                return str(v)
            elif v is None:
                return 'null'
            else:
                return str(v)
        
        def write_dict(d: dict, indent: int = 0) -> None:
            """Recursively write a dict as YAML key-value pairs.

            Args:
                d: Dictionary to write as YAML.
                indent: Current indentation level (number of 2-space increments).
            """
            for k, v in d.items():
                prefix = '  ' * indent
                if isinstance(v, dict):
                    lines.append(f"{prefix}{k}:")
                    write_dict(v, indent + 1)
                elif isinstance(v, list):
                    lines.append(f"{prefix}{k}:")
                    write_list(v, indent + 1)
                else:
                    lines.append(f"{prefix}{k}: {yaml_value(v)}")
        
        def write_list(lst, indent=0):
            prefix = '  ' * indent
            for item in lst:
                if isinstance(item, dict):
                    lines.append(f"{prefix}-")
                    for k, v in item.items():
                        item_prefix = '  ' * (indent + 1)
                        if isinstance(v, dict):
                            lines.append(f"{item_prefix}{k}:")
                            write_dict(v, indent + 2)
                        elif isinstance(v, list):
                            lines.append(f"{item_prefix}{k}:")
                            write_list(v, indent + 2)
                        else:
                            lines.append(f"{item_prefix}{k}: {yaml_value(v)}")
                else:
                    lines.append(f"{prefix}- {yaml_value(item)}")
        
        write_dict(manifest)
        
        with open(out_path, 'w') as f:
            f.write('\n'.join(lines) + '\n')


def validate_manifest(manifest_path: str) -> tuple[bool, list[str]]:
    """Load manifest.yaml, validate required fields + checksum integrity. Return (ok, errors)."""
    errors = []
    
    # Try to load YAML
    try:
        import yaml
        with open(manifest_path, 'r') as f:
            manifest = yaml.safe_load(f)
    except ImportError:
        # Manual YAML parse (simple key-value)
        manifest = {}
        with open(manifest_path, 'r') as f:
            content = f.read()
        # Simple parsing for basic validation
        current_key = None
        for line in content.split('\n'):
            line = line.rstrip()
            if not line or line.startswith('#'):
                continue
            if ':' in line and not line.startswith(' '):
                parts = line.split(':', 1)
                key = parts[0].strip()
                value = parts[1].strip() if len(parts) > 1 else ''
                manifest[key] = value
    
    if not manifest:
        errors.append("Failed to load manifest or manifest is empty")
        return (False, errors)
    
    # Check required fields
    required_fields = ['batch_id', 'vendor_id', 'spec_version', 'upload_date', 
                       'total_clips', 'total_size_gb', 'operators', 'clips', 'manifest_sha256']
    
    for field in required_fields:
        if field not in manifest:
            errors.append(f"Missing required field: {field}")
    
    # Validate clips
    if 'clips' in manifest and isinstance(manifest['clips'], list):
        clip_required = ['clip_id', 'sha256', 'size_bytes']
        for i, clip in enumerate(manifest['clips']):
            if not isinstance(clip, dict):
                errors.append(f"Clip {i} is not a valid object")
                continue
            for field in clip_required:
                if field not in clip:
                    errors.append(f"Clip {i} missing field: {field}")
    
    # Validate manifest_sha256
    if 'manifest_sha256' in manifest:
        stored_hash = manifest['manifest_sha256']
        # Recompute hash
        manifest_for_hash = {k: v for k, v in manifest.items() if k != 'manifest_sha256'}
        manifest_json = json.dumps(manifest_for_hash, sort_keys=True, default=str)
        computed_hash = hashlib.sha256(manifest_json.encode()).hexdigest()
        
        if stored_hash != computed_hash:
            errors.append(f"Manifest SHA256 mismatch: stored={stored_hash[:16]}..., computed={computed_hash[:16]}...")
    
    return (len(errors) == 0, errors)


def main(argv: list[str] | None = None) -> int:
    """CLI entry point."""
    if argv is None:
        argv = sys.argv[1:]
    
    import argparse
    
    parser = argparse.ArgumentParser(description='Generate or validate vendor batch manifest')
    parser.add_argument('--batch-dir', help='Batch directory containing tarballs')
    parser.add_argument('--vendor-id', help='Vendor ID')
    parser.add_argument('--batch-id', help='Batch ID')
    parser.add_argument('--output', '-o', help='Output manifest path')
    parser.add_argument('--validate', help='Validate existing manifest')
    parser.add_argument('--spec-version', default='v1', help='Spec version (default: v1)')
    parser.add_argument('--notes', default='', help='Notes to include in manifest')
    
    args = parser.parse_args(argv)
    
    if args.validate:
        ok, errors = validate_manifest(args.validate)
        if ok:
            print(f"✓ Manifest {args.validate} is valid")
            return 0
        else:
            print(f"✗ Manifest {args.validate} has errors:")
            for err in errors:
                print(f"  - {err}")
            return 1
    
    if not args.batch_dir or not args.vendor_id or not args.batch_id:
        parser.error("--batch-dir, --vendor-id, and --batch-id are required for manifest generation")
    
    batch_dir = args.batch_dir
    vendor_id = args.vendor_id
    batch_id = args.batch_id
    output_path = args.output or os.path.join(batch_dir, 'manifest.yaml')
    
    print(f"Scanning batch directory: {batch_dir}")
    tarballs = list_batch_tarballs(batch_dir)
    print(f"Found {len(tarballs)} tarballs")
    
    if not tarballs:
        print("No tarballs found matching pattern vendor-*_batch-*_clip-*_v*.tar.gz")
        return 1
    
    print("Building manifest...")
    manifest = build_manifest(batch_dir, vendor_id, batch_id, args.spec_version, args.notes)
    
    print(f"Writing manifest to: {output_path}")
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)
    write_manifest_yaml(manifest, output_path)
    
    print(f"✓ Manifest written: {manifest['total_clips']} clips, {manifest['total_size_gb']} GB")
    return 0


if __name__ == '__main__':
    sys.exit(main())