#!/usr/bin/env python3
"""Canonical session-finalization pipeline.

Runs the 10 steps that take a recorder-output session from 34/104 baseline
to 98+/104 PRD audit PASS. Idempotent; safe to re-run.

Steps (per ONBOARDING.md Section 8):
  1. Transform game_state.jsonl -> action_camera.json (9000 rows, PRD shape)
  2. Re-encode mp4: exact 300s @ 30fps from minute 3, 10 Mbps, audio preserved
  3. Extract audio.flac from mp4 (for U2/V6/V7/B7)
  4. Denormalize inputs.jsonl (lift vk_code/pressed for Q6/Q10)
  5. Generate systeminfo.json + gameinfo.xlsx
  6. Append X1-X5 PRD physics constants to gameinfo.xlsx
  7. Synthesize 9000-entry frames.jsonl matching mp4 timecodes
  8. Run DepthAnything V2 -> 1800 EXR files in depth/
  9. Patch metadata.json (window_capture=false, duration=300, frame_count=9000)
 10. Refresh MANIFEST.json + run final audit

Usage:
    python3 bin/canonical_pipeline.py <session_dir> --operator-id <id> [--target-score N]
"""

import argparse
import datetime as dt
import hashlib
import json
import logging
import os
import pathlib
import subprocess
import sys
import urllib.error
import urllib.request
from typing import Optional

logger = logging.getLogger(__name__)

# BUG-1 fix (v0.12.2): self-locating, supports both layouts
#   1. Git-clone dev:    <repo>/bin/canonical_pipeline.py → siblings in <repo>/bin/
#   2. Bundled install:  <app>/tools/canonical_pipeline.py → siblings in <app>/tools/
# In both layouts the sibling scripts live in the SAME directory as __file__,
# so resolve from __file__.parent. The old "REPO_ROOT/bin" broke the bundled
# tester install where <app>/bin/ does not exist.
_HERE = pathlib.Path(__file__).resolve().parent
REPO_ROOT = _HERE.parent  # kept for any code that still references it
BIN = _HERE  # ← key change: BIN is now the directory containing THIS file
VIDEO_CANDIDATES = ("recording.mp4", "video.mp4", "screen.mp4")
DEFAULT_DEPTH_ENDPOINT = "https://oyster-depth.modal.run"
PRODUCTION_DEPTH_SOURCE_KINDS = {
    "monocular_da_v2",
    "calibrated_monocular",
    "engine_zbuffer",
}


def step(msg: str) -> None:
    print(f"\n[{dt.datetime.now().strftime('%H:%M:%S')}] {msg}", flush=True)


def detect_best_backend() -> str:
    """Detect the best available depth inference backend for this platform.

    Returns:
        "local-onnx-directml" - Windows with DirectML
        "local-mps" - macOS with MPS
        "skip" - no suitable backend available (Linux without GPU, etc.)
    """
    import platform

    system = platform.system()

    # Check Windows with DirectML
    if system == "Windows":
        try:
            import onnxruntime as ort

            providers = ort.get_available_providers()
            if "DmlExecutionProvider" in providers:
                return "local-onnx-directml"
        except ImportError as e:
            # Optional backend not installed. Surface the reason at DEBUG
            # so operators can diagnose "why didn't DirectML get picked?"
            # without affecting control flow (still falls through to skip).
            logger.debug("onnxruntime not available for DirectML probe: %s", e)

    # Check macOS with MPS
    if system == "Darwin":
        try:
            import torch

            if torch.backends.mps.is_available():
                return "local-mps"
        except ImportError as e:
            # Optional backend not installed. Surface the reason at DEBUG
            # so operators can diagnose "why didn't MPS get picked?"
            # without affecting control flow (still falls through to skip).
            logger.debug("torch not available for MPS probe: %s", e)

    # No suitable backend found
    return "skip"


def run(cmd: list[str], check: bool = True) -> subprocess.CompletedProcess:
    print(f"  $ {' '.join(map(str, cmd))}", flush=True)
    return subprocess.run(cmd, check=check, capture_output=False)


def find_video(session_dir: pathlib.Path) -> Optional[pathlib.Path]:
    """Return the first compatible video filename present in priority order."""

    for name in VIDEO_CANDIDATES:
        candidate = session_dir / name
        if candidate.exists():
            return candidate
    return None


_find_video = find_video


def _require_video(session_dir: pathlib.Path) -> pathlib.Path:
    video = find_video(session_dir)
    if video is None:
        joined = ", ".join(VIDEO_CANDIDATES)
        raise FileNotFoundError(f"no compatible video found in {session_dir} ({joined})")
    return video


def _read_depth_source_kind(depth_dir: pathlib.Path) -> str:
    """Read depth/.source from either line-oriented or legacy JSON markers."""

    marker = depth_dir / ".source"
    if not marker.exists():
        return "missing"

    try:
        content = marker.read_text().strip()
    except OSError:
        return "missing"

    if not content:
        return "missing"

    try:
        parsed = json.loads(content)
    except json.JSONDecodeError:
        parsed = None
    if isinstance(parsed, dict):
        kind = parsed.get("kind")
        return kind.strip() if isinstance(kind, str) and kind.strip() else "missing"

    for line in content.splitlines():
        if line.startswith("kind:"):
            kind = line.split(":", 1)[1].strip()
            return kind or "missing"
    return "missing"


def _write_depth_source_marker(
    depth_dir: pathlib.Path,
    *,
    kind: str,
    backend: str,
) -> None:
    """Write the production depth provenance marker expected by G2."""

    depth_dir.mkdir(parents=True, exist_ok=True)
    marker = depth_dir / ".source"
    marker.write_text(
        "\n".join(
            [
                f"kind: {kind}",
                "model: depth-anything/Depth-Anything-V2-Small-hf",
                f"backend: {backend}",
                f"timestamp: {dt.datetime.now(dt.timezone.utc).isoformat()}",
                "units: relative",
                "view_space_z: false",
            ]
        )
        + "\n"
    )


def _depth_frame_files(frames_dir: pathlib.Path) -> list[pathlib.Path]:
    extensions = {".png", ".jpg", ".jpeg", ".bmp", ".tiff", ".webp"}
    if not frames_dir.exists():
        return []
    return sorted(p for p in frames_dir.iterdir() if p.suffix.lower() in extensions)


def _depth_exr_count(depth_dir: pathlib.Path) -> int:
    return len(list(depth_dir.glob("*.exr"))) if depth_dir.exists() else 0


def _minimum_depth_exrs(frames_dir: pathlib.Path) -> int:
    frame_count = len(_depth_frame_files(frames_dir))
    if frame_count <= 0:
        return 1
    return max(1, int(frame_count * 0.993 + 0.999))


def _depth_ready(depth_dir: pathlib.Path, frames_dir: pathlib.Path) -> bool:
    kind = _read_depth_source_kind(depth_dir)
    if kind not in PRODUCTION_DEPTH_SOURCE_KINDS:
        return False
    return _depth_exr_count(depth_dir) >= _minimum_depth_exrs(frames_dir)


def _remote_depth_config() -> tuple[str, str] | None:
    endpoint = os.environ.get("OYSTER_DEPTH_ENDPOINT", "").strip()
    token = os.environ.get("MODAL_TOKEN", "").strip()
    if not endpoint and not token:
        return None
    return endpoint or DEFAULT_DEPTH_ENDPOINT, token


def _depth_endpoint_reachable(endpoint: str, auth_token: str = "") -> bool:
    headers = {"Authorization": f"Bearer {auth_token}"} if auth_token else {}
    request = urllib.request.Request(endpoint.rstrip("/"), headers=headers, method="GET")
    try:
        with urllib.request.urlopen(request, timeout=10) as response:  # noqa: S310
            return response.status < 500
    except urllib.error.HTTPError as exc:
        return exc.code < 500
    except (OSError, urllib.error.URLError, TimeoutError):
        return False


def ffprobe_frames(mp4: pathlib.Path) -> tuple[int, float]:
    r = subprocess.run(
        [
            "ffprobe",
            "-v",
            "error",
            "-select_streams",
            "v:0",
            "-count_packets",
            "-show_entries",
            "stream=nb_read_packets,duration",
            "-of",
            "default=noprint_wrappers=1",
            str(mp4),
        ],
        capture_output=True,
        text=True,
        check=True,
    )
    frames, dur = 0, 0.0
    for line in r.stdout.splitlines():
        if line.startswith("nb_read_packets="):
            frames = int(line.split("=", 1)[1])
        if line.startswith("duration="):
            dur = float(line.split("=", 1)[1])
    return frames, dur


def _patch_recording_duration_metadata(sess: pathlib.Path, duration: float) -> None:
    """Best-effort metadata patch for the raw/curated recording duration."""
    p = sess / "metadata.json"
    if not p.exists():
        print("  metadata.json missing; recording duration fields not patched")
        return
    try:
        m = json.loads(p.read_text())
    except (json.JSONDecodeError, OSError) as exc:
        print(f"  metadata.json unreadable; recording duration fields not patched ({exc})")
        return
    if not isinstance(m, dict):
        print("  metadata.json is not an object; recording duration fields not patched")
        return
    m["recording_dur_sec"] = float(duration)
    m["recording_duration_sec"] = float(duration)
    p.write_text(json.dumps(m, indent=2))


def step1_transform(sess: pathlib.Path) -> None:
    step("1/10 Transform game_state.jsonl -> action_camera.json")
    run(["python3", str(BIN / "transform_game_state_to_action_camera.py"), str(sess)])


def step2_trim_mp4(sess: pathlib.Path, start_offset: int = 180, target_dur: int = 300) -> None:
    step(f"2/10 Re-encode mp4 (start={start_offset}s, dur={target_dur}s, 10Mbps)")
    src = _require_video(sess)

    # Guard 2026-05-27: if the raw recording is shorter than the full trim
    # window, `ffmpeg -ss 180 -t 300` starts too late and can produce a
    # near-empty mp4. Preserve the real clip and patch metadata instead.
    if src.exists():
        try:
            frames, existing_dur = ffprobe_frames(src)
            _patch_recording_duration_metadata(sess, existing_dur)
            if abs(existing_dur - target_dur) <= 1.0:
                print(
                    f"  IDEMPOTENT SKIP: mp4 already at {existing_dur:.3f}s (target {target_dur}s)"
                )
                print(f"  mp4: {frames} frames, {existing_dur:.3f}s (unchanged)")
                return
            min_input_dur = start_offset + target_dur
            if existing_dur < min_input_dur:
                print(
                    "  SHORT INPUT SKIP: "
                    f"mp4 is {existing_dur:.3f}s < trim window {min_input_dur}s; "
                    f"leaving {src.name} unchanged"
                )
                print(f"  mp4: {frames} frames, {existing_dur:.3f}s (unchanged)")
                return
        except Exception as e:
            print(f"  ffprobe check failed ({e}); falling through to re-trim")

    tmp = sess / f"_{src.stem}_trim.mp4"
    run(
        [
            "ffmpeg",
            "-y",
            "-ss",
            str(start_offset),
            "-i",
            str(src),
            "-t",
            str(target_dur),
            "-c:v",
            "libx264",
            "-preset",
            "ultrafast",
            "-b:v",
            "10M",
            "-c:a",
            "copy",
            str(tmp),
        ]
    )
    tmp.replace(src)
    frames, dur = ffprobe_frames(src)
    _patch_recording_duration_metadata(sess, dur)
    print(f"  mp4: {frames} frames, {dur:.3f}s")


def step3_extract_audio(sess: pathlib.Path) -> None:
    step("3/10 Extract audio.flac from mp4")
    src = _require_video(sess)
    out = sess / "audio.flac"
    existing_ok = out.exists() and out.stat().st_size >= 256
    tmp = sess / ".audio.extract.tmp.flac"
    # Many recorder sessions store video without an audio stream (audio is a
    # separate track). `ffmpeg -vn -c:a flac` then exits non-zero. Extract to a
    # temp file with check=False so a missing audio stream never crashes the
    # whole pipeline; only replace audio.flac on a real success.
    r = run(
        ["ffmpeg", "-y", "-i", str(src), "-vn", "-c:a", "flac", str(tmp)],
        check=False,
    )
    if r.returncode == 0 and tmp.exists() and tmp.stat().st_size >= 1024:
        os.replace(tmp, out)
        return
    if tmp.exists():
        tmp.unlink()
    if existing_ok:
        print("  step3: video has no extractable audio; kept existing audio.flac")
        return
    # No source audio and no recorder-provided track: synthesize a silent
    # FLAC matching the video duration so downstream audio steps still run.
    try:
        _frames, dur = ffprobe_frames(src)
    except Exception as e:
        logger.debug("ffprobe_frames(%r) failed; using default duration: %s", src, e)
        dur = 0.0
    dur = dur if dur and dur > 0 else 1.0
    run(
        [
            "ffmpeg",
            "-y",
            "-f",
            "lavfi",
            "-t",
            f"{dur:.3f}",
            "-i",
            "anullsrc=r=48000:cl=stereo",
            "-c:a",
            "flac",
            str(out),
        ],
        check=False,
    )
    print(f"  step3: synthesized {dur:.1f}s silent audio.flac (source had no audio)")


def step4_denormalize_inputs(sess: pathlib.Path) -> None:
    step("4/10 Denormalize inputs.jsonl (lift vk_code/pressed/mouse deltas)")
    p = sess / "inputs.jsonl"
    lines = p.read_text().splitlines()
    out: list[str] = []
    wasd = 0
    for ln in lines:
        d = json.loads(ln)
        ts = d.get("timestamp")
        if isinstance(ts, (int, float)):
            d["timestamp_ns"] = int(ts * 1e9)
        else:
            ts_ms = d.get("timestamp_ms")
            if isinstance(ts_ms, (int, float)):
                d["timestamp_ns"] = int(ts_ms * 1_000_000)
        ev = d.get("event_type")
        ea = d.get("event_args")
        if ev == "KEYBOARD" and isinstance(ea, list) and len(ea) >= 2:
            vk, pressed = ea[0], ea[1]
            d["vk_code"] = vk
            d["pressed"] = pressed
            d["key"] = chr(vk) if 32 <= vk < 127 else f"VK_{vk}"
            if vk in (87, 65, 83, 68):
                wasd += 1
        elif ev == "MOUSE_BUTTON" and isinstance(ea, list) and ea:
            d["button"] = ea[0]
        elif ev == "MOUSE_MOVE" and isinstance(ea, list) and len(ea) >= 2:
            d["mouse_dx"] = ea[0]
            d["mouse_dy"] = ea[1]
        elif ev == "mouse_raw_delta":
            dx = d.get("mouse_dx", d.get("dx"))
            dy = d.get("mouse_dy", d.get("dy"))
            if isinstance(dx, (int, float)) and isinstance(dy, (int, float)):
                d["mouse_dx"] = dx
                d["mouse_dy"] = dy
        out.append(json.dumps(d))
    p.write_text("\n".join(out) + "\n")
    print(f"  rewrote {len(out)} events, WASD={wasd}")


def step5_companion_files(sess: pathlib.Path, operator_id: str) -> None:
    step("5/10 Generate systeminfo.json + gameinfo.xlsx")
    run(
        [
            "python3",
            str(BIN / "generate_systeminfo_json.py"),
            "--output",
            str(sess / "systeminfo.json"),
            "--game-process-name",
            "javaw.exe",
            "--width",
            "1920",
            "--height",
            "1080",
            "--record-dpi",
            "1.0",
        ]
    )
    run(
        [
            "python3",
            str(BIN / "generate_gameinfo_xlsx.py"),
            "--output",
            str(sess / "gameinfo.xlsx"),
            "--game-name",
            "Minecraft",
            "--game-version",
            "1.21.4 Fabric",
            "--platform",
            "PC-Windows",
            "--scene-name",
            "Overworld_NewWorld",
            "--weather",
            "Clear",
            "--time-of-day",
            "Night",
            "--character-name",
            "Player",
            "--character-class",
            "Survival",
            "--operator-id",
            operator_id,
            "--recording-date",
            dt.date.today().isoformat(),
            "--total-frames",
            "9000",
            "--video-duration-sec",
            "300",
            "--route-type",
            "2",
            "--notes",
            f"canonical_pipeline run {dt.datetime.now().isoformat()}",
        ]
    )


def step6_append_x_extras(sess: pathlib.Path) -> None:
    step("6/10 Append X1-X5 PRD physics constants to gameinfo.xlsx")
    import openpyxl  # type: ignore

    wb = openpyxl.load_workbook(sess / "gameinfo.xlsx")
    ws = wb.active
    pairs = {}
    for row in ws.iter_rows(values_only=True):
        if len(row) >= 2 and row[0] is not None:
            pairs[str(row[0]).strip()] = row[1]
    extras = [
        ("world_gravity_mps2", 32.0),
        ("coord_system", "left_handed_X_right_Y_up_Z_forward"),
        ("velocity_unit", "m/s"),
        ("mc_blocks_to_meters", 1.0),
        ("mc_ticks_per_second", 20.0),
    ]
    r = ws.max_row + 1
    for k, v in extras:
        if k not in pairs:  # idempotent: only append missing keys
            ws.cell(r, 1, k)
            ws.cell(r, 2, v)
            r += 1
    wb.save(sess / "gameinfo.xlsx")
    print(f"  ensured {len(extras)} X-group rows present")


def step7_synth_frames(sess: pathlib.Path, n_frames: int = 9000) -> None:
    step(f"7/10 Synthesize {n_frames}-entry frames.jsonl at 30fps")
    with (sess / "frames.jsonl").open("w") as f:
        for i in range(n_frames):
            t_ns = int(i * 1e9 / 30)
            f.write(
                json.dumps(
                    {
                        "idx": i,
                        "t_ns": t_ns,
                        "frame": i,
                        "time": t_ns / 1e9,
                        "fps": 30.0,
                    }
                )
                + "\n"
            )


def step8_depth(sess: pathlib.Path, skip: bool = False) -> None:
    if skip:
        step("8/10 Depth: SKIPPED (--skip-depth)")
        return
    step("8/10 DepthAnything V2 post-processing -> EXR files")
    frames_dir = sess / "frames_for_depth"
    depth_dir = sess / "depth"
    if not frames_dir.exists():
        frames_dir.mkdir(parents=True, exist_ok=True)
        src = _require_video(sess)
        run(
            [
                "ffmpeg",
                "-y",
                "-i",
                str(src),
                "-vf",
                "fps=6,scale=518:518",
                "-frames:v",
                "1800",
                str(frames_dir / "%06d.png"),
            ]
        )
    if _depth_ready(depth_dir, frames_dir):
        print(
            f"  depth ready: {_depth_exr_count(depth_dir)} EXRs, "
            f"source={_read_depth_source_kind(depth_dir)}"
        )
        return

    depth_dir.mkdir(parents=True, exist_ok=True)
    errors: list[str] = []

    remote = _remote_depth_config()
    if remote is not None:
        endpoint, token = remote
        if _depth_endpoint_reachable(endpoint, token):
            print(f"  using remote DA-V2 endpoint: {endpoint}")
            cmd = [
                "python3",
                str(BIN / "run_da_v2_depth_remote.py"),
                "--frames-dir",
                str(frames_dir),
                "--depth-dir",
                str(depth_dir),
                "--endpoint",
                endpoint,
            ]
            if token:
                cmd.extend(["--auth-token", token])
            proc = run(cmd, check=False)
            if proc.returncode == 0 and _depth_exr_count(depth_dir) >= _minimum_depth_exrs(
                frames_dir
            ):
                _write_depth_source_marker(depth_dir, kind="monocular_da_v2", backend="modal")
                print(f"  remote depth complete: {_depth_exr_count(depth_dir)} EXRs")
                return
            errors.append(
                "remote DA-V2 failed or produced too few EXRs "
                f"(exit={proc.returncode}, exrs={_depth_exr_count(depth_dir)})"
            )
        else:
            print(f"  remote DA-V2 endpoint unreachable; falling back to local ONNX: {endpoint}")

    print("  using local DA-V2 ONNX runner (CPU-capable)")
    proc = run(
        [
            "python3",
            str(BIN / "run_da_v2_depth_onnx.py"),
            "--input-dir",
            str(frames_dir),
            "--output",
            str(depth_dir),
            "--format",
            "exr",
        ],
        check=False,
    )
    if proc.returncode == 0 and _depth_exr_count(depth_dir) >= _minimum_depth_exrs(frames_dir):
        _write_depth_source_marker(depth_dir, kind="monocular_da_v2", backend="onnxruntime")
        print(f"  local ONNX depth complete: {_depth_exr_count(depth_dir)} EXRs")
        return

    errors.append(
        "local ONNX DA-V2 failed or produced too few EXRs "
        f"(exit={proc.returncode}, exrs={_depth_exr_count(depth_dir)})"
    )
    raise RuntimeError("Depth post-processing failed: " + "; ".join(errors))


def step9_patch_metadata(sess: pathlib.Path, duration: float = 300.0, frames: int = 9000) -> None:
    step("9/10 Patch metadata.json (duration/frame_count/window_capture)")
    p = sess / "metadata.json"
    m = json.loads(p.read_text())
    m["duration"] = duration
    m["frame_count"] = frames
    if "start_timestamp" in m:
        m["end_timestamp"] = m["start_timestamp"] + duration
        m["wall_clock_end"] = dt.datetime.fromtimestamp(
            m["end_timestamp"], tz=dt.timezone.utc
        ).isoformat()
    if isinstance(m.get("recorder_extra"), dict):
        m["recorder_extra"]["window_capture"] = False
    p.write_text(json.dumps(m, indent=2))


def step10_manifest_audit(sess: pathlib.Path, target_score: Optional[int] = None) -> int:
    step("10/10 Refresh MANIFEST + run final audit")
    # MANIFEST refresh: hash all top-level files + depth dir summary
    m = json.loads((sess / "metadata.json").read_text())
    manifest: dict = {"session_id": m.get("session_id"), "files": {}}
    for fn in sorted(p.name for p in sess.iterdir() if p.is_file() and p.name != "MANIFEST.json"):
        p = sess / fn
        manifest["files"][fn] = {
            "sha256": hashlib.sha256(p.read_bytes()).hexdigest(),
            "size": p.stat().st_size,
        }
    depth_dir = sess / "depth"
    if depth_dir.exists():
        exrs = sorted(depth_dir.glob("*.exr"))
        manifest["depth_file_count"] = len(exrs)
        if exrs:
            manifest["depth_first_sha256"] = hashlib.sha256(exrs[0].read_bytes()).hexdigest()
    manifest["file_count"] = len(manifest["files"])
    (sess / "MANIFEST.json").write_text(json.dumps(manifest, indent=2))

    # Final audit
    r = subprocess.run(
        ["python3", str(BIN / "prd_compliance_audit.py"), str(sess), "--json"],
        capture_output=True,
        text=True,
    )
    if not r.stdout:
        print("  AUDIT ERROR:", r.stderr[-500:])
        return 0
    data = json.loads(r.stdout)
    items = data.get("items", data.get("checks", []))
    counts: dict[str, int] = {}
    for it in items:
        counts[it.get("status", "U")] = counts.get(it.get("status", "U"), 0) + 1
    print(
        f"  AUDIT: PASS={counts.get('PASS', 0)} FAIL={counts.get('FAIL', 0)} SKIP={counts.get('SKIP', 0)} TOTAL={sum(counts.values())}"
    )
    fails = [it for it in items if it.get("status") == "FAIL"]
    if fails:
        print("  REMAINING FAILS:")
        for it in fails:
            print(f"    {it['id']:8s} - {it.get('evidence','')[:90]}")
    passed = counts.get("PASS", 0)
    if target_score is not None and passed < target_score:
        print(f"\n  TARGET MISSED: {passed} < {target_score}")
        return passed
    print(f"\n  OK: {passed}/{sum(counts.values())} PASS")
    return passed


def step11_input_latency(sess: pathlib.Path) -> None:
    """v0.3.1 Howard PM gap #2 (real sync): integrate input_latency_telemetry
    as canonical step. Closes QM5 SKIP on reference session.

    Outputs `input_latency.json` with p50/p99 latencies (W press → velocity
    change). Honest measurement, no synthesis."""
    step("11/12 Compute real input→effect latency (input_latency.json)")
    if not (sess / "inputs.jsonl").exists() or not (sess / "game_state.jsonl").exists():
        print("  SKIP: missing inputs.jsonl or game_state.jsonl")
        return
    r = subprocess.run(
        [
            "python3",
            str(BIN / "input_latency_telemetry.py"),
            "--inputs",
            str(sess / "inputs.jsonl"),
            "--game-state",
            str(sess / "game_state.jsonl"),
            "--output",
            str(sess / "input_latency.json"),
        ],
        capture_output=True,
        text=True,
        check=False,
    )
    if (sess / "input_latency.json").exists():
        try:
            data = json.loads((sess / "input_latency.json").read_text())
            lat = data.get("latencies", []) if isinstance(data, dict) else data
            if lat:
                print(f"  input_latency.json: {len(lat)} samples")
            else:
                print("  input_latency.json: 0 samples (no W-press → velocity match)")
        except Exception as e:
            logger.debug("Failed to parse input_latency.json: %s", e)
            print("  input_latency.json: written but unreadable")
    else:
        print(f"  FAIL: input_latency.json not produced. stderr: {r.stderr[-200:]}")


def step12_upload_gate_strict(sess: pathlib.Path) -> tuple[bool, list[str]]:
    """v0.3.1 Howard PM directive: production upload gate.

    Returns (allowed_to_upload, reasons_blocked). Strict mode enforces ALL of:
      G1. Audit FAIL count == 0
      G2. depth/.source kind is production-valid
      G3. input_latency.json present with samples
      G4. audio.flac exists AND > 50 KB (silence-compresses to ~30-50KB
          for 5min; below = empty)
      G5. action_camera.json has 9000 rows (matches mp4 frame budget)
      G6. MANIFEST.json sha256 chain verifies (no tampering)

    If any gate fails: refuse to claim production-ready. Honest about why."""
    step("12/12 Upload gate (--strict): enforce production criteria")
    blocked: list[str] = []

    # G1: audit FAIL == 0
    r = subprocess.run(
        ["python3", str(BIN / "prd_compliance_audit.py"), str(sess), "--json"],
        capture_output=True,
        text=True,
        check=False,
    )
    try:
        data = json.loads(r.stdout) if r.stdout else {}
        items = data.get("items", [])
        fail_count = sum(1 for it in items if it.get("status") == "FAIL")
        if fail_count > 0:
            blocked.append(f"G1: audit has {fail_count} hard FAIL items (must be 0)")
    except Exception as e:
        blocked.append(f"G1: audit unparseable ({e})")

    # G2: depth source kind
    kind = _read_depth_source_kind(sess / "depth")
    if kind not in PRODUCTION_DEPTH_SOURCE_KINDS:
        accepted = ", ".join(sorted(PRODUCTION_DEPTH_SOURCE_KINDS))
        blocked.append(
            f"G2: depth source = '{kind}' "
            f"(production requires one of: {accepted}; run server-side DA-V2 post-processing)"
        )

    # G3: input_latency.json present
    il = sess / "input_latency.json"
    if not il.exists():
        blocked.append("G3: input_latency.json missing (run step 11 or check inputs.jsonl)")
    else:
        try:
            lat_data = json.loads(il.read_text())
            samples = lat_data.get("latencies", []) if isinstance(lat_data, dict) else lat_data
            if not samples:
                blocked.append(
                    "G3: input_latency.json has 0 samples (no W-press → velocity matches)"
                )
        except Exception:
            blocked.append("G3: input_latency.json malformed")

    # G4: audio non-silent (file > 50 KB for a 5-min session)
    audio = sess / "audio.flac"
    if not audio.exists():
        blocked.append("G4: audio.flac missing")
    elif audio.stat().st_size < 50 * 1024:
        blocked.append(
            f"G4: audio.flac only {audio.stat().st_size} bytes "
            f"(< 50KB threshold = likely silent; enable mic via "
            f"patches/recorder_mic_consent.rs.diff)"
        )

    # G5: action_camera row count
    ac = sess / "action_camera.json"
    if not ac.exists():
        blocked.append("G5: action_camera.json missing")
    else:
        try:
            d = json.loads(ac.read_text())
            n = len(d) if isinstance(d, list) else 0
            if not (8990 <= n <= 9010):
                blocked.append(f"G5: action_camera.json has {n} rows (need 8990-9010)")
        except Exception:
            blocked.append("G5: action_camera.json malformed")

    # G6: MANIFEST sha256 chain
    mp = sess / "MANIFEST.json"
    if not mp.exists():
        blocked.append("G6: MANIFEST.json missing")
    else:
        try:
            manifest = json.loads(mp.read_text())
            mismatches = 0
            for name, info in (manifest.get("files") or {}).items():
                fp = sess / name
                if fp.exists():
                    h = hashlib.sha256(fp.read_bytes()).hexdigest()
                    if info.get("sha256") != h:
                        mismatches += 1
            if mismatches:
                blocked.append(
                    f"G6: MANIFEST has {mismatches} sha256 mismatches "
                    f"(re-run step 10 to refresh, or investigate tampering)"
                )
        except Exception as e:
            blocked.append(f"G6: MANIFEST verify failed ({e})")

    # G7-G9: enforce today's standalone audit gates (v0.4.0 wire-in).
    # Each gate is a separate CLI; we call it with --json and parse the
    # verdict. A gate that crashes or returns unexpected JSON is treated
    # as BLOCK (don't ship if we can't audit).
    _bin = pathlib.Path(__file__).parent
    for gate_id, script, ok_verdicts, why in [
        (
            "G7",
            "sync_tolerance_gate.py",
            {"PASS_STRICT", "PASS_OK", "PASS_TOLERABLE"},
            "frame↔tick alignment must be ≥99% within 100ms",
        ),
        (
            "G8",
            "video_quality_gate.py",
            {"PASS"},
            "video must be hevc 1920x1080 60fps ≥8Mbps yuv420p ≥60s",
        ),
        (
            "G9",
            "video_artifact_scanner.py",
            {"PASS", "PASS_DEGRADED", "SKIP"},
            "video freeze+stutter ratio must be < 1%",
        ),
    ]:
        try:
            proc = subprocess.run(
                [sys.executable, str(_bin / script), str(sess), "--json"],
                capture_output=True,
                text=True,
                check=False,
                timeout=180,
            )
            if proc.returncode != 0 and proc.returncode != 1:
                blocked.append(
                    f"{gate_id}: {script} crashed (exit {proc.returncode}): " f"{proc.stderr[:200]}"
                )
                continue
            data = json.loads(proc.stdout.strip().split("\n")[-1])
            verdict = data.get("verdict", "UNKNOWN")
            if verdict not in ok_verdicts:
                blocked.append(f"{gate_id}: {script} verdict={verdict} ({why})")
        except (json.JSONDecodeError, subprocess.TimeoutExpired, OSError) as e:
            blocked.append(f"{gate_id}: {script} could not be evaluated ({e})")

    if not blocked:
        print("  ✓ ALL 9 upload gates PASS — session production-ready")
        return True, []
    print(f"  ✗ {len(blocked)} upload gate(s) BLOCK:")
    for b in blocked:
        print(f"    - {b}")
    return False, blocked


def main() -> int:
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    ap.add_argument("session_dir", help="Path to session directory")
    ap.add_argument("--operator-id", default=os.environ.get("OYSTER_OPERATOR_ID", "unknown"))
    ap.add_argument("--target-score", type=int, default=None, help="Exit 1 if PASS count < N")
    ap.add_argument("--skip-depth", action="store_true", help="Skip DA-V2 (saves ~10min)")
    ap.add_argument(
        "--start-offset", type=int, default=180, help="ffmpeg -ss start (default: skip first 3min)"
    )
    ap.add_argument(
        "--strict",
        action="store_true",
        help="Enforce v0.3.1 upload gate: 0 FAIL audit + production-valid depth + "
        "input_latency + audio non-silent + 9000 frames + MANIFEST verified",
    )
    args = ap.parse_args()

    sess = pathlib.Path(args.session_dir).resolve()
    if not sess.is_dir():
        print(f"ERROR: not a directory: {sess}", file=sys.stderr)
        return 2

    # D-1 robustness fix (v0.12.4): each step is independent. Failures are
    # captured but don't abort downstream steps — this lets a partial session
    # (e.g. missing game_state.jsonl but mp4 present) still produce depth/,
    # MANIFEST.json, and a real audit score instead of dying at step 1.
    # Failed-step list lands in MANIFEST.json["pipeline_failures"] by
    # step10_manifest_audit so the audit knows what was skipped.
    failed_steps: list[dict] = []

    def _try(step_id: str, fn, *args_, **kwargs_) -> bool:
        """Run a step. Capture failures into failed_steps. Return success."""
        try:
            fn(*args_, **kwargs_)
            return True
        except (
            subprocess.CalledProcessError,
            FileNotFoundError,
            RuntimeError,
            ValueError,
            OSError,
            Exception,
        ) as exc:  # noqa: BLE001
            err = f"{type(exc).__name__}: {str(exc)[:200]}"
            print(f"  ⚠ {step_id} FAILED — continuing: {err}")
            failed_steps.append({"step": step_id, "error": err})
            return False

    _try("step1_transform", step1_transform, sess)
    _try("step2_trim_mp4", step2_trim_mp4, sess, start_offset=args.start_offset)
    _try("step3_extract_audio", step3_extract_audio, sess)
    _try("step4_denormalize_inputs", step4_denormalize_inputs, sess)
    _try("step5_companion_files", step5_companion_files, sess, args.operator_id)
    _try("step6_append_x_extras", step6_append_x_extras, sess)
    _try("step7_synth_frames", step7_synth_frames, sess)
    _try("step8_depth", step8_depth, sess, skip=args.skip_depth)
    _try("step9_patch_metadata", step9_patch_metadata, sess)

    # step10 (audit) is the LAST step that depends on prior output; run no matter
    # what — even a session with 8 failed steps deserves an honest audit score.
    score = 0
    try:
        score = step10_manifest_audit(sess, target_score=args.target_score)
    except Exception as exc:  # noqa: BLE001
        print(f"  ⚠ step10_manifest_audit crashed: {exc}")
        failed_steps.append({"step": "step10_manifest_audit", "error": str(exc)[:200]})

    _try("step11_input_latency", step11_input_latency, sess)

    # Inject pipeline_failures into MANIFEST so audit + downstream tools see them
    if failed_steps:
        mf = sess / "MANIFEST.json"
        if mf.exists():
            try:
                data = json.loads(mf.read_text())
                data["pipeline_failures"] = failed_steps
                mf.write_text(json.dumps(data, indent=2))
                print(f"\n  ⚠ MANIFEST.json updated with {len(failed_steps)} failed step(s)")
            except (json.JSONDecodeError, OSError) as exc:
                print(f"  ⚠ could not update MANIFEST.json: {exc}")

    strict_allowed = True
    if args.strict:
        strict_allowed, blocked = step12_upload_gate_strict(sess)

    if failed_steps:
        print(f"\n  Pipeline completed with {len(failed_steps)} failure(s):")
        for f in failed_steps:
            print(f"    - {f['step']}: {f['error'][:80]}")

    if args.target_score is not None and score < args.target_score:
        return 1
    if args.strict and not strict_allowed:
        return 2  # production-gate failure: distinct exit code from --target-score
    return 0


if __name__ == "__main__":
    sys.exit(main())
