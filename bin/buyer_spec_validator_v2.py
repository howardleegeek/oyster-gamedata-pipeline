#!/usr/bin/env python3
"""
G038 · bin/buyer_spec_validator_v2.py

Strict independent v2 validator for buyer specifications.
Performs cross-file consistency checks beyond lint_buyer_spec:
- Manifest SHA vs actual files
- XLSX semantic validation
- EXR temporal alignment
- Cross-file consistency validation

This is a standalone validator that can be run independently of the main pipeline.
"""

import argparse
import hashlib
import json
import os
import sys
import tempfile
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any, Set
import warnings

# Optional imports with lazy loading
try:
    import yaml
    YAML_AVAILABLE = True
except ImportError:
    YAML_AVAILABLE = False

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import numpy as np
    NUMPY_AVAILABLE = True
except ImportError:
    NUMPY_AVAILABLE = False

try:
    from PIL import Image
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False


class BuyerSpecValidatorV2:
    """Main validator class for buyer specification v2."""
    
    def __init__(self, root_dir: str = "."):
        """
        Initialize validator with root directory.
        
        Args:
            root_dir: Root directory to validate
        """
        self.root_dir = Path(root_dir).resolve()
        self.errors: List[str] = []
        self.warnings: List[str] = []
        
    def validate_all(self) -> bool:
        """
        Run all validation checks.
        
        Returns:
            True if all checks pass, False otherwise
        """
        print(f"Validating buyer specification in: {self.root_dir}")
        
        # Run all validation checks
        checks = [
            self._validate_directory_structure,
            self._validate_manifest_files,
            self._validate_xlsx_semantics,
            self._validate_exr_temporal_alignment,
            self._validate_cross_file_consistency,
        ]
        
        for check in checks:
            try:
                check()
            except Exception as e:
                self.errors.append(f"Check failed with exception: {e}")
                
        # Print results
        self._print_results()
        
        return len(self.errors) == 0
        
    def _validate_directory_structure(self) -> None:
        """Validate basic directory structure."""
        required_dirs = ["docs", "data", "specs"]
        
        for dir_name in required_dirs:
            dir_path = self.root_dir / dir_name
            if not dir_path.exists():
                self.warnings.append(f"Directory '{dir_name}' not found")
            elif not dir_path.is_dir():
                self.errors.append(f"'{dir_name}' exists but is not a directory")
                
    def _validate_manifest_files(self) -> None:
        """
        Validate manifest SHA checksums against actual files.
        
        Checks for manifest files and verifies file integrity.
        """
        manifest_files = list(self.root_dir.glob("**/*manifest*.json")) + \
                        list(self.root_dir.glob("**/*manifest*.yaml")) + \
                        list(self.root_dir.glob("**/*manifest*.yml"))
        
        if not manifest_files:
            self.warnings.append("No manifest files found")
            return
            
        for manifest_path in manifest_files:
            self._validate_single_manifest(manifest_path)
            
    def _validate_single_manifest(self, manifest_path: Path) -> None:
        """Validate a single manifest file."""
        try:
            if manifest_path.suffix in ['.yaml', '.yml']:
                if not YAML_AVAILABLE:
                    self.warnings.append(f"PyYAML not available, skipping YAML manifest: {manifest_path}")
                    return
                with open(manifest_path, 'r') as f:
                    manifest = yaml.safe_load(f)
            else:  # Assume JSON
                with open(manifest_path, 'r') as f:
                    manifest = json.load(f)
        except Exception as e:
            self.errors.append(f"Failed to parse manifest {manifest_path}: {e}")
            return
            
        # Check for file entries with checksums
        if isinstance(manifest, dict):
            self._validate_manifest_entries(manifest, manifest_path)
            
    def _validate_manifest_entries(self, manifest: Dict, manifest_path: Path) -> None:
        """Validate individual entries in a manifest."""
        if 'files' in manifest:
            files = manifest['files']
        elif 'entries' in manifest:
            files = manifest['entries']
        else:
            # Try to find any list of files
            files = []
            for key, value in manifest.items():
                if isinstance(value, list) and value and isinstance(value[0], dict):
                    if 'path' in value[0] or 'file' in value[0]:
                        files = value
                        break
                        
        if not files:
            self.warnings.append(f"No file entries found in manifest: {manifest_path}")
            return
            
        for i, file_entry in enumerate(files):
            if not isinstance(file_entry, dict):
                continue
                
            # Get file path
            file_path = file_entry.get('path') or file_entry.get('file') or file_entry.get('name')
            if not file_path:
                continue
                
            # Resolve relative to manifest location
            abs_path = manifest_path.parent / file_path
            
            # Check if file exists
            if not abs_path.exists():
                self.errors.append(f"File referenced in manifest does not exist: {file_path}")
                continue
                
            # Verify checksum if present
            checksum = file_entry.get('sha256') or file_entry.get('sha1') or file_entry.get('md5')
            if checksum:
                self._verify_checksum(abs_path, checksum, file_entry.get('algorithm', 'sha256'))
                
    def _verify_checksum(self, file_path: Path, expected_hash: str, algorithm: str = 'sha256') -> None:
        """Verify file checksum."""
        try:
            hash_func = getattr(hashlib, algorithm, None)
            if not hash_func:
                self.warnings.append(f"Unsupported hash algorithm: {algorithm}")
                return
                
            with open(file_path, 'rb') as f:
                file_hash = hash_func(f.read()).hexdigest()
                
            if file_hash != expected_hash:
                self.errors.append(f"Checksum mismatch for {file_path}: expected {expected_hash}, got {file_hash}")
        except Exception as e:
            self.errors.append(f"Failed to compute checksum for {file_path}: {e}")
            
    def _validate_xlsx_semantics(self) -> None:
        """Validate XLSX file semantics and structure."""
        if not OPENPYXL_AVAILABLE:
            self.warnings.append("openpyxl not available, skipping XLSX validation")
            return
            
        xlsx_files = list(self.root_dir.glob("**/*.xlsx"))
        
        for xlsx_path in xlsx_files:
            self._validate_single_xlsx(xlsx_path)
            
    def _validate_single_xlsx(self, xlsx_path: Path) -> None:
        """Validate a single XLSX file."""
        try:
            workbook = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
            
            # Check for required sheets
            sheet_names = workbook.sheetnames
            
            # Semantic checks based on common buyer spec patterns
            if 'Metadata' in sheet_names:
                metadata_sheet = workbook['Metadata']
                self._validate_metadata_sheet(metadata_sheet, xlsx_path)
                
            if 'Data' in sheet_names or 'Sheet1' in sheet_names:
                data_sheet = workbook['Data'] if 'Data' in sheet_names else workbook['Sheet1']
                self._validate_data_sheet(data_sheet, xlsx_path)
                
            workbook.close()
            
        except Exception as e:
            self.errors.append(f"Failed to validate XLSX {xlsx_path}: {e}")
            
    def _validate_metadata_sheet(self, sheet, xlsx_path: Path) -> None:
        """Validate metadata sheet in XLSX."""
        # Check for required metadata fields
        required_fields = ['version', 'created_date', 'author', 'description']
        found_fields = set()
        
        for row in sheet.iter_rows(min_row=1, max_row=20, max_col=2, values_only=True):
            if row[0] and isinstance(row[0], str):
                field_name = row[0].lower().strip()
                if any(req in field_name for req in required_fields):
                    found_fields.add(field_name)
                    
        missing = [f for f in required_fields if not any(f in found for found in found_fields)]
        if missing:
            self.warnings.append(f"Missing metadata fields in {xlsx_path}: {missing}")
            
    def _validate_data_sheet(self, sheet, xlsx_path: Path) -> None:
        """Validate data sheet in XLSX."""
        # Check for non-empty data
        has_data = False
        for row in sheet.iter_rows(min_row=2, max_row=100, values_only=True):
            if any(cell is not None for cell in row):
                has_data = True
                break
                
        if not has_data:
            self.warnings.append(f"Data sheet appears empty in {xlsx_path}")
            
    def _validate_exr_temporal_alignment(self) -> None:
        """Validate EXR file temporal alignment and consistency."""
        exr_files = list(self.root_dir.glob("**/*.exr"))
        
        if not exr_files:
            # EXR files are optional
            return
            
        # Group EXR files by sequence
        sequences: Dict[str, List[Path]] = {}
        for exr_path in exr_files:
            # Simple sequence detection - look for frame numbers
            stem = exr_path.stem
            # Try to find pattern like name.####.exr
            import re
            match = re.search(r'(\D+)(\d+)\.exr$', exr_path.name)
            if match:
                seq_name = match.group(1)
                frame_num = int(match.group(2))
                if seq_name not in sequences:
                    sequences[seq_name] = []
                sequences[seq_name].append(exr_path)
                
        # Validate each sequence
        for seq_name, files in sequences.items():
            if len(files) > 1:
                # Check for consistent frame numbering
                frames = []
                for file_path in files:
                    match = re.search(r'(\D+)(\d+)\.exr$', file_path.name)
                    if match:
                        frames.append(int(match.group(2)))
                        
                if frames:
                    frames.sort()
                    # Check for gaps in frame sequence
                    expected = list(range(min(frames), max(frames) + 1))
                    missing = set(expected) - set(frames)
                    if missing:
                        self.warnings.append(f"Missing frames in EXR sequence {seq_name}: {sorted(missing)}")
                        
    def _validate_cross_file_consistency(self) -> None:
        """Validate consistency across different files."""
        # Look for configuration files
        config_files = list(self.root_dir.glob("**/config*.yaml")) + \
                      list(self.root_dir.glob("**/config*.yml")) + \
                      list(self.root_dir.glob("**/config*.json"))
                      
        if len(config_files) > 1:
            # Compare configurations
            configs = []
            for config_path in config_files:
                try:
                    if config_path.suffix in ['.yaml', '.yml']:
                        if YAML_AVAILABLE:
                            with open(config_path, 'r') as f:
                                configs.append(yaml.safe_load(f))
                    else:
                        with open(config_path, 'r') as f:
                            configs.append(json.load(f))
                except Exception:
                    continue
                    
            # Check for inconsistencies
            if len(configs) > 1:
                # Simple comparison of top-level keys
                keys_sets = [set(config.keys()) for config in configs if isinstance(config, dict)]
                if keys_sets:
                    common_keys = set.intersection(*keys_sets)
                    for i, keys in enumerate(keys_sets):
                        extra_keys = keys - common_keys
                        if extra_keys:
                            self.warnings.append(f"Config file {config_files[i]} has extra keys: {extra_keys}")
                            
    def _print_results(self) -> None:
        """Print validation results."""
        print("\n" + "="*60)
        print("VALIDATION RESULTS")
        print("="*60)
        
        if self.warnings:
            print(f"\nWarnings ({len(self.warnings)}):")
            for warning in self.warnings:
                print(f"  ⚠  {warning}")
                
        if self.errors:
            print(f"\nErrors ({len(self.errors)}):")
            for error in self.errors:
                print(f"  ✗ {error}")
        else:
            print("\n✓ No errors found")
            
        if not self.warnings and not self.errors:
            print("\n✓ All checks passed successfully!")
        elif self.errors:
            print(f"\n✗ Validation failed with {len(self.errors)} error(s)")
        else:
            print(f"\n✓ Validation passed with {len(self.warnings)} warning(s)")
            
        print("="*60)
        

def main(argv: Optional[List[str]] = None) -> int:
    """
    Main entry point for the validator.
    
    Args:
        argv: Command line arguments (defaults to sys.argv[1:])
        
    Returns:
        Exit code (0 for success, non-zero for failure)
    """
    parser = argparse.ArgumentParser(
        description="Strict independent v2 validator for buyer specifications",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  %(prog)s                          # Validate current directory
  %(prog)s /path/to/spec            # Validate specific directory
  %(prog)s --strict                 # Treat warnings as errors
  %(prog)s --no-xlsx                # Skip XLSX validation
        """
    )
    
    parser.add_argument(
        "directory",
        nargs="?",
        default=".",
        help="Directory to validate (default: current directory)"
    )
    
    parser.add_argument(
        "--strict",
        action="store_true",
        help="Treat warnings as errors"
    )
    
    parser.add_argument(
        "--no-xlsx",
        action="store_true",
        help="Skip XLSX validation"
    )
    
    parser.add_argument(
        "--no-exr",
        action="store_true",
        help="Skip EXR validation"
    )
    
    parser.add_argument(
        "--verbose",
        "-v",
        action="store_true",
        help="Verbose output"
    )
    
    args = parser.parse_args(argv)
    
    # Create validator
    validator = BuyerSpecValidatorV2(args.directory)
    
    # Run validation
    success = validator.validate_all()
    
    # Determine exit code
    if args.strict and validator.warnings:
        return 1
    elif not success:
        return 1
    else:
        return 0


if __name__ == "__main__":
    sys.exit(main())