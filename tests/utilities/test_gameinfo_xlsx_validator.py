#!/usr/bin/env python3
"""Tests for bin/gameinfo_xlsx_validator.py"""

import os
import sys
import tempfile
import unittest

# Ensure bin/ is importable
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "bin"))

from gameinfo_xlsx_validator import validate_xlsx, REQUIRED_FIELDS, EXPECTED_SHEETS


def _write_xlsx(path: str, sheets: dict) -> None:
    """Helper: write an xlsx file with given sheet data.

    *sheets* maps sheet_name -> list-of-rows (each row is a list of values).
    """
    import openpyxl
    wb = openpyxl.Workbook()
    # Remove default sheet
    default = wb.sheetnames[0]
    del wb[default]

    for sname, rows in sheets.items():
        ws = wb.create_sheet(sname)
        for row in rows:
            ws.append(row)
    wb.save(path)
    wb.close()


class TestValidateXlsx(unittest.TestCase):

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()

    def _path(self, name="gameinfo.xlsx"):
        return os.path.join(self.tmpdir, name)

    # -- happy path ----------------------------------------------------------

    def test_valid_file(self):
        """All 3 sheets present with all required fields."""
        sheets = {}
        for sname, fields in REQUIRED_FIELDS.items():
            sheets[sname] = [fields, ["v1", "v2", "v3", "v4", "v5"]]
        p = self._path()
        _write_xlsx(p, sheets)

        result = validate_xlsx(p)
        self.assertTrue(result["ok"])
        self.assertEqual(sorted(result["sheets_found"]), sorted(EXPECTED_SHEETS))
        self.assertEqual(result["missing_sheets"], [])
        self.assertEqual(result["field_errors"], {})

    # -- missing sheets ------------------------------------------------------

    def test_missing_one_sheet(self):
        """Only metadata and scene_table present; asset_ramp missing."""
        sheets = {}
        for sname in ["metadata", "scene_table"]:
            sheets[sname] = [REQUIRED_FIELDS[sname]]
        p = self._path()
        _write_xlsx(p, sheets)

        result = validate_xlsx(p)
        self.assertFalse(result["ok"])
        self.assertEqual(result["missing_sheets"], ["asset_ramp"])

    def test_missing_all_sheets(self):
        """Workbook has no expected sheets at all."""
        p = self._path()
        _write_xlsx(p, {"unrelated_sheet": [["foo", "bar"]]})

        result = validate_xlsx(p)
        self.assertFalse(result["ok"])
        self.assertEqual(sorted(result["missing_sheets"]), sorted(EXPECTED_SHEETS))

    # -- missing fields ------------------------------------------------------

    def test_missing_fields_in_metadata(self):
        """metadata sheet is missing 'developer' and 'publisher'."""
        sheets = {}
        for sname, fields in REQUIRED_FIELDS.items():
            if sname == "metadata":
                sheets[sname] = [["game_id", "game_name", "version"]]
            else:
                sheets[sname] = [fields]
        p = self._path()
        _write_xlsx(p, sheets)

        result = validate_xlsx(p)
        self.assertFalse(result["ok"])
        self.assertEqual(result["missing_sheets"], [])
        self.assertEqual(result["field_errors"]["metadata"], ["developer", "publisher"])

    def test_missing_fields_in_multiple_sheets(self):
        """Both scene_table and asset_ramp have missing fields."""
        sheets = {}
        for sname, fields in REQUIRED_FIELDS.items():
            if sname == "scene_table":
                sheets[sname] = [["scene_id", "scene_name"]]  # missing 2
            elif sname == "asset_ramp":
                sheets[sname] = [["asset_id"]]  # missing 4
            else:
                sheets[sname] = [fields]
        p = self._path()
        _write_xlsx(p, sheets)

        result = validate_xlsx(p)
        self.assertFalse(result["ok"])
        self.assertEqual(len(result["field_errors"]), 2)
        self.assertEqual(result["field_errors"]["scene_table"], ["scene_type", "background_asset"])
        self.assertEqual(
            result["field_errors"]["asset_ramp"],
            ["asset_name", "asset_type", "priority", "load_order"],
        )

    # -- extra sheets don't hurt ---------------------------------------------

    def test_extra_sheets_ignored(self):
        """Extra sheets beyond the 3 expected should not cause failure."""
        sheets = {}
        for sname, fields in REQUIRED_FIELDS.items():
            sheets[sname] = [fields]
        sheets["bonus_sheet"] = [["a", "b"]]
        p = self._path()
        _write_xlsx(p, sheets)

        result = validate_xlsx(p)
        self.assertTrue(result["ok"])
        self.assertEqual(sorted(result["sheets_found"]), sorted(EXPECTED_SHEETS))


if __name__ == "__main__":
    unittest.main()
