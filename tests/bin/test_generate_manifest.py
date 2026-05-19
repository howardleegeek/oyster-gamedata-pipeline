#!/usr/bin/env python3
"""Tests for bin/generate_manifest.py"""

import hashlib
import json
import os
import sys
import tarfile
import tempfile
from io import BytesIO
from pathlib import Path
from unittest import mock

import pytest

# Add bin to path for imports
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from bin.generate_manifest import (
    build_manifest,
    compute_sha256,
    extract_clip_metadata,
    list_batch_tarballs,
    main,
    parse_clip_id,
    validate_manifest,
    write_manifest_yaml,
)


class TestComputeSha256:
    """Tests for compute_sha256 function."""

    def test_compute_sha256_known_value(self):
        """Write b'hello' to tmp file and verify hex digest."""
        expected = hashlib.sha256(b"hello").hexdigest()

        with tempfile.NamedTemporaryFile(delete=False) as f:
            f.write(b"hello")
            f.flush()
            temp_path = f.name

        try:
            result = compute_sha256(temp_path)
            assert result == expected
            assert result == "2cf24dba5fb0a30e26e83b2ac5b9e29e1b161e5c1fa7425e73043362938b9824"
        finally:
            os.unlink(temp_path)

    def test_compute_sha256_empty_file(self):
        """Test SHA256 of empty file."""
        expected = hashlib.sha256(b"").hexdigest()

        with tempfile.NamedTemporaryFile(delete=False) as f:
            temp_path = f.name

        try:
            result = compute_sha256(temp_path)
            assert result == expected
        finally:
            os.unlink(temp_path)

    def test_compute_sha256_large_file_streaming(self):
        """Test that large files are processed in chunks (streaming)."""
        # Create a 1MB+ file
        data = b"x" * (1024 * 1024 + 1)
        expected = hashlib.sha256(data).hexdigest()

        with tempfile.NamedTemporaryFile(delete=False) as f:
            f.write(data)
            f.flush()
            temp_path = f.name

        try:
            result = compute_sha256(temp_path, chunk_size=65536)
            assert result == expected
        finally:
            os.unlink(temp_path)


class TestParseClipId:
    """Tests for parse_clip_id function."""

    def test_parse_clip_id_extracts_correctly(self):
        """Test various filename variants."""
        test_cases = [
            ("vendor-001_batch-2026-05-A_clip-00042_v1.tar.gz", "clip-00042"),
            ("vendor-abc_batch-xyz_clip-12345_v2.tar.gz", "clip-12345"),
            ("some_prefix_clip-99999_v3.tar.gz", "clip-99999"),
            ("clip-00001_v1.tar.gz", "clip-00001"),
            ("vendor-001_batch-2026-05-A_clip-00042_v1", "clip-00042"),  # No extension
        ]

        for filename, expected in test_cases:
            result = parse_clip_id(filename)
            assert result == expected, f"Failed for {filename}: got {result}, expected {expected}"

    def test_parse_clip_id_no_match(self):
        """Test filename without clip pattern."""
        result = parse_clip_id("vendor-001_batch-2026-05-A_v1.tar.gz")
        assert result == ""


class TestExtractMetadata:
    """Tests for extract_clip_metadata function."""

    def test_extract_metadata_from_real_tarball(self):
        """Create fake tarball with action_camera.json and extract metadata."""
        # Use a proper filename that matches the pattern
        with tempfile.TemporaryDirectory() as tmpdir:
            temp_path = os.path.join(tmpdir, "vendor-001_batch-2026-05-A_clip-00042_v1.tar.gz")

            # Create tarball with action_camera.json
            action_camera_data = {"duration_sec": 120.5, "frame_count": 3600, "route_type": 1}

            with tarfile.open(temp_path, "w:gz") as tar:
                # Add action_camera.json
                json_bytes = json.dumps(action_camera_data).encode("utf-8")
                json_info = tarfile.TarInfo(name="clip-00042/action_camera.json")
                json_info.size = len(json_bytes)
                tar.addfile(json_info, BytesIO(json_bytes))

                # Add some depth EXRs
                for i in range(5):
                    exr_info = tarfile.TarInfo(name=f"clip-00042/depth_{i:04d}.exr")
                    exr_info.size = 0
                    tar.addfile(exr_info, BytesIO(b""))

            result = extract_clip_metadata(temp_path)

            assert result["clip_id"] == "clip-00042"
            assert result["duration_sec"] == 120.5
            assert result["frame_count"] == 3600
            assert result["route_type"] == 1
            assert result["depth_count"] == 5
            assert result["scene"] == ""  # No gameinfo.xlsx
            assert result["operator_id"] == ""  # No gameinfo.xlsx

    def test_extract_metadata_with_gameinfo_xlsx(self):
        """Test extraction with gameinfo.xlsx (requires openpyxl)."""
        pytest.importorskip("openpyxl")
        from openpyxl import Workbook

        with tempfile.TemporaryDirectory() as tmpdir:
            temp_path = os.path.join(tmpdir, "vendor-001_batch-2026-05-A_clip-00100_v1.tar.gz")

            # Create a simple xlsx in memory
            wb = Workbook()
            ws = wb.active
            ws["A1"] = "scene"
            ws["B1"] = "outdoor_forest"
            ws["A2"] = "operator_id"
            ws["B2"] = "operator-001"

            xlsx_buffer = BytesIO()
            wb.save(xlsx_buffer)
            xlsx_data = xlsx_buffer.getvalue()

            action_camera_data = {"duration_sec": 60.0, "frame_count": 1800, "route_type": 0}

            with tarfile.open(temp_path, "w:gz") as tar:
                json_bytes = json.dumps(action_camera_data).encode("utf-8")
                json_info = tarfile.TarInfo(name="clip-00100/action_camera.json")
                json_info.size = len(json_bytes)
                tar.addfile(json_info, BytesIO(json_bytes))

                xlsx_info = tarfile.TarInfo(name="clip-00100/gameinfo.xlsx")
                xlsx_info.size = len(xlsx_data)
                tar.addfile(xlsx_info, BytesIO(xlsx_data))

            result = extract_clip_metadata(temp_path)

            assert result["clip_id"] == "clip-00100"
            assert result["scene"] == "outdoor_forest"
            assert result["operator_id"] == "operator-001"


class TestListBatchTarballs:
    """Tests for list_batch_tarballs function."""

    def test_list_batch_tarballs_filters_pattern(self):
        """Create temp directory with various tar.gz files and test filtering."""
        with tempfile.TemporaryDirectory() as tmpdir:
            # Create matching files
            matching_files = [
                "vendor-001_batch-2026-05-A_clip-00001_v1.tar.gz",
                "vendor-001_batch-2026-05-A_clip-00002_v1.tar.gz",
                "vendor-001_batch-2026-05-A_clip-00003_v1.tar.gz",
            ]

            # Create non-matching files
            non_matching_files = [
                "random_file.tar.gz",
                "vendor-001_batch-2026-05-A.tar.gz",  # Missing clip
                "clip-00001_v1.tar.gz",  # Missing vendor/batch
                "not_a_tarball.txt",
            ]

            for f in matching_files + non_matching_files:
                Path(tmpdir, f).touch()

            result = list_batch_tarballs(tmpdir)

            assert len(result) == 3
            # Check sorted by clip_id
            assert "clip-00001" in result[0]
            assert "clip-00002" in result[1]
            assert "clip-00003" in result[2]

    def test_list_batch_tarballs_empty_dir(self):
        """Test with empty directory."""
        with tempfile.TemporaryDirectory() as tmpdir:
            result = list_batch_tarballs(tmpdir)
            assert result == []

    def test_list_batch_tarballs_nonexistent_dir(self):
        """Test with non-existent directory."""
        result = list_batch_tarballs("/nonexistent/path")
        assert result == []


class TestBuildManifest:
    """Tests for build_manifest function."""

    def test_build_manifest_aggregates_operators(self):
        """Mock 5 clips with different operators and verify aggregation."""
        with tempfile.TemporaryDirectory() as tmpdir:
            # Create 5 fake tarballs
            # route_type values will be: 1%3=1, 2%3=2, 3%3=0, 4%3=1, 5%3=2
            for i in range(1, 6):
                clip_id = f"{i:05d}"
                filename = f"vendor-001_batch-2026-05-A_clip-{clip_id}_v1.tar.gz"
                filepath = os.path.join(tmpdir, filename)

                # Create minimal tarball
                action_camera_data = {
                    "duration_sec": 60.0 + i,
                    "frame_count": 1800 + i * 30,
                    "route_type": i % 3,  # 1, 2, 0, 1, 2
                }

                with tarfile.open(filepath, "w:gz") as tar:
                    json_bytes = json.dumps(action_camera_data).encode("utf-8")
                    json_info = tarfile.TarInfo(name=f"clip-{clip_id}/action_camera.json")
                    json_info.size = len(json_bytes)
                    tar.addfile(json_info, BytesIO(json_bytes))

            # Mock extract_clip_metadata to add operator info
            original_extract = extract_clip_metadata
            operators = ["op-001", "op-001", "op-002", "op-001", "op-002"]

            def mock_extract(path):
                result = original_extract(path)
                idx = int(result["clip_id"].split("-")[1]) - 1
                if idx < len(operators):
                    result["operator_id"] = operators[idx]
                return result

            with mock.patch(
                "bin.generate_manifest.extract_clip_metadata", side_effect=mock_extract
            ):
                manifest = build_manifest(tmpdir, "vendor-001", "vendor-001_batch-2026-05-A")

            assert manifest["total_clips"] == 5
            assert manifest["vendor_id"] == "vendor-001"
            assert manifest["batch_id"] == "vendor-001_batch-2026-05-A"
            assert manifest["spec_version"] == "v1"

            # Check operator aggregation
            operators_dict = {op["operator_id"]: op for op in manifest["operators"]}
            assert operators_dict["op-001"]["clip_count"] == 3
            assert operators_dict["op-002"]["clip_count"] == 2

            # Check route aggregation
            # op-001 has clips 1, 2, 4 with route_types 1, 2, 1 -> special:2, loop:1, normal:0
            assert operators_dict["op-001"]["routes"]["special"] == 2
            assert operators_dict["op-001"]["routes"]["loop"] == 1
            assert operators_dict["op-001"]["routes"]["normal"] == 0

            # op-002 has clips 3, 5 with route_types 0, 2 -> normal:1, loop:1, special:0
            assert operators_dict["op-002"]["routes"]["normal"] == 1
            assert operators_dict["op-002"]["routes"]["loop"] == 1
            assert operators_dict["op-002"]["routes"]["special"] == 0


class TestWriteManifestYaml:
    """Tests for write_manifest_yaml function."""

    def test_write_manifest_yaml_with_pyyaml(self):
        """Test writing YAML with PyYAML available."""
        pytest.importorskip("yaml")

        manifest = {
            "batch_id": "test-batch",
            "vendor_id": "vendor-001",
            "total_clips": 2,
            "clips": [{"clip_id": "clip-00001", "sha256": "abc123", "size_bytes": 1000}],
        }

        with tempfile.NamedTemporaryFile(suffix=".yaml", delete=False) as tmp:
            temp_path = tmp.name

        try:
            write_manifest_yaml(manifest, temp_path)

            with open(temp_path) as f:
                content = f.read()

            assert "batch_id" in content
            assert "test-batch" in content
            assert "vendor_id" in content
            assert "vendor-001" in content
            assert "total_clips" in content
            assert "2" in content

        finally:
            os.unlink(temp_path)

    def test_write_manifest_yaml_no_pyyaml(self):
        """Test writing YAML without PyYAML (manual fallback)."""
        manifest = {
            "batch_id": "test-batch",
            "vendor_id": "vendor-001",
            "total_clips": 2,
            "notes": "some notes",
            "clips": [
                {"clip_id": "clip-00001", "sha256": "abc123", "size_bytes": 1000, "route_type": 0}
            ],
        }

        with tempfile.NamedTemporaryFile(suffix=".yaml", delete=False) as tmp:
            temp_path = tmp.name

        try:
            # Mock the yaml module to raise ImportError
            import builtins

            real_import = builtins.__import__

            def mock_import(name, *args, **kwargs):
                if name == "yaml":
                    raise ImportError("No yaml module")
                return real_import(name, *args, **kwargs)

            with mock.patch("builtins.__import__", side_effect=mock_import):
                write_manifest_yaml(manifest, temp_path)

            with open(temp_path) as f:
                content = f.read()

            # Check that key content is present (format may vary)
            assert "batch_id" in content
            assert "test-batch" in content
            assert "vendor_id" in content
            assert "vendor-001" in content
            assert "total_clips" in content
            assert "clips" in content

        finally:
            os.unlink(temp_path)


class TestValidateManifest:
    """Tests for validate_manifest function."""

    def test_validate_manifest_passes_well_formed(self):
        """Use build_manifest output and validate it."""
        pytest.importorskip("yaml")

        with tempfile.TemporaryDirectory() as tmpdir:
            # Create a minimal tarball
            filepath = os.path.join(tmpdir, "vendor-001_batch-2026-05-A_clip-00001_v1.tar.gz")
            action_camera_data = {"duration_sec": 60.0, "frame_count": 1800, "route_type": 0}

            with tarfile.open(filepath, "w:gz") as tar:
                json_bytes = json.dumps(action_camera_data).encode("utf-8")
                json_info = tarfile.TarInfo(name="clip-00001/action_camera.json")
                json_info.size = len(json_bytes)
                tar.addfile(json_info, BytesIO(json_bytes))

            manifest = build_manifest(tmpdir, "vendor-001", "vendor-001_batch-2026-05-A")
            manifest_path = os.path.join(tmpdir, "manifest.yaml")
            write_manifest_yaml(manifest, manifest_path)

            ok, errors = validate_manifest(manifest_path)

            assert ok, f"Validation failed with errors: {errors}"

    def test_validate_manifest_fails_on_bad_checksum(self):
        """Modify a sha256 in manifest and verify validation fails."""
        pytest.importorskip("yaml")

        with tempfile.TemporaryDirectory() as tmpdir:
            # Create a minimal tarball
            filepath = os.path.join(tmpdir, "vendor-001_batch-2026-05-A_clip-00001_v1.tar.gz")
            action_camera_data = {"duration_sec": 60.0, "frame_count": 1800, "route_type": 0}

            with tarfile.open(filepath, "w:gz") as tar:
                json_bytes = json.dumps(action_camera_data).encode("utf-8")
                json_info = tarfile.TarInfo(name="clip-00001/action_camera.json")
                json_info.size = len(json_bytes)
                tar.addfile(json_info, BytesIO(json_bytes))

            manifest = build_manifest(tmpdir, "vendor-001", "vendor-001_batch-2026-05-A")

            # Corrupt the manifest_sha256
            manifest["manifest_sha256"] = "0" * 64

            manifest_path = os.path.join(tmpdir, "manifest.yaml")
            write_manifest_yaml(manifest, manifest_path)

            ok, errors = validate_manifest(manifest_path)

            assert not ok
            assert any(
                "SHA256 mismatch" in err or "manifest_sha256" in err.lower() for err in errors
            )

    def test_validate_manifest_missing_required_field(self):
        """Test validation fails when required field is missing."""
        with tempfile.NamedTemporaryFile(suffix=".yaml", delete=False, mode="w") as tmp:
            tmp.write("batch_id: test-batch\n")
            tmp.write("vendor_id: vendor-001\n")
            # Missing spec_version, upload_date, etc.
            temp_path = tmp.name

        try:
            ok, errors = validate_manifest(temp_path)
            assert not ok
            assert any("spec_version" in err for err in errors)
        finally:
            os.unlink(temp_path)


class TestMain:
    """Tests for main CLI function."""

    def test_main_generate_manifest(self):
        """Test CLI manifest generation."""
        with tempfile.TemporaryDirectory() as tmpdir:
            # Create a tarball
            filepath = os.path.join(tmpdir, "vendor-001_batch-2026-05-A_clip-00001_v1.tar.gz")
            action_camera_data = {"duration_sec": 60.0, "frame_count": 1800, "route_type": 0}

            with tarfile.open(filepath, "w:gz") as tar:
                json_bytes = json.dumps(action_camera_data).encode("utf-8")
                json_info = tarfile.TarInfo(name="clip-00001/action_camera.json")
                json_info.size = len(json_bytes)
                tar.addfile(json_info, BytesIO(json_bytes))

            output_path = os.path.join(tmpdir, "manifest.yaml")

            argv = [
                "--batch-dir",
                tmpdir,
                "--vendor-id",
                "vendor-001",
                "--batch-id",
                "vendor-001_batch-2026-05-A",
                "--output",
                output_path,
            ]

            result = main(argv)

            assert result == 0
            assert os.path.exists(output_path)

    def test_main_validate_manifest(self):
        """Test CLI validation."""
        pytest.importorskip("yaml")

        with tempfile.TemporaryDirectory() as tmpdir:
            # Create a tarball
            filepath = os.path.join(tmpdir, "vendor-001_batch-2026-05-A_clip-00001_v1.tar.gz")
            action_camera_data = {"duration_sec": 60.0, "frame_count": 1800, "route_type": 0}

            with tarfile.open(filepath, "w:gz") as tar:
                json_bytes = json.dumps(action_camera_data).encode("utf-8")
                json_info = tarfile.TarInfo(name="clip-00001/action_camera.json")
                json_info.size = len(json_bytes)
                tar.addfile(json_info, BytesIO(json_bytes))

            manifest = build_manifest(tmpdir, "vendor-001", "vendor-001_batch-2026-05-A")
            manifest_path = os.path.join(tmpdir, "manifest.yaml")
            write_manifest_yaml(manifest, manifest_path)

            argv = ["--validate", manifest_path]
            result = main(argv)

            assert result == 0

    def test_main_no_tarballs_found(self):
        """Test CLI with empty directory."""
        with tempfile.TemporaryDirectory() as tmpdir:
            argv = [
                "--batch-dir",
                tmpdir,
                "--vendor-id",
                "vendor-001",
                "--batch-id",
                "vendor-001_batch-2026-05-A",
            ]

            result = main(argv)
            assert result == 1  # Error code for no tarballs


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
