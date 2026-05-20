#!/usr/bin/env python3
"""
Tests for bin/sample_tarball_builder.py
"""

import hashlib
import os
import sys
import tarfile
import tempfile
from pathlib import Path
from unittest import mock

import pytest

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from bin.sample_tarball_builder import (
    build_sample_tarball,
    synthesize_action_camera,
    synthesize_depth_dir,
    synthesize_video,
)

# Handle optional imports like the main module does
try:
    from openpyxl import Workbook
except ImportError:
    Workbook = None

try:
    import openpyxl  # noqa: F401
except ImportError:
    openpyxl = None


class TestSynthesizeVideo:
    """Tests for synthesize_video function."""

    def test_synthesize_video_creates_file(self):
        """Test that synthesize_video creates a video file (mock subprocess)."""
        with tempfile.TemporaryDirectory() as tmpdir:
            video_path = os.path.join(tmpdir, "test_video.mp4")

            # Mock subprocess.run to avoid actually calling ffmpeg
            with mock.patch("subprocess.run") as mock_run:
                mock_run.return_value = mock.Mock(returncode=0, stderr="")

                # Create a dummy file to simulate ffmpeg output
                Path(video_path).parent.mkdir(parents=True, exist_ok=True)
                with open(video_path, "wb") as f:
                    f.write(b"\x00" * 1024)  # Dummy video data

                result = synthesize_video(video_path, duration_sec=1, fps=30)

                # Verify ffmpeg was called with correct arguments
                mock_run.assert_called_once()
                call_args = mock_run.call_args[0][0]
                assert "ffmpeg" in call_args[0]
                assert "-c:v" in call_args
                # Sample fix #1: H.265 (libx265), not libx264 — buyer accepts both
                # but H.265 halves bitrate at same quality.
                assert "libx265" in call_args

                # Verify output path
                assert result == video_path
                assert os.path.exists(video_path)

    def test_synthesize_video_ffmpeg_failure(self):
        """Test that synthesize_video raises on ffmpeg failure."""
        with tempfile.TemporaryDirectory() as tmpdir:
            video_path = os.path.join(tmpdir, "test_video.mp4")

            with mock.patch("subprocess.run") as mock_run:
                mock_run.return_value = mock.Mock(
                    returncode=1, stderr="ffmpeg error: invalid codec"
                )

                with pytest.raises(RuntimeError, match="ffmpeg failed"):
                    synthesize_video(video_path)


class TestSynthesizeActionCamera:
    """Tests for synthesize_action_camera function."""

    def test_synthesize_action_camera_returns_valid_records(self):
        """Test synthesize_action_camera produces PDF-compliant JSON records.

        rc4 changed format .bin → .json (PDF p7 file 2). Test now validates
        JSON structure + required fields per actual schema.
        """
        import json

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, "action_camera.json")
            result = synthesize_action_camera(output_path, frame_count=10)

            # Verify file was created
            assert os.path.exists(result)

            # Load and validate JSON structure
            with open(result, "r") as f:
                data = json.load(f)

            # Should be a list of records
            assert isinstance(data, list)
            assert len(data) == 10

            # Each record should have required fields per actual schema
            # (matches synthesize_action_camera output in bin/sample_tarball_builder.py)
            required_fields = [
                "frame",
                "time",
                "fps",
                "route_type",
                "mouse_x",
                "mouse_y",
                "mouse_dx",
                "mouse_dy",
                "keyCode",
                "camera_position",
                "camera_rotation_oula",
                "camera_rotation_quaternion",
                "camera_Follow Offset",
                "camera_intrinsics",
                "camera_speed",
                "player_position",
                "player_rotation_oula",
                "player_rotation_quaternion",
                "player_speed",
                "metric_scale",
            ]
            for record in data:
                for field in required_fields:
                    assert field in record, f"Missing required field: {field}"

                # camera_position should be [x, y, z]
                assert isinstance(record["camera_position"], list)
                assert len(record["camera_position"]) == 3

                # camera_rotation_quaternion should be [x, y, z, w]
                assert isinstance(record["camera_rotation_quaternion"], list)
                assert len(record["camera_rotation_quaternion"]) == 4

    def test_synthesize_action_camera_creates_file(self):
        """Test that synthesize_action_camera creates a valid JSON file."""
        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, "action_camera.json")
            result = synthesize_action_camera(output_path, frame_count=5)

            assert result == output_path
            assert os.path.exists(output_path)

            # Verify it's valid JSON
            import json

            with open(output_path, "r") as f:
                data = json.load(f)
            assert isinstance(data, list)


class TestSynthesizeDepthDir:
    """Tests for synthesize_depth_dir function."""

    def test_synthesize_depth_dir_requires_openexr(self):
        """Test that synthesize_depth_dir raises without OpenEXR."""
        with tempfile.TemporaryDirectory() as tmpdir:
            # Mock OpenEXR import to fail
            with mock.patch.dict(
                sys.modules, {"OpenEXR": None, "Imath": None}
            ):
                with pytest.raises(RuntimeError, match="OpenEXR"):
                    synthesize_depth_dir(tmpdir, count=1)

    def test_synthesize_depth_dir_creates_files(self):
        """Test that synthesize_depth_dir creates EXR files (with OpenEXR available)."""
        # Skip if OpenEXR not installed
        try:
            import OpenEXR  # noqa: F401
            import Imath  # noqa: F401
        except ImportError:
            pytest.skip("OpenEXR not installed")

        with tempfile.TemporaryDirectory() as tmpdir:
            count = synthesize_depth_dir(tmpdir, count=3)

            assert count == 3
            exr_files = list(Path(tmpdir).glob("*.exr"))
            assert len(exr_files) == 3


class TestCreateGameinfoXlsx:
    """Tests for create_gameinfo_xlsx function."""

    def test_create_gameinfo_xlsx_creates_file(self):
        """Test that create_gameinfo_xlsx creates a valid xlsx file."""
        # Skip if openpyxl not installed
        if Workbook is None:
            pytest.skip("openpyxl not installed")

        from bin.sample_tarball_builder import create_gameinfo_xlsx

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, "gameinfo.xlsx")
            result = create_gameinfo_xlsx(output_path)

            assert result == output_path
            assert os.path.exists(output_path)

            # Verify it's a valid xlsx file
            from openpyxl import load_workbook

            wb = load_workbook(output_path)
            assert "gameinfo" in wb.sheetnames

    def test_create_gameinfo_xlsx_without_openpyxl(self):
        """Test that create_gameinfo_xlsx raises without openpyxl."""
        with mock.patch.dict(sys.modules, {"openpyxl": None}):
            from bin.sample_tarball_builder import create_gameinfo_xlsx

            with tempfile.TemporaryDirectory() as tmpdir:
                output_path = os.path.join(tmpdir, "gameinfo.xlsx")
                with pytest.raises(RuntimeError, match="openpyxl"):
                    create_gameinfo_xlsx(output_path)


class TestBuildSampleTarball:
    """Tests for build_sample_tarball function."""

    def test_build_sample_tarball_creates_tarball(self):
        """Test that build_sample_tarball creates a tarball with expected contents."""
        # Skip if openpyxl not installed
        if Workbook is None:
            pytest.skip("openpyxl not installed")

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, "test_sample.tar.gz")

            # Mock synthesize_video to avoid ffmpeg dependency
            with mock.patch("bin.sample_tarball_builder.synthesize_video") as mock_video:
                mock_video.return_value = os.path.join(tmpdir, "video.mp4")
                # Create dummy video file
                Path(mock_video.return_value).parent.mkdir(parents=True, exist_ok=True)
                with open(mock_video.return_value, "wb") as f:
                    f.write(b"\x00" * 1024)

                # Mock synthesize_depth_dir to avoid OpenEXR dependency
                with mock.patch(
                    "bin.sample_tarball_builder.synthesize_depth_dir"
                ) as mock_depth:
                    mock_depth.return_value = 10

                    result = build_sample_tarball(
                        output_path,
                        video_duration_sec=1,
                        frame_count=10,
                        depth_count=10,
                    )

            assert result == output_path
            assert os.path.exists(output_path)

            # Verify tarball contents
            with tarfile.open(output_path, "r:gz") as tar:
                names = tar.getnames()
                assert "video.mp4" in names
                assert "action_camera.json" in names
                assert "gameinfo.xlsx" in names


class TestMain:
    """Tests for main CLI function."""

    def test_main_creates_default_tarball(self):
        """Test that main creates a tarball with default arguments."""
        # Skip if openpyxl not installed
        if Workbook is None:
            pytest.skip("openpyxl not installed")

        from bin.sample_tarball_builder import main

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, "sample.tar.gz")

            # Mock synthesize_video and synthesize_depth_dir
            with mock.patch("bin.sample_tarball_builder.synthesize_video") as mock_video:
                mock_video.return_value = os.path.join(tmpdir, "video.mp4")
                Path(mock_video.return_value).parent.mkdir(parents=True, exist_ok=True)
                with open(mock_video.return_value, "wb") as f:
                    f.write(b"\x00" * 1024)

                with mock.patch(
                    "bin.sample_tarball_builder.synthesize_depth_dir"
                ) as mock_depth:
                    mock_depth.return_value = 10

                    # Mock sys.argv
                    with mock.patch(
                        "sys.argv",
                        ["sample_tarball_builder.py", "-o", output_path],
                    ):
                        main()

            assert os.path.exists(output_path)

    def test_main_with_custom_args(self):
        """Test that main respects custom arguments."""
        # Skip if openpyxl not installed
        if Workbook is None:
            pytest.skip("openpyxl not installed")

        from bin.sample_tarball_builder import main

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, "custom.tar.gz")

            with mock.patch("bin.sample_tarball_builder.synthesize_video") as mock_video:
                mock_video.return_value = os.path.join(tmpdir, "video.mp4")
                Path(mock_video.return_value).parent.mkdir(parents=True, exist_ok=True)
                with open(mock_video.return_value, "wb") as f:
                    f.write(b"\x00" * 1024)

                with mock.patch(
                    "bin.sample_tarball_builder.synthesize_depth_dir"
                ) as mock_depth:
                    mock_depth.return_value = 5

                    with mock.patch(
                        "sys.argv",
                        [
                            "sample_tarball_builder.py",
                            "-o",
                            output_path,
                            "--duration",
                            "60",
                            "--frames",
                            "100",
                            "--depth-count",
                            "5",
                        ],
                    ):
                        main()

            assert os.path.exists(output_path)

            # Verify custom frame count was used
            with tarfile.open(output_path, "r:gz") as tar:
                # Extract and check action_camera.json
                member = tar.getmember("action_camera.json")
                f = tar.extractfile(member)
                import json

                data = json.load(f)
                assert len(data) == 100


class TestActionCameraQuaternionNorm:
    """Tests for quaternion normalization in action camera records."""

    def test_quaternion_is_unit_norm(self):
        """Test that all quaternions have unit norm (within tolerance)."""
        import json
        import math

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, "action_camera.json")
            synthesize_action_camera(output_path, frame_count=100)

            with open(output_path) as f:
                data = json.load(f)

            for record in data:
                q = record["camera_rotation_quaternion"]
                norm = math.sqrt(sum(x * x for x in q))
                assert (
                    abs(norm - 1.0) < 0.001
                ), f"Quaternion {q} has norm {norm}, expected 1.0"


class TestActionCameraWASDDistribution:
    """Tests for WASD key distribution in action camera records."""

    def test_wasd_distribution(self):
        """Test that WASD keys follow expected 40/20/20/20 distribution."""
        import json
        from collections import Counter

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, "action_camera.json")
            synthesize_action_camera(output_path, frame_count=1000)

            with open(output_path) as f:
                data = json.load(f)

            # Count key codes (W=87, A=65, S=83, D=68)
            key_counts = Counter()
            for record in data:
                codes = record["keyCode"]
                for code in codes:
                    key_counts[code] += 1

            # Check distribution (should be roughly 40/20/20/20)
            total = sum(key_counts.values())
            w_ratio = key_counts[87] / total
            a_ratio = key_counts[65] / total
            s_ratio = key_counts[83] / total
            d_ratio = key_counts[68] / total

            # Allow 5% tolerance
            assert 0.35 <= w_ratio <= 0.45, f"W ratio {w_ratio} not in expected range"
            assert 0.15 <= a_ratio <= 0.25, f"A ratio {a_ratio} not in expected range"
            assert 0.15 <= s_ratio <= 0.25, f"S ratio {s_ratio} not in expected range"
            assert 0.15 <= d_ratio <= 0.25, f"D ratio {d_ratio} not in expected range"


class TestActionCameraIntrinsics:
    """Tests for camera intrinsics in action camera records."""

    def test_intrinsics_structure(self):
        """Test that camera_intrinsics has required fields."""
        import json

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, "action_camera.json")
            synthesize_action_camera(output_path, frame_count=10)

            with open(output_path) as f:
                data = json.load(f)

            for record in data:
                intrinsics = record["camera_intrinsics"]
                assert "fx" in intrinsics, "Missing fx in camera_intrinsics"
                assert "fy" in intrinsics, "Missing fy in camera_intrinsics"
                assert "cx" in intrinsics, "Missing cx in camera_intrinsics"
                assert "cy" in intrinsics, "Missing cy in camera_intrinsics"

                # fx should equal fy (square pixels)
                assert intrinsics["fx"] == intrinsics["fy"]

                # Values should be positive
                assert intrinsics["fx"] > 0
                assert intrinsics["fy"] > 0
                assert intrinsics["cx"] > 0
                assert intrinsics["cy"] > 0


class TestActionCameraMouseCoordinates:
    """Tests for mouse coordinates in action camera records."""

    def test_mouse_coordinates_in_range(self):
        """Test that mouse_x/y are in [0,1] and mouse_dx/dy in [-1,1]."""
        import json

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, "action_camera.json")
            synthesize_action_camera(output_path, frame_count=100)

            with open(output_path) as f:
                data = json.load(f)

            for record in data:
                # mouse_x/y should be list of values in [0, 1]
                for val in record["mouse_x"]:
                    assert 0 <= val <= 1, f"mouse_x {val} not in [0,1]"
                for val in record["mouse_y"]:
                    assert 0 <= val <= 1, f"mouse_y {val} not in [0,1]"

                # mouse_dx/dy should be list of values in [-1, 1]
                for val in record["mouse_dx"]:
                    assert -1 <= val <= 1, f"mouse_dx {val} not in [-1,1]"
                for val in record["mouse_dy"]:
                    assert -1 <= val <= 1, f"mouse_dy {val} not in [-1,1]"


class TestActionCameraSpeedConsistency:
    """Tests for speed/position consistency in action camera records."""

    def test_speed_matches_position_delta(self):
        """Test that camera_speed equals Δposition × fps."""
        import json
        import math

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, "action_camera.json")
            synthesize_action_camera(output_path, frame_count=100)

            with open(output_path) as f:
                data = json.load(f)

            fps = 30.0
            tolerance = 0.01  # Allow small floating point errors

            for i, record in enumerate(data[:-1]):  # Skip last record
                pos = record["camera_position"]
                speed = record["camera_speed"]

                # Speed should be approximately (next_pos - pos) * fps
                next_record = data[i + 1]
                next_pos = next_record["camera_position"]

                expected_speed = [
                    (next_pos[0] - pos[0]) * fps,
                    (next_pos[1] - pos[1]) * fps,
                    (next_pos[2] - pos[2]) * fps,
                ]

                for j in range(3):
                    diff = abs(speed[j] - expected_speed[j])
                    assert (
                        diff < tolerance
                    ), f"Frame {i}: speed[{j}]={speed[j]} != expected {expected_speed[j]}"


class TestActionCameraFrameContinuity:
    """Tests for frame continuity in action camera records."""

    def test_frames_are_sequential(self):
        """Test that frame numbers are sequential starting from 0."""
        import json

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, "action_camera.json")
            synthesize_action_camera(output_path, frame_count=50)

            with open(output_path) as f:
                data = json.load(f)

            for i, record in enumerate(data):
                assert record["frame"] == i, f"Frame {record['frame']} != expected {i}"

    def test_time_strings_are_valid(self):
        """Test that time strings are valid timestamps."""
        import json
        from datetime import datetime

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = os.path.join(tmpdir, "action_camera.json")
            synthesize_action_camera(output_path, frame_count=10)

            with open(output_path) as f:
                data = json.load(f)

            for record in data:
                time_str = record["time"]
                # Should parse as datetime
                try:
                    datetime.strptime(time_str, "%Y-%m-%d %H:%M:%S.%f")
                except ValueError:
                    pytest.fail(f"Invalid time string: {time_str}")