#!/usr/bin/env python3
"""Tests for bin/gameinfo_xlsx_validator.py — validates gameinfo.xlsx against buyer-spec."""

from __future__ import annotations

# Import the module under test
import importlib.util
import sys
import tempfile
from pathlib import Path
from unittest.mock import patch

import pytest

spec = importlib.util.spec_from_file_location(
    "gameinfo_xlsx_validator",
    Path(__file__).parent.parent.parent / "bin" / "gameinfo_xlsx_validator.py",
)
validator_module = importlib.util.module_from_spec(spec)
spec.loader.exec_module(validator_module)

validate_xlsx = validator_module.validate_xlsx
main = validator_module.main
REQUIRED_FIELDS = validator_module.REQUIRED_FIELDS
EXPECTED_SHEETS = validator_module.EXPECTED_SHEETS


class TestValidateXlsx:
    """Tests for validate_xlsx() function."""

    def test_valid_xlsx_all_sheets_and_fields(self):
        """Test that a valid xlsx with all sheets and fields passes."""
        # Create a minimal valid xlsx using openpyxl
        import openpyxl

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)

            # Create each required sheet with required fields
            for sheet_name in ["metadata", "scene_table", "asset_ramp"]:
                ws = wb.create_sheet(sheet_name)
                headers = REQUIRED_FIELDS[sheet_name]
                ws.append(headers)
                # Add one data row
                ws.append(["test"] * len(headers))

            wb.save(f.name)

        try:
            result = validate_xlsx(f.name)
            assert result["ok"] is True
            assert result["missing_sheets"] == []
            assert result["field_errors"] == {}
            assert set(result["sheets_found"]) == EXPECTED_SHEETS
        finally:
            Path(f.name).unlink()

    def test_missing_sheets(self):
        """Test detection of missing sheets."""
        import openpyxl

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            # Only add metadata sheet
            ws = wb.create_sheet("metadata")
            ws.append(REQUIRED_FIELDS["metadata"])
            wb.save(f.name)

        try:
            result = validate_xlsx(f.name)
            assert result["ok"] is False
            assert "scene_table" in result["missing_sheets"]
            assert "asset_ramp" in result["missing_sheets"]
        finally:
            Path(f.name).unlink()

    def test_missing_fields_in_sheet(self):
        """Test detection of missing required fields in a sheet."""
        import openpyxl

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)

            # Create metadata sheet with only SOME required fields
            ws = wb.create_sheet("metadata")
            ws.append(["game_id", "game_name"])  # Missing: version, developer, publisher
            ws.append(["test", "test"])

            # Add other sheets to avoid "missing sheets" error
            for sheet_name in ["scene_table", "asset_ramp"]:
                ws = wb.create_sheet(sheet_name)
                ws.append(REQUIRED_FIELDS[sheet_name])

            wb.save(f.name)

        try:
            result = validate_xlsx(f.name)
            assert result["ok"] is False
            assert "metadata" in result["field_errors"]
            missing = result["field_errors"]["metadata"]
            assert "version" in missing
            assert "developer" in missing
            assert "publisher" in missing
        finally:
            Path(f.name).unlink()

    def test_empty_workbook(self):
        """Test that workbook with only default sheet fails with missing sheets."""
        import openpyxl

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb = openpyxl.Workbook()
            # Keep default "Sheet" - this is not a required sheet
            # We need to add at least one sheet to save
            wb.create_sheet("metadata")
            wb.save(f.name)

        try:
            result = validate_xlsx(f.name)
            assert result["ok"] is False
            assert "scene_table" in result["missing_sheets"]
            assert "asset_ramp" in result["missing_sheets"]
        finally:
            Path(f.name).unlink()

    def test_extra_sheets_allowed(self):
        """Test that extra sheets beyond required ones are allowed (not reported in sheets_found)."""
        import openpyxl

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)

            # Create required sheets
            for sheet_name in ["metadata", "scene_table", "asset_ramp"]:
                ws = wb.create_sheet(sheet_name)
                ws.append(REQUIRED_FIELDS[sheet_name])

            # Add extra sheet (not required)
            wb.create_sheet("custom_sheet")

            wb.save(f.name)

        try:
            result = validate_xlsx(f.name)
            assert result["ok"] is True
            # Extra sheets are ignored (sheets_found only contains required sheets)
            assert set(result["sheets_found"]) == EXPECTED_SHEETS
        finally:
            Path(f.name).unlink()


class TestMain:
    """Tests for main() CLI function."""

    def test_main_valid_xlsx(self, capsys):
        """Test main() with valid xlsx returns 0."""
        import openpyxl

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)

            for sheet_name in ["metadata", "scene_table", "asset_ramp"]:
                ws = wb.create_sheet(sheet_name)
                ws.append(REQUIRED_FIELDS[sheet_name])
                ws.append(["test"] * len(REQUIRED_FIELDS[sheet_name]))

            wb.save(f.name)
            xlsx_path = f.name

        try:
            with patch.object(sys, "argv", ["gameinfo_xlsx_validator", "--xlsx", xlsx_path]):
                result = main()
                assert result is None  # main() returns None
        finally:
            Path(xlsx_path).unlink()

    def test_main_missing_xlsx_argument(self, capsys):
        """Test main() with missing --xlsx argument exits with error."""
        with patch.object(sys, "argv", ["gameinfo_xlsx_validator"]):
            with pytest.raises(SystemExit) as exc_info:
                main()
            assert exc_info.value.code == 2  # argparse error exit code

    def test_main_invalid_xlsx_path(self, capsys):
        """Test main() with non-existent file raises FileNotFoundError."""
        with patch.object(sys, "argv", ["gameinfo_xlsx_validator", "--xlsx", "/nonexistent/path.xlsx"]):
            with pytest.raises(FileNotFoundError):
                main()

    def test_main_missing_sheets_exits_1(self, capsys):
        """Test main() with missing sheets exits with code 1."""
        import openpyxl

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            # Only create one sheet, missing others
            ws = wb.create_sheet("metadata")
            ws.append(REQUIRED_FIELDS["metadata"])
            wb.save(f.name)
            xlsx_path = f.name

        try:
            with patch.object(sys, "argv", ["gameinfo_xlsx_validator", "--xlsx", xlsx_path]):
                with pytest.raises(SystemExit) as exc_info:
                    main()
                assert exc_info.value.code == 1
        finally:
            Path(xlsx_path).unlink()

    def test_main_missing_fields_exits_1(self, capsys):
        """Test main() with missing fields exits with code 1."""
        import openpyxl

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)

            # Create all sheets but with missing fields
            ws = wb.create_sheet("metadata")
            ws.append(["game_id"])  # Missing other required fields

            ws = wb.create_sheet("scene_table")
            ws.append(REQUIRED_FIELDS["scene_table"])

            ws = wb.create_sheet("asset_ramp")
            ws.append(REQUIRED_FIELDS["asset_ramp"])

            wb.save(f.name)
            xlsx_path = f.name

        try:
            with patch.object(sys, "argv", ["gameinfo_xlsx_validator", "--xlsx", xlsx_path]):
                with pytest.raises(SystemExit) as exc_info:
                    main()
                assert exc_info.value.code == 1
        finally:
            Path(xlsx_path).unlink()
