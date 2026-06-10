#!/usr/bin/env python3
"""
Build a minimal synthetic session for CI smoke testing.

Creates the smallest possible files that satisfy the audit shape:
- recording.mp4 (1s black, made via ffmpeg)
- action_camera.json (1 row)
- game_state.jsonl (20 ticks)
- frames.jsonl (1 entry)
- inputs.jsonl (1 KEYBOARD event)
- metadata.json (full keys)
- gameinfo.xlsx (14 fields + X1-X5)
- systeminfo.json
- audio_check.json
- MANIFEST.json
- depth/.source marker (kind: ci_fixture)

Audit on this will mostly FAIL on content checks but TOTAL should be 105.
"""

import json
import subprocess
import sys
from pathlib import Path

# Try to import openpyxl, but provide fallback
try:
    from openpyxl import Workbook

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def build_minimal_session(output_dir: str):
    """Build a minimal session fixture for CI testing."""
    base = Path(output_dir)
    base.mkdir(parents=True, exist_ok=True)

    # 1. recording.mp4 - 1 second black video via ffmpeg
    recording_path = base / "recording.mp4"
    subprocess.run(
        [
            "ffmpeg",
            "-f",
            "lavfi",
            "-i",
            "color=c=black:s=640x480:d=1",
            "-c:v",
            "libx264",
            "-pix_fmt",
            "yuv420p",
            "-y",
            str(recording_path),
        ],
        check=True,
        capture_output=True,
    )

    # 2. action_camera.json - 1 row
    action_camera = [
        {
            "timestamp": "2024-01-01T00:00:00.000Z",
            "action": "START_RECORDING",
            "camera_id": "cam_001",
        }
    ]
    with open(base / "action_camera.json", "w") as f:
        json.dump(action_camera, f, indent=2)

    # 3. game_state.jsonl - 20 ticks
    game_state_lines = []
    for i in range(20):
        state = {
            "tick": i,
            "timestamp": f"2024-01-01T00:00:{i:02d}.000Z",
            "player_position": {"x": 0.0, "y": 0.0, "z": 0.0},
            "player_health": 100,
            "game_phase": "PLAYING",
        }
        game_state_lines.append(json.dumps(state))
    with open(base / "game_state.jsonl", "w") as f:
        f.write("\n".join(game_state_lines) + "\n")

    # 4. frames.jsonl - 1 entry
    frames = [
        {
            "frame_id": 0,
            "timestamp": "2024-01-01T00:00:00.000Z",
            "width": 640,
            "height": 480,
            "format": "RGB24",
        }
    ]
    with open(base / "frames.jsonl", "w") as f:
        for frame in frames:
            f.write(json.dumps(frame) + "\n")

    # 5. inputs.jsonl - 1 KEYBOARD event
    inputs = [
        {
            "timestamp": "2024-01-01T00:00:00.500Z",
            "device": "KEYBOARD",
            "action": "PRESS",
            "key": "SPACE",
            "frame_id": 0,
        }
    ]
    with open(base / "inputs.jsonl", "w") as f:
        for inp in inputs:
            f.write(json.dumps(inp) + "\n")

    # 6. metadata.json - full keys
    metadata = {
        "session_id": "ci-test-session-001",
        "game_name": "TestGame",
        "game_version": "1.0.0",
        "recorder_version": "1.0.0",
        "start_time": "2024-01-01T00:00:00.000Z",
        "end_time": "2024-01-01T00:00:01.000Z",
        "duration_seconds": 1.0,
        "platform": "Windows",
        "player_id": "test-player-001",
        "resolution": "640x480",
        "fps": 30,
        "audio_enabled": True,
        "video_enabled": True,
        "input_capture_enabled": True,
    }
    with open(base / "metadata.json", "w") as f:
        json.dump(metadata, f, indent=2)

    # 7. gameinfo.xlsx - 14 fields + X1-X5
    if HAS_OPENPYXL:
        wb = Workbook()
        ws = wb.active
        ws.title = "GameInfo"

        # Headers
        headers = [
            "game_name",
            "game_version",
            "session_id",
            "player_id",
            "start_time",
            "end_time",
            "duration",
            "platform",
            "resolution",
            "fps",
            "audio",
            "video",
            "input",
            "notes",
            "X1",
            "X2",
            "X3",
            "X4",
            "X5",
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Data row
        values = [
            "TestGame",
            "1.0.0",
            "ci-test-session-001",
            "test-player-001",
            "2024-01-01T00:00:00.000Z",
            "2024-01-01T00:00:01.000Z",
            "1.0",
            "Windows",
            "640x480",
            "30",
            "true",
            "true",
            "true",
            "CI test session",
            "",
            "",
            "",
            "",
            "",
        ]
        for col, value in enumerate(values, 1):
            ws.cell(row=2, column=col, value=value)

        wb.save(base / "gameinfo.xlsx")
    else:
        # Create a minimal xlsx file without openpyxl
        # This is a fallback - create a simple zip-based xlsx
        import zipfile

        xlsx_path = base / "gameinfo.xlsx"
        with zipfile.ZipFile(xlsx_path, "w") as zf:
            # Minimal content types
            zf.writestr(
                "[Content_Types].xml",
                """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
<Default Extension="xml" ContentType="application/xml"/>
<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
</Types>""",
            )
            # Minimal workbook
            zf.writestr(
                "xl/workbook.xml",
                """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
<sheets><sheet name="GameInfo" sheetId="1" r:id="rId1"/></sheets>
</workbook>""",
            )
            # Minimal worksheet
            zf.writestr(
                "xl/worksheets/sheet1.xml",
                """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><sheetData/></worksheet>""",
            )
            # Relationships
            zf.writestr(
                "xl/_rels/workbook.xml.rels",
                """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
</Relationships>""",
            )
            zf.writestr(
                "_rels/.rels",
                """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
</Relationships>""",
            )

    # 8. systeminfo.json
    systeminfo = {
        "os_name": "Windows 10",
        "os_version": "10.0.19045",
        "cpu": "Intel Core i7-9700K",
        "gpu": "NVIDIA GeForce RTX 2080",
        "ram_gb": 16,
        "disk_space_gb": 500,
        "python_version": "3.11.0",
        "ffmpeg_version": "5.1.2",
    }
    with open(base / "systeminfo.json", "w") as f:
        json.dump(systeminfo, f, indent=2)

    # 9. audio_check.json
    audio_check = {
        "audio_present": True,
        "audio_format": "AAC",
        "sample_rate": 44100,
        "channels": 2,
        "duration_seconds": 1.0,
        "check_passed": True,
    }
    with open(base / "audio_check.json", "w") as f:
        json.dump(audio_check, f, indent=2)

    # 10. MANIFEST.json
    manifest = {
        "version": "1.0",
        "created": "2024-01-01T00:00:00.000Z",
        "files": [
            {"name": "recording.mp4", "type": "video", "size": 10000},
            {"name": "action_camera.json", "type": "metadata", "size": 200},
            {"name": "game_state.jsonl", "type": "telemetry", "size": 2000},
            {"name": "frames.jsonl", "type": "metadata", "size": 150},
            {"name": "inputs.jsonl", "type": "input", "size": 100},
            {"name": "metadata.json", "type": "metadata", "size": 500},
            {"name": "gameinfo.xlsx", "type": "metadata", "size": 5000},
            {"name": "systeminfo.json", "type": "metadata", "size": 300},
            {"name": "audio_check.json", "type": "metadata", "size": 150},
        ],
        "total_files": 9,
        "checksum_algorithm": "sha256",
    }
    with open(base / "MANIFEST.json", "w") as f:
        json.dump(manifest, f, indent=2)

    # 11. depth/.source marker
    depth_dir = base / "depth"
    depth_dir.mkdir(exist_ok=True)
    source_marker = {
        "kind": "ci_fixture",
        "created": "2024-01-01T00:00:00.000Z",
        "generator": "build_minimal_session.py",
        "purpose": "CI smoke testing",
    }
    with open(depth_dir / ".source", "w") as f:
        json.dump(source_marker, f, indent=2)

    print(f"Built minimal session at: {base}")
    print("Files created:")
    for item in sorted(base.rglob("*")):
        if item.is_file():
            print(f"  {item.relative_to(base)}: {item.stat().st_size} bytes")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print(f"Usage: {sys.argv[0]} <output_dir>", file=sys.stderr)
        sys.exit(1)

    output_dir = sys.argv[1]
    build_minimal_session(output_dir)
