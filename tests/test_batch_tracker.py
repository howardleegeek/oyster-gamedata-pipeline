#!/usr/bin/env python3
"""
Test fixture for batch_tracker.py
"""
import json
import tempfile
import shutil
from pathlib import Path
from unittest.mock import patch, MagicMock
import pytest
import sys
import os

# Add bin to path
sys.path.insert(0, str(Path(__file__).parent.parent / "bin"))

from batch_tracker import (
    find_session_dirs,
    read_route_type_from_gameinfo,
    read_route_type_from_action_camera,
    get_route_type,
    compute_distribution,
    check_distribution,
    TARGET_DISTRIBUTION,
    TOLERANCE_PERCENT,
)


def test_find_session_dirs(tmp_path):
    """Test finding session directories."""
    # Create a recordings root directory
    recordings_root = tmp_path / "recordings"
    recordings_root.mkdir()
    
    # Create some session directories with gameinfo.xlsx
    session1 = recordings_root / "session_20240101_120000"
    session1.mkdir()
    (session1 / "gameinfo.xlsx").write_text("dummy")
    
    session2 = recordings_root / "session_20240101_130000"
    session2.mkdir()
    (session2 / "action_camera.json").write_text("{}")
    
    # Create a non-session directory (no gameinfo or action_camera)
    non_session = recordings_root / "not_a_session"
    non_session.mkdir()
    
    # Create a nested session
    nested_parent = recordings_root / "parent_dir"
    nested_parent.mkdir()
    nested_session = nested_parent / "nested_session"
    nested_session.mkdir()
    (nested_session / "gameinfo.xlsx").write_text("dummy")
    
    # Test finding sessions
    sessions = find_session_dirs(recordings_root)
    session_names = {s.name for s in sessions}
    
    assert "session_20240101_120000" in session_names
    assert "session_20240101_130000" in session_names
    assert "nested_session" in session_names
    assert "not_a_session" not in session_names
    assert "parent_dir" not in session_names  # parent dir doesn't have gameinfo/action_camera


def test_find_session_dirs_nonexistent():
    """Test finding session directories when root doesn't exist."""
    non_existent = Path("/non/existent/path")
    sessions = find_session_dirs(non_existent)
    assert sessions == []


def test_read_route_type_from_action_camera(tmp_path):
    """Test reading route_type from action_camera.json."""
    session_dir = tmp_path / "test_session"
    session_dir.mkdir()
    
    # Test with route_type in first frame
    action_camera_data = [
        {"route_type": 1, "timestamp": 0},
        {"route_type": 1, "timestamp": 1},
    ]
    action_camera_path = session_dir / "action_camera.json"
    with open(action_camera_path, "w") as f:
        json.dump(action_camera_data, f)
    
    assert read_route_type_from_action_camera(session_dir) == 1
    
    # Test with route_type at top level
    action_camera_data = {"route_type": 2, "frames": []}
    with open(action_camera_path, "w") as f:
        json.dump(action_camera_data, f)
    
    assert read_route_type_from_action_camera(session_dir) == 2
    
    # Test with no route_type
    action_camera_data = {"other_field": "value"}
    with open(action_camera_path, "w") as f:
        json.dump(action_camera_data, f)
    
    assert read_route_type_from_action_camera(session_dir) is None
    
    # Test with invalid JSON
    action_camera_path.write_text("{invalid json")
    assert read_route_type_from_action_camera(session_dir) is None
    
    # Test with non-existent file
    non_existent_dir = tmp_path / "non_existent"
    non_existent_dir.mkdir()
    assert read_route_type_from_action_camera(non_existent_dir) is None


@patch('batch_tracker.openpyxl')
def test_read_route_type_from_gameinfo(mock_openpyxl, tmp_path):
    """Test reading route_type from gameinfo.xlsx."""
    session_dir = tmp_path / "test_session"
    session_dir.mkdir()
    gameinfo_path = session_dir / "gameinfo.xlsx"
    gameinfo_path.touch()
    
    # Mock workbook and worksheet
    mock_ws = MagicMock()
    mock_wb = MagicMock()
    mock_wb.active = mock_ws
    mock_openpyxl.load_workbook.return_value = mock_wb
    
    # Mock headers with route_type in column 3
    mock_ws[1] = [MagicMock(value="col1"), MagicMock(value="col2"), MagicMock(value="route_type")]
    
    # Mock data row with route_type = 3
    mock_cell = MagicMock()
    mock_cell.value = 3
    mock_ws.iter_rows.return_value = [[mock_cell]]
    
    result = read_route_type_from_gameinfo(session_dir)
    assert result == 3
    
    # Test with route_type as float
    mock_cell.value = 2.0
    result = read_route_type_from_gameinfo(session_dir)
    assert result == 2
    
    # Test with non-existent file
    non_existent_dir = tmp_path / "non_existent"
    non_existent_dir.mkdir()
    result = read_route_type_from_gameinfo(non_existent_dir)
    assert result is None
    
    # Test with openpyxl not available
    mock_openpyxl.load_workbook.side_effect = ImportError("No module named 'openpyxl'")
    result = read_route_type_from_gameinfo(session_dir)
    assert result is None


def test_get_route_type(tmp_path):
    """Test getting route_type with fallback logic."""
    session_dir = tmp_path / "test_session"
    session_dir.mkdir()
    
    # Test with gameinfo.xlsx (preferred)
    gameinfo_path = session_dir / "gameinfo.xlsx"
    gameinfo_path.touch()
    
    with patch('batch_tracker.read_route_type_from_gameinfo') as mock_gameinfo:
        with patch('batch_tracker.read_route_type_from_action_camera') as mock_action:
            mock_gameinfo.return_value = 1
            mock_action.return_value = 2
            
            result = get_route_type(session_dir)
            assert result == 1  # Should use gameinfo value
            mock_gameinfo.assert_called_once_with(session_dir)
            mock_action.assert_not_called()
    
    # Test fallback to action_camera.json when gameinfo returns None
    with patch('batch_tracker.read_route_type_from_gameinfo') as mock_gameinfo:
        with patch('batch_tracker.read_route_type_from_action_camera') as mock_action:
            mock_gameinfo.return_value = None
            mock_action.return_value = 2
            
            result = get_route_type(session_dir)
            assert result == 2  # Should use action_camera value
            mock_gameinfo.assert_called_once_with(session_dir)
            mock_action.assert_called_once_with(session_dir)
    
    # Test both return None
    with patch('batch_tracker.read_route_type_from_gameinfo') as mock_gameinfo:
        with patch('batch_tracker.read_route_type_from_action_camera') as mock_action:
            mock_gameinfo.return_value = None
            mock_action.return_value = None
            
            result = get_route_type(session_dir)
            assert result is None


def test_compute_distribution(tmp_path):
    """Test computing route_type distribution."""
    # Create mock session directories
    sessions = []
    for i in range(10):
        session_dir = tmp_path / f"session_{i}"
        session_dir.mkdir()
        sessions.append(session_dir)
    
    # Mock get_route_type to return specific values
    route_types = [1, 1, 1, 1, 1, 2, 2, 3, 3, None]  # 5x type1, 2x type2, 2x type3, 1x None
    
    with patch('batch_tracker.get_route_type') as mock_get_route_type:
        mock_get_route_type.side_effect = route_types
        
        counts = compute_distribution(sessions, verbose=False)
        
        assert counts[1] == 5
        assert counts[2] == 2
        assert counts[3] == 2
        # Unknown types are not counted in distribution


def test_check_distribution_within_tolerance():
    """Test distribution check when within tolerance."""
    # Perfect distribution: 50% type1, 25% type2, 25% type3
    counts = {1: 50, 2: 25, 3: 25}
    total = 100
    
    alerts = check_distribution(counts, total)
    assert len(alerts) == 0  # No alerts when within tolerance
    
    # Slightly off but within 10% tolerance
    counts = {1: 55, 2: 20, 3: 25}  # 55%, 20%, 25%
    total = 100
    
    alerts = check_distribution(counts, total)
    assert len(alerts) == 0  # Still within tolerance


def test_check_distribution_outside_tolerance():
    """Test distribution check when outside tolerance."""
    # Type 1 at 65% (should be 50%, deviation +15%)
    counts = {1: 65, 2: 20, 3: 15}
    total = 100
    
    alerts = check_distribution(counts, total)
    assert len(alerts) == 3  # All three are outside tolerance
    
    # Check alert messages
    alert_types = {alert["route_type"] for alert in alerts}
    assert alert_types == {1, 2, 3}
    
    # Check type 1 alert
    type1_alert = [a for a in alerts if a["route_type"] == 1][0]
    assert type1_alert["actual_pct"] == 65.0
    assert type1_alert["target_pct"] == 50.0
    assert type1_alert["deviation"] == 15.0
    assert "exceeds 10% tolerance" in type1_alert["message"]


def test_check_distribution_edge_cases():
    """Test edge cases for distribution check."""
    # Zero total sessions
    counts = {1: 0, 2: 0, 3: 0}
    total = 0
    
    alerts = check_distribution(counts, total)
    # Should handle division by zero gracefully
    assert len(alerts) == 0  # Or maybe 3 alerts for 0% vs target?
    
    # Missing route types
    counts = {1: 100}  # Only type 1
    total = 100
    
    alerts = check_distribution(counts, total)
    assert len(alerts) >= 2  # Should alert for missing types 2 and 3


def test_target_distribution_constants():
    """Test that target distribution constants are correct."""
    assert TARGET_DISTRIBUTION[1] == 0.50  # 50%
    assert TARGET_DISTRIBUTION[2] == 0.25  # 25%
    assert TARGET_DISTRIBUTION[3] == 0.25  # 25%
    assert TOLERANCE_PERCENT == 10  # 10 percentage points


def test_integration_with_mock_files(tmp_path):
    """Integration test with mock Excel files."""
    # Skip if openpyxl is not available
    try:
        import openpyxl
    except ImportError:
        pytest.skip("openpyxl not available")
    
    # Create a recordings directory structure
    recordings_root = tmp_path / "recordings"
    recordings_root.mkdir()
    
    # Create session directories with gameinfo.xlsx files
    for i, route_type in enumerate([1, 1, 1, 2, 3, 1, 2, 3, 1, 2]):
        session_dir = recordings_root / f"session_{i:03d}"
        session_dir.mkdir()
        
        # Create gameinfo.xlsx
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["timestamp", "player_x", "player_y", "player_z", "route_type"])
        ws.append([0, 100, 64, 200, route_type])
        wb.save(str(session_dir / "gameinfo.xlsx"))
    
    # Test the full flow
    sessions = find_session_dirs(recordings_root)
    assert len(sessions) == 10
    
    counts = compute_distribution(sessions, verbose=False)
    total = sum(counts.values())
    
    # Expected: 5x type1, 3x type2, 2x type3
    assert counts[1] == 5  # 50%
    assert counts[2] == 3  # 30%
    assert counts[3] == 2  # 20%
    assert total == 10
    
    alerts = check_distribution(counts, total)
    # Type 2: 30% vs 25% target = +5% (within tolerance)
    # Type 3: 20% vs 25% target = -5% (within tolerance)
    # Type 1: 50% vs 50% target = 0% (within tolerance)
    assert len(alerts) == 0  # All within tolerance


if __name__ == "__main__":
    pytest.main([__file__, "-v"])