#!/usr/bin/env python3
"""
lint_buyer_spec.py — buyer-runnable validator for a Buyer-Spec v1 bundle.

The buyer's 2026-05-01 acceptance spec asks for **four** sibling deliverables
per recording session:

    buyer-bundle/
    ├── video.mp4                   # 5-minute, 1920x1080, 30 fps, h264/hevc
    ├── systeminfo.json             # window placement + DPI
    ├── action_camera.json          # per-frame telemetry (20 fields)
    ├── gameinfo.xlsx               # operator-annotated scene metadata
    └── depth/                      # ≥6 EXR frames per second of video,
        ├── frame_000000.exr        # single-channel float32 metric depth
        ├── frame_000010.exr
        └── ...

This linter validates a candidate buyer bundle against that spec end-to-end.
It runs offline before delivery so we (and the buyer) catch schema drift
before the file leaves our hands.

Checks performed
----------------
 1. ``video.mp4``: ffprobe → duration ∈ [300, 360]s, resolution 1920×1080,
    fps == 30 (within 0.5 tol), codec ∈ {h264, hevc}.
 2. ``systeminfo.json``: required ``gameProcessName`` (str), ``x``/``y``/
    ``width``/``height`` (ints), ``recordDpi`` (float).
 3. ``action_camera.json``: top-level array; each record carries the 20
    buyer-spec fields with correct types/ranges (see Buyer-Spec v1).
    Includes the literal-spaced key ``camera_Follow Offset``.
 4. ``gameinfo.xlsx``: file present and openable via openpyxl (lazy import,
    skipped with hint if openpyxl is absent).
 5. ``depth/*.exr``: at least ``ceil(6 * video_duration)`` files; sample one
    frame via OpenEXR (lazy import, skipped with hint if absent) and check
    it is single-channel, float32.
 6. Cross-deliverable consistency: action_camera frame count matches the
    video frame count (within 1 frame slack).

Exit codes
----------
    0 — bundle is conformant (or warnings only without ``--strict``)
    1 — at least one check failed (or warning under ``--strict``)
    2 — bad usage / bundle directory missing

Usage
-----
    python bin/lint_buyer_spec.py /path/to/buyer-bundle
    python bin/lint_buyer_spec.py /path/to/buyer-bundle --json
    python bin/lint_buyer_spec.py /path/to/buyer-bundle --strict
    python bin/lint_buyer_spec.py /path/to/buyer-bundle --allow-null-engine-fields

The ``--allow-null-engine-fields`` flag is the **Week 1 mode**: engine hooks
have not landed yet, so ``player_position``, ``player_rotation_oula``,
``player_rotation_quaternion``, ``camera_Follow Offset`` and ``metric_scale``
are allowed to be null. Once engine hooks ship in Week 2, drop the flag and
the linter requires real values.
"""

from __future__ import annotations

import argparse
import dataclasses
import json
import math
import shutil
import struct
import subprocess
import sys
from collections.abc import Iterable
from datetime import datetime
from pathlib import Path
from typing import Any

# Optional rich rendering — same fallback pattern as bin/lint_bundle.py so the
# linter still runs in a stripped buyer environment.
try:
    from rich.console import Console
    from rich.table import Table

    _RICH = True
except ImportError:  # pragma: no cover - exercised only on stripped envs
    _RICH = False


# --------------------------------------------------------------------------- types


@dataclasses.dataclass(frozen=True)
class LintIssue:
    """One actionable issue found in the bundle.

    ``severity`` is either ``"error"`` (always fails the lint) or
    ``"warning"`` (only fails under ``--strict``).
    """

    code: str
    file: str
    message: str
    location: str = ""
    severity: str = "error"

    def to_dict(self) -> dict[str, str]:
        """Convert this LintIssue to a dictionary representation.

        Returns:
            Dictionary with keys: code, file, location, message, severity.
        """
        return {
            "code": self.code,
            "file": self.file,
            "location": self.location,
            "message": self.message,
            "severity": self.severity,
        }


@dataclasses.dataclass
class LintResult:
    """Aggregate result of linting one buyer-spec bundle."""

    bundle_dir: Path
    issues: list[LintIssue] = dataclasses.field(default_factory=list)
    checked_files: list[str] = dataclasses.field(default_factory=list)
    skipped_files: list[str] = dataclasses.field(default_factory=list)
    video_duration_sec: float | None = None
    video_frame_count: int | None = None
    action_camera_records: int | None = None
    depth_file_count: int | None = None

    @property
    def errors(self) -> list[LintIssue]:
        """Return all issues with severity 'error'."""
        return [i for i in self.issues if i.severity == "error"]

    @property
    def warnings(self) -> list[LintIssue]:
        """Return all issues with severity 'warning'."""
        return [i for i in self.issues if i.severity == "warning"]

    def ok(self, *, strict: bool = False) -> bool:
        """Check if the lint result passes validation.

        Args:
            strict: If True, warnings also cause failure. If False, only
                errors cause failure.

        Returns:
            True if validation passes, False otherwise.
        """
        if self.errors:
            return False
        return not (strict and self.warnings)

    def to_dict(self) -> dict[str, Any]:
        """Convert lint result to a dictionary for serialization.

        Returns:
            Dictionary containing all lint result data including status,
            bundle info, and collected issues.
        """
        return {
            "ok_strict": self.ok(strict=True),
            "ok_lenient": self.ok(strict=False),
            "bundle_dir": str(self.bundle_dir),
            "video_duration_sec": self.video_duration_sec,
            "video_frame_count": self.video_frame_count,
            "action_camera_records": self.action_camera_records,
            "depth_file_count": self.depth_file_count,
            "checked_files": self.checked_files,
            "skipped_files": self.skipped_files,
            "errors": [i.to_dict() for i in self.errors],
            "warnings": [i.to_dict() for i in self.warnings],
        }


# --------------------------------------------------------------------------- helpers


_INSTALL_HINTS: dict[str, str] = {
    "openpyxl": "pip install openpyxl  # for gameinfo.xlsx validation",
    "OpenEXR": (
        "pip install OpenEXR Imath  # for depth/*.exr validation; "
        "on macOS may also need: brew install openexr"
    ),
}


def _add(
    issues: list[LintIssue],
    *,
    code: str,
    file: str,
    message: str,
    location: str = "",
    severity: str = "error",
) -> None:
    issues.append(
        LintIssue(
            code=code,
            file=file,
            message=message,
            location=location,
            severity=severity,
        )
    )


def _is_finite_number(v: Any) -> bool:
    """Return True iff ``v`` is a finite int/float (and not a bool)."""
    if isinstance(v, bool):
        return False
    if isinstance(v, int):
        return True
    if isinstance(v, float):
        return math.isfinite(v)
    return False


def _is_int(v: Any) -> bool:
    return isinstance(v, int) and not isinstance(v, bool)


def _is_unit_quaternion(values: list[float], tol: float = 1e-3) -> bool:
    if len(values) != 4:
        return False
    norm_sq = sum(float(v) * float(v) for v in values)
    return math.isclose(norm_sq, 1.0, rel_tol=0.0, abs_tol=tol)


def _quaternion_norm_error(values: list[float]) -> float | None:
    if len(values) != 4:
        return None
    norm_sq = sum(float(v) * float(v) for v in values)
    return abs(norm_sq - 1.0)


# --------------------------------------------------------------------------- ffprobe


def _run_ffprobe(path: Path) -> dict[str, Any] | None:
    """Run ffprobe on ``path`` and return parsed JSON, or None on hard failure."""
    if shutil.which("ffprobe") is None:
        return None
    try:
        proc = subprocess.run(
            [
                "ffprobe",
                "-v",
                "error",
                "-print_format",
                "json",
                "-show_format",
                "-show_streams",
                str(path),
            ],
            capture_output=True,
            text=True,
            check=False,
            timeout=60,
        )
    except (subprocess.TimeoutExpired, OSError):
        return None
    if proc.returncode != 0:
        return None
    try:
        return json.loads(proc.stdout)
    except json.JSONDecodeError:
        return None


def _parse_fps(rate_str: str) -> float | None:
    """ffprobe reports rates as ``"30000/1001"`` etc. Reduce to a float."""
    if not rate_str:
        return None
    if "/" in rate_str:
        num, _, den = rate_str.partition("/")
        try:
            n = float(num)
            d = float(den)
        except ValueError:
            return None
        if d == 0:
            return None
        return n / d
    try:
        return float(rate_str)
    except ValueError:
        return None


# --------------------------------------------------------------------------- check: video


_VIDEO_MIN_SEC = 300.0
_VIDEO_MAX_SEC = 360.0
_VIDEO_REQ_WIDTH = 1920
_VIDEO_REQ_HEIGHT = 1080
_VIDEO_REQ_FPS = 30.0
_VIDEO_FPS_TOL = 0.5
_ACCEPTED_CODECS: frozenset[str] = frozenset({"h264", "hevc"})


def _check_video(bundle: Path, result: LintResult) -> None:
    path = bundle / "video.mp4"
    here = "video.mp4"
    if not path.exists():
        _add(
            result.issues,
            code="VIDEO_MISSING",
            file=here,
            message=f"required file not found at {path}",
        )
        return

    if shutil.which("ffprobe") is None:
        _add(
            result.issues,
            code="FFPROBE_NOT_INSTALLED",
            file=here,
            message=(
                "ffprobe is not on PATH; cannot validate video metadata. "
                "Install via: brew install ffmpeg (macOS) or apt-get install ffmpeg (Linux)."
            ),
        )
        return

    probe = _run_ffprobe(path)
    if probe is None:
        _add(
            result.issues,
            code="VIDEO_FFPROBE_FAILED",
            file=here,
            message="ffprobe could not parse the file (corrupt mp4 or unsupported container?)",
        )
        return

    # Find the first video stream.
    streams = probe.get("streams", [])
    video_stream: dict[str, Any] | None = None
    for s in streams:
        if s.get("codec_type") == "video":
            video_stream = s
            break
    if video_stream is None:
        _add(
            result.issues,
            code="VIDEO_NO_STREAM",
            file=here,
            message="ffprobe reports no video stream in the container",
        )
        return

    # Duration may live on the format dict OR on the stream — prefer format.
    duration: float | None = None
    fmt = probe.get("format", {})
    for source in (fmt.get("duration"), video_stream.get("duration")):
        if source is None:
            continue
        try:
            duration = float(source)
            break
        except (TypeError, ValueError):
            continue
    if duration is None:
        _add(
            result.issues,
            code="VIDEO_NO_DURATION",
            file=here,
            message="ffprobe did not report a duration for the video",
        )
        return

    result.video_duration_sec = duration
    if duration < _VIDEO_MIN_SEC:
        _add(
            result.issues,
            code="VIDEO_TOO_SHORT",
            file=here,
            message=(
                f"duration={duration:.2f}s is below buyer minimum "
                f"{_VIDEO_MIN_SEC:.0f}s (5 minutes)"
            ),
            location="format.duration",
        )
    if duration > _VIDEO_MAX_SEC:
        _add(
            result.issues,
            code="VIDEO_TOO_LONG",
            file=here,
            message=(
                f"duration={duration:.2f}s exceeds buyer maximum "
                f"{_VIDEO_MAX_SEC:.0f}s (6 minutes)"
            ),
            location="format.duration",
        )

    # Resolution
    width = video_stream.get("width")
    height = video_stream.get("height")
    if width != _VIDEO_REQ_WIDTH or height != _VIDEO_REQ_HEIGHT:
        _add(
            result.issues,
            code="VIDEO_RESOLUTION",
            file=here,
            message=(
                f"resolution {width}x{height} does not match required "
                f"{_VIDEO_REQ_WIDTH}x{_VIDEO_REQ_HEIGHT}"
            ),
            location="streams[0]",
        )

    # FPS — prefer avg_frame_rate, fall back to r_frame_rate.
    fps_raw = video_stream.get("avg_frame_rate") or video_stream.get("r_frame_rate")
    fps = _parse_fps(str(fps_raw) if fps_raw else "")
    if fps is None:
        _add(
            result.issues,
            code="VIDEO_FPS_UNKNOWN",
            file=here,
            message="could not parse avg_frame_rate / r_frame_rate from ffprobe output",
            location="streams[0].avg_frame_rate",
        )
    elif abs(fps - _VIDEO_REQ_FPS) > _VIDEO_FPS_TOL:
        _add(
            result.issues,
            code="VIDEO_FPS",
            file=here,
            message=(
                f"fps={fps:.3f} differs from required {_VIDEO_REQ_FPS} "
                f"by more than {_VIDEO_FPS_TOL}"
            ),
            location="streams[0].avg_frame_rate",
        )

    # Codec
    codec_name = (video_stream.get("codec_name") or "").lower()
    if codec_name not in _ACCEPTED_CODECS:
        _add(
            result.issues,
            code="VIDEO_CODEC",
            file=here,
            message=(
                f"codec_name={codec_name!r} is not in accepted set " f"{sorted(_ACCEPTED_CODECS)}"
            ),
            location="streams[0].codec_name",
        )

    # Frame count: prefer the explicit nb_frames, else duration * fps.
    nb_frames_raw = video_stream.get("nb_frames")
    if nb_frames_raw is not None:
        try:
            result.video_frame_count = int(nb_frames_raw)
        except (TypeError, ValueError):
            result.video_frame_count = None
    if result.video_frame_count is None and fps is not None:
        result.video_frame_count = int(round(duration * fps))

    result.checked_files.append(here)


# --------------------------------------------------------------------------- check: systeminfo.json


_SYSTEMINFO_REQUIRED: tuple[tuple[str, type | tuple[type, ...]], ...] = (
    ("gameProcessName", str),
    ("x", int),
    ("y", int),
    ("width", int),
    ("height", int),
    ("recordDpi", (int, float)),
)


def _check_systeminfo(bundle: Path, result: LintResult) -> None:
    path = bundle / "systeminfo.json"
    here = "systeminfo.json"
    if not path.exists():
        _add(
            result.issues,
            code="SYSTEMINFO_MISSING",
            file=here,
            message=f"required file not found at {path}",
        )
        return
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        _add(
            result.issues,
            code="MALFORMED_JSON",
            file=here,
            message=f"file is not valid JSON: {exc}",
            location=f"{here}:{exc.lineno}",
        )
        return
    if not isinstance(data, dict):
        _add(
            result.issues,
            code="SYSTEMINFO_BAD_TOPLEVEL",
            file=here,
            message="top-level value must be a JSON object",
        )
        return

    for field, expected_type in _SYSTEMINFO_REQUIRED:
        if field not in data:
            _add(
                result.issues,
                code="SYSTEMINFO_MISSING_FIELD",
                file=here,
                message=f"missing required field {field!r}",
                location=f"{here}.{field}",
            )
            continue
        v = data[field]
        # bool slips through int isinstance — exclude explicitly.
        if isinstance(v, bool):
            _add(
                result.issues,
                code="SYSTEMINFO_BAD_TYPE",
                file=here,
                message=f"{field!r} must be {expected_type}, got bool",
                location=f"{here}.{field}",
            )
            continue
        if not isinstance(v, expected_type):
            _add(
                result.issues,
                code="SYSTEMINFO_BAD_TYPE",
                file=here,
                message=(f"{field!r} must be {expected_type}, " f"got {type(v).__name__}={v!r}"),
                location=f"{here}.{field}",
            )

    result.checked_files.append(here)


# --------------------------------------------------------------------------- check: action_camera.json

# Fields that the engine hook will populate; nullable in Week 1 mode.
# camera_speed and player_speed are also derived from engine telemetry
# (camera_speed needs camera_position deltas; player_speed needs the
# engine's velocity vector). Until engine hooks land they are emitted
# as null too.
_ENGINE_NULLABLE_FIELDS: frozenset[str] = frozenset(
    {
        "player_position",
        "player_rotation_oula",
        "player_rotation_quaternion",
        "player_speed",
        "camera_Follow Offset",
        "camera_speed",
        "metric_scale",
    }
)

_ROUTE_TYPES: frozenset[int] = frozenset({1, 2, 3})


def _check_vector(
    rec: dict[str, Any],
    field: str,
    expected_len: int,
    here: str,
    issues: list[LintIssue],
    *,
    nullable: bool = False,
    elem_range: tuple[float, float] | None = None,
) -> list[float] | None:
    """Validate that ``rec[field]`` is a list of ``expected_len`` finite numbers.

    Returns the parsed list of floats on success, None on failure.
    Emits issues for missing/malformed fields.
    """
    if field not in rec:
        _add(
            issues,
            code="ACTION_CAMERA_MISSING_FIELD",
            file="action_camera.json",
            message=f"required field {field!r} is missing",
            location=f"{here}.{field}",
        )
        return None
    v = rec[field]
    if v is None:
        if nullable:
            return None
        _add(
            issues,
            code="ACTION_CAMERA_BAD_FIELD",
            file="action_camera.json",
            message=f"{field!r} must not be null (engine fields not yet allowed null)",
            location=f"{here}.{field}",
        )
        return None
    if not isinstance(v, list) or len(v) != expected_len:
        _add(
            issues,
            code="ACTION_CAMERA_BAD_FIELD",
            file="action_camera.json",
            message=f"{field!r} must be a list of {expected_len} finite numbers",
            location=f"{here}.{field}",
        )
        return None
    if not all(_is_finite_number(elem) for elem in v):
        _add(
            issues,
            code="ACTION_CAMERA_BAD_FIELD",
            file="action_camera.json",
            message=f"{field!r} must contain only finite numbers (no NaN/inf/bool)",
            location=f"{here}.{field}",
        )
        return None
    floats = [float(elem) for elem in v]
    if elem_range is not None:
        lo, hi = elem_range
        for j, elem in enumerate(floats):
            if not lo <= elem <= hi:
                _add(
                    issues,
                    code="ACTION_CAMERA_RANGE",
                    file="action_camera.json",
                    message=(f"{field!r}[{j}]={elem} is outside required range [{lo}, {hi}]"),
                    location=f"{here}.{field}",
                )
    return floats


def _check_scalar_range(
    rec: dict[str, Any],
    field: str,
    here: str,
    issues: list[LintIssue],
    *,
    nullable: bool = False,
    rng: tuple[float, float] | None = None,
    positive: bool = False,
) -> float | None:
    """Validate a scalar float field with optional range / positive constraints."""
    if field not in rec:
        _add(
            issues,
            code="ACTION_CAMERA_MISSING_FIELD",
            file="action_camera.json",
            message=f"required field {field!r} is missing",
            location=f"{here}.{field}",
        )
        return None
    v = rec[field]
    if v is None:
        if nullable:
            return None
        _add(
            issues,
            code="ACTION_CAMERA_BAD_FIELD",
            file="action_camera.json",
            message=f"{field!r} must not be null",
            location=f"{here}.{field}",
        )
        return None
    if not _is_finite_number(v):
        _add(
            issues,
            code="ACTION_CAMERA_BAD_FIELD",
            file="action_camera.json",
            message=f"{field!r} must be a finite number",
            location=f"{here}.{field}",
        )
        return None
    fv = float(v)
    if rng is not None:
        lo, hi = rng
        if not lo <= fv <= hi:
            _add(
                issues,
                code="ACTION_CAMERA_RANGE",
                file="action_camera.json",
                message=f"{field!r}={fv} is outside required range [{lo}, {hi}]",
                location=f"{here}.{field}",
            )
    if positive and fv <= 0:
        _add(
            issues,
            code="ACTION_CAMERA_RANGE",
            file="action_camera.json",
            message=f"{field!r}={fv} must be > 0",
            location=f"{here}.{field}",
        )
    return fv


def _check_intrinsics(
    rec: dict[str, Any],
    here: str,
    issues: list[LintIssue],
) -> None:
    """``camera_intrinsics`` is a dict with fx, fy, cx, cy floats; fx == fy required."""
    field = "camera_intrinsics"
    if field not in rec:
        _add(
            issues,
            code="ACTION_CAMERA_MISSING_FIELD",
            file="action_camera.json",
            message=f"required field {field!r} is missing",
            location=f"{here}.{field}",
        )
        return
    v = rec[field]
    if not isinstance(v, dict):
        _add(
            issues,
            code="ACTION_CAMERA_BAD_FIELD",
            file="action_camera.json",
            message=f"{field!r} must be a JSON object with fx, fy, cx, cy",
            location=f"{here}.{field}",
        )
        return
    parsed: dict[str, float] = {}
    for sub in ("fx", "fy", "cx", "cy"):
        if sub not in v:
            _add(
                issues,
                code="ACTION_CAMERA_MISSING_FIELD",
                file="action_camera.json",
                message=f"{field}.{sub} is required",
                location=f"{here}.{field}.{sub}",
            )
            continue
        sv = v[sub]
        if not _is_finite_number(sv):
            _add(
                issues,
                code="ACTION_CAMERA_BAD_FIELD",
                file="action_camera.json",
                message=f"{field}.{sub} must be a finite number",
                location=f"{here}.{field}.{sub}",
            )
            continue
        parsed[sub] = float(sv)
    if "fx" in parsed and "fy" in parsed and parsed["fx"] != parsed["fy"]:
        _add(
            issues,
            code="ACTION_CAMERA_FX_FY_MISMATCH",
            file="action_camera.json",
            message=(
                f"camera_intrinsics.fx ({parsed['fx']}) must equal camera_intrinsics.fy "
                f"({parsed['fy']}) per buyer-spec (square pixels)"
            ),
            location=f"{here}.{field}",
        )


def _check_action_camera_record(
    rec: Any,
    idx: int,
    issues: list[LintIssue],
    *,
    allow_null_engine_fields: bool,
    prev_frame: int | None,
) -> int | None:
    """Validate one record. Returns its frame number on success, else None."""
    here = f"action_camera.json[{idx}]"
    if not isinstance(rec, dict):
        _add(
            issues,
            code="ACTION_CAMERA_BAD_RECORD",
            file="action_camera.json",
            message="record must be a JSON object",
            location=here,
        )
        return None

    # time — string parseable as datetime with ms precision.
    t = rec.get("time", "__missing__")
    if t == "__missing__":
        _add(
            issues,
            code="ACTION_CAMERA_MISSING_FIELD",
            file="action_camera.json",
            message="required field 'time' is missing",
            location=f"{here}.time",
        )
    elif not isinstance(t, str):
        _add(
            issues,
            code="ACTION_CAMERA_BAD_FIELD",
            file="action_camera.json",
            message="'time' must be a string in ISO-like datetime format with ms",
            location=f"{here}.time",
        )
    else:
        # Try a few parse forms — fromisoformat (3.11+) handles most modern variants.
        parsed_ok = False
        for fmt in (
            "%Y-%m-%d %H:%M:%S.%f",
            "%Y-%m-%dT%H:%M:%S.%f",
            "%Y-%m-%dT%H:%M:%S.%fZ",
            "%Y/%m/%d %H:%M:%S.%f",
        ):
            try:
                datetime.strptime(t, fmt)
                parsed_ok = True
                break
            except ValueError:
                continue
        if not parsed_ok:
            try:
                datetime.fromisoformat(t.replace("Z", "+00:00"))
                parsed_ok = True
            except ValueError:
                pass
        if not parsed_ok:
            _add(
                issues,
                code="ACTION_CAMERA_BAD_FIELD",
                file="action_camera.json",
                message=f"'time'={t!r} is not parseable as a datetime with ms precision",
                location=f"{here}.time",
            )

    # fps — float > 0
    _check_scalar_range(rec, "fps", here, issues, positive=True)

    # frame — int ≥ 0, monotonic
    frame_v = rec.get("frame", "__missing__")
    frame_int: int | None = None
    if frame_v == "__missing__":
        _add(
            issues,
            code="ACTION_CAMERA_MISSING_FIELD",
            file="action_camera.json",
            message="required field 'frame' is missing",
            location=f"{here}.frame",
        )
    elif not _is_int(frame_v):
        _add(
            issues,
            code="ACTION_CAMERA_BAD_FIELD",
            file="action_camera.json",
            message="'frame' must be an integer",
            location=f"{here}.frame",
        )
    elif frame_v < 0:
        _add(
            issues,
            code="ACTION_CAMERA_RANGE",
            file="action_camera.json",
            message=f"'frame'={frame_v} must be ≥ 0",
            location=f"{here}.frame",
        )
    else:
        frame_int = frame_v
        if prev_frame is not None and frame_int <= prev_frame:
            _add(
                issues,
                code="ACTION_CAMERA_NON_MONOTONIC",
                file="action_camera.json",
                message=(
                    f"'frame'={frame_int} is not strictly greater than "
                    f"previous record's frame={prev_frame}"
                ),
                location=f"{here}.frame",
            )

    # route_type — int ∈ {1,2,3}
    rt = rec.get("route_type", "__missing__")
    if rt == "__missing__":
        _add(
            issues,
            code="ACTION_CAMERA_MISSING_FIELD",
            file="action_camera.json",
            message="required field 'route_type' is missing",
            location=f"{here}.route_type",
        )
    elif not _is_int(rt):
        _add(
            issues,
            code="ACTION_CAMERA_BAD_FIELD",
            file="action_camera.json",
            message="'route_type' must be an integer",
            location=f"{here}.route_type",
        )
    elif rt not in _ROUTE_TYPES:
        _add(
            issues,
            code="ACTION_CAMERA_BAD_FIELD",
            file="action_camera.json",
            message=f"'route_type'={rt} must be one of {sorted(_ROUTE_TYPES)}",
            location=f"{here}.route_type",
        )

    # Camera vectors / quaternions
    _check_vector(rec, "camera_position", 3, here, issues)
    _check_vector(
        rec,
        "camera_rotation_oula",
        3,
        here,
        issues,
        elem_range=(-180.0, 180.0),
    )
    cq = _check_vector(rec, "camera_rotation_quaternion", 4, here, issues)
    if cq is not None and not _is_unit_quaternion(cq):
        norm_err = _quaternion_norm_error(cq)
        _add(
            issues,
            code="ACTION_CAMERA_NON_UNIT_QUAT",
            file="action_camera.json",
            message=(
                f"camera_rotation_quaternion is not unit-norm "
                f"(|q|^2 - 1 = {norm_err:.6g}, tolerance 1e-3)"
            ),
            location=f"{here}.camera_rotation_quaternion",
            severity="warning",
        )

    _check_vector(
        rec,
        "camera_Follow Offset",
        3,
        here,
        issues,
        nullable=allow_null_engine_fields,
    )

    _check_intrinsics(rec, here, issues)

    _check_vector(
        rec,
        "camera_speed",
        3,
        here,
        issues,
        nullable=allow_null_engine_fields,
    )

    # Player vectors / quaternions — engine-side, nullable in Week 1 mode.
    _check_vector(
        rec,
        "player_position",
        3,
        here,
        issues,
        nullable=allow_null_engine_fields,
    )
    _check_vector(
        rec,
        "player_rotation_oula",
        3,
        here,
        issues,
        nullable=allow_null_engine_fields,
        elem_range=(-180.0, 180.0),
    )
    pq = _check_vector(
        rec,
        "player_rotation_quaternion",
        4,
        here,
        issues,
        nullable=allow_null_engine_fields,
    )
    if pq is not None and not _is_unit_quaternion(pq):
        norm_err = _quaternion_norm_error(pq)
        _add(
            issues,
            code="ACTION_CAMERA_NON_UNIT_QUAT",
            file="action_camera.json",
            message=(
                f"player_rotation_quaternion is not unit-norm "
                f"(|q|^2 - 1 = {norm_err:.6g}, tolerance 1e-3)"
            ),
            location=f"{here}.player_rotation_quaternion",
            severity="warning",
        )

    _check_vector(
        rec,
        "player_speed",
        3,
        here,
        issues,
        nullable=allow_null_engine_fields,
    )

    # Mouse normalized coords / deltas
    _check_scalar_range(rec, "mouse_x", here, issues, rng=(0.0, 1.0))
    _check_scalar_range(rec, "mouse_y", here, issues, rng=(0.0, 1.0))
    _check_scalar_range(rec, "mouse_dx", here, issues, rng=(-1.0, 1.0))
    _check_scalar_range(rec, "mouse_dy", here, issues, rng=(-1.0, 1.0))

    # keyCode — list of ints
    kc = rec.get("keyCode", "__missing__")
    if kc == "__missing__":
        _add(
            issues,
            code="ACTION_CAMERA_MISSING_FIELD",
            file="action_camera.json",
            message="required field 'keyCode' is missing",
            location=f"{here}.keyCode",
        )
    elif not isinstance(kc, list):
        _add(
            issues,
            code="ACTION_CAMERA_BAD_FIELD",
            file="action_camera.json",
            message="'keyCode' must be a list of integers (use [] for empty)",
            location=f"{here}.keyCode",
        )
    else:
        for j, code in enumerate(kc):
            if not _is_int(code):
                _add(
                    issues,
                    code="ACTION_CAMERA_BAD_FIELD",
                    file="action_camera.json",
                    message="'keyCode' entries must be integers",
                    location=f"{here}.keyCode[{j}]",
                )

    # metric_scale — engine-side, nullable in Week 1 mode; > 0 otherwise.
    _check_scalar_range(
        rec,
        "metric_scale",
        here,
        issues,
        nullable=allow_null_engine_fields,
        positive=True,
    )

    return frame_int


def _check_action_camera(
    bundle: Path,
    result: LintResult,
    *,
    allow_null_engine_fields: bool,
) -> None:
    path = bundle / "action_camera.json"
    here = "action_camera.json"
    if not path.exists():
        _add(
            result.issues,
            code="ACTION_CAMERA_MISSING",
            file=here,
            message=f"required file not found at {path}",
        )
        return
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        _add(
            result.issues,
            code="MALFORMED_JSON",
            file=here,
            message=f"file is not valid JSON: {exc}",
            location=f"{here}:{exc.lineno}",
        )
        return
    if not isinstance(data, list):
        _add(
            result.issues,
            code="ACTION_CAMERA_BAD_TOPLEVEL",
            file=here,
            message="top-level value must be a JSON array",
        )
        return

    prev_frame: int | None = None
    for i, rec in enumerate(data):
        returned = _check_action_camera_record(
            rec,
            i,
            result.issues,
            allow_null_engine_fields=allow_null_engine_fields,
            prev_frame=prev_frame,
        )
        if returned is not None:
            prev_frame = returned

    result.action_camera_records = len(data)
    result.checked_files.append(here)


# --------------------------------------------------------------------------- check: gameinfo.xlsx


def _check_gameinfo(bundle: Path, result: LintResult) -> None:
    # Spec mentions "gameinfo.xlsx (or .xlsx)" — accept the canonical name.
    path = bundle / "gameinfo.xlsx"
    here = "gameinfo.xlsx"
    if not path.exists():
        _add(
            result.issues,
            code="GAMEINFO_MISSING",
            file=here,
            message=f"required file not found at {path}",
        )
        return

    try:
        import openpyxl  # type: ignore[import-not-found]
    except ImportError:
        _add(
            result.issues,
            code="GAMEINFO_OPENPYXL_MISSING",
            file=here,
            message=(
                "openpyxl is not installed; cannot validate xlsx structure. "
                f"Install hint: {_INSTALL_HINTS['openpyxl']}"
            ),
            severity="warning",
        )
        result.skipped_files.append(here)
        return

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # pragma: no cover - depends on bad file shape
        _add(
            result.issues,
            code="GAMEINFO_UNREADABLE",
            file=here,
            message=f"openpyxl failed to load workbook: {exc}",
        )
        return

    sheet_names = wb.sheetnames
    if not sheet_names:
        _add(
            result.issues,
            code="GAMEINFO_NO_SHEETS",
            file=here,
            message="workbook has no sheets",
        )
        wb.close()
        return

    # Spec says "sheet 0 readable" — we just confirm it has at least one row.
    first_sheet = wb[sheet_names[0]]
    has_row = False
    for _ in first_sheet.iter_rows(min_row=1, max_row=1, values_only=True):
        has_row = True
        break
    if not has_row:
        _add(
            result.issues,
            code="GAMEINFO_EMPTY_SHEET",
            file=here,
            message=f"sheet 0 ({sheet_names[0]!r}) has no rows",
            severity="warning",
        )
    wb.close()
    result.checked_files.append(here)


# --------------------------------------------------------------------------- check: depth/*.exr


def _read_exr_lazy(path: Path) -> tuple[bool, str]:
    """Open one EXR file via OpenEXR (and Imath if available).

    Returns ``(ok, message)``. ``ok=True`` means the file is single-channel,
    float32. ``ok=False`` means a structural issue (with details in
    ``message``). Caller decides whether to add an issue or skip.

    Handles two API generations:

    * OpenEXR 3.x: ``OpenEXR.FLOAT`` is a ``PixelType`` enum the channel
      type compares equal to. Imath is not required.
    * OpenEXR 1.x/2.x: pixel types come from ``Imath.PixelType.FLOAT`` and
      expose ``.v`` ints.
    """
    try:
        import OpenEXR  # type: ignore[import-not-found]
    except ImportError as exc:
        return False, f"OpenEXR import failed: {exc}"

    try:
        Imath = __import__("Imath")  # noqa: N806 - third-party module name
    except ImportError:
        Imath = None  # noqa: N806

    try:
        f = OpenEXR.InputFile(str(path))
    except Exception as exc:  # pragma: no cover - depends on bad file shape
        return False, f"OpenEXR could not open {path.name}: {exc}"
    try:
        header = f.header()
        channels = header.get("channels", {})
        names = list(channels.keys())
        if len(names) != 1:
            return (
                False,
                f"expected 1 channel, got {len(names)}: {names}",
            )
        chan_name = names[0]
        chan = channels[chan_name]
        ptype = getattr(chan, "type", None)
        if ptype is None:
            return False, f"channel {chan_name!r} has no type info"

        # Acceptance: ptype identifies float32 in any of the supported
        # OpenEXR/Imath pixel-type representations.
        # FLOAT canonical int value across all APIs is 2.
        is_float = False

        # 1) Imath legacy: ``.v`` attribute on PixelType objects.
        type_v = getattr(ptype, "v", None)
        if type_v == 2:
            is_float = True

        # 2) OpenEXR 3.x: the channel.type IS an OpenEXR.PixelType enum
        #    whose int() value is 2 (FLOAT). Wrap in try/except because
        #    enum classes from the two libraries are not cross-comparable.
        if not is_float:
            try:
                if int(ptype) == 2:
                    is_float = True
            except (TypeError, ValueError):
                pass

        # 3) Direct equality to the FLOAT sentinel from either library.
        if not is_float:
            ekey = getattr(OpenEXR, "FLOAT", None)
            if ekey is not None:
                try:
                    if ptype == ekey:
                        is_float = True
                except (AttributeError, TypeError):
                    pass
        if not is_float and Imath is not None:
            ikey = getattr(getattr(Imath, "PixelType", None), "FLOAT", None)
            if ikey is not None:
                try:
                    if ptype == ikey:
                        is_float = True
                except (AttributeError, TypeError):
                    pass

        # 4) Last resort: name-based identification.
        if not is_float and "FLOAT" in repr(ptype).upper():
            is_float = True

        if not is_float:
            return (
                False,
                f"channel {chan_name!r} is not float32 (type={ptype!r})",
            )
        return True, f"single-channel float32 ({chan_name})"
    finally:
        f.close()


def _looks_like_exr(path: Path) -> bool:
    """Cheap magic-byte check — EXR starts with 0x762f3101."""
    try:
        with path.open("rb") as fh:
            magic = fh.read(4)
    except OSError:
        return False
    return magic == b"\x76\x2f\x31\x01"


def _check_depth(bundle: Path, result: LintResult) -> None:
    depth_dir = bundle / "depth"
    here = "depth/"
    if not depth_dir.exists() or not depth_dir.is_dir():
        _add(
            result.issues,
            code="DEPTH_DIR_MISSING",
            file=here,
            message=f"required directory not found at {depth_dir}",
        )
        return

    exr_files = sorted(p for p in depth_dir.iterdir() if p.suffix.lower() == ".exr")
    result.depth_file_count = len(exr_files)
    if not exr_files:
        _add(
            result.issues,
            code="DEPTH_NO_FILES",
            file=here,
            message="depth/ has no .exr files",
        )
        return

    # 6 frames per second of video — only enforceable if we know the duration.
    if result.video_duration_sec is not None:
        required = math.ceil(6 * result.video_duration_sec)
        if len(exr_files) < required:
            _add(
                result.issues,
                code="DEPTH_INSUFFICIENT_FRAMES",
                file=here,
                message=(
                    f"depth/ has {len(exr_files)} EXR files; buyer requires ≥6/sec "
                    f"of video (={required} for {result.video_duration_sec:.1f}s)"
                ),
            )

    # Magic-byte sanity check on every file (cheap).
    for path in exr_files:
        if not _looks_like_exr(path):
            _add(
                result.issues,
                code="DEPTH_BAD_MAGIC",
                file=str(path.relative_to(bundle)),
                message="file does not start with the EXR magic bytes 0x762f3101",
            )

    # Sample one EXR via OpenEXR for type/channel verification. Pick the
    # middle file to give a representative read.
    sample = exr_files[len(exr_files) // 2]
    ok, message = _read_exr_lazy(sample)
    if not ok:
        if message.startswith("OpenEXR import failed"):
            _add(
                result.issues,
                code="DEPTH_OPENEXR_MISSING",
                file=here,
                message=(
                    f"OpenEXR import failed; cannot validate EXR structure. "
                    f"Install hint: {_INSTALL_HINTS['OpenEXR']}"
                ),
                severity="warning",
            )
            result.skipped_files.append("depth/*.exr structural check")
        else:
            _add(
                result.issues,
                code="DEPTH_BAD_FORMAT",
                file=str(sample.relative_to(bundle)),
                message=message,
            )
    result.checked_files.append(here)


# --------------------------------------------------------------------------- cross-deliverable consistency


_FRAME_COUNT_SLACK = 1


def _check_consistency(result: LintResult) -> None:
    if result.video_frame_count is not None and result.action_camera_records is not None:
        diff = abs(result.video_frame_count - result.action_camera_records)
        if diff > _FRAME_COUNT_SLACK:
            _add(
                result.issues,
                code="FRAME_COUNT_MISMATCH",
                file="action_camera.json",
                message=(
                    f"action_camera.json has {result.action_camera_records} records but "
                    f"video.mp4 has {result.video_frame_count} frames "
                    f"(diff={diff} > slack={_FRAME_COUNT_SLACK})"
                ),
            )


# --------------------------------------------------------------------------- top-level


def lint_buyer_spec(
    bundle_dir: Path,
    *,
    allow_null_engine_fields: bool = False,
) -> LintResult:
    """Lint one buyer-spec bundle directory. Returns the aggregate result."""
    bundle_dir = Path(bundle_dir).resolve()
    result = LintResult(bundle_dir=bundle_dir)

    if not bundle_dir.exists() or not bundle_dir.is_dir():
        _add(
            result.issues,
            code="BUNDLE_DIR_MISSING",
            file=str(bundle_dir),
            message="bundle directory does not exist or is not a directory",
        )
        return result

    _check_video(bundle_dir, result)
    _check_systeminfo(bundle_dir, result)
    _check_action_camera(
        bundle_dir,
        result,
        allow_null_engine_fields=allow_null_engine_fields,
    )
    _check_gameinfo(bundle_dir, result)
    _check_depth(bundle_dir, result)
    _check_consistency(result)
    return result


# --------------------------------------------------------------------------- reporting


def _file_status_lines(result: LintResult) -> Iterable[str]:
    issues_by_file: dict[str, list[LintIssue]] = {}
    for issue in result.issues:
        issues_by_file.setdefault(issue.file, []).append(issue)
    for f in result.checked_files:
        if f in issues_by_file and any(i.severity == "error" for i in issues_by_file[f]):
            yield f"[FAIL] {f}"
        elif f in issues_by_file:
            yield f"[WARN] {f}"
        else:
            yield f"[OK]   {f}"
    for f in result.skipped_files:
        yield f"[SKIP] {f}"


def _emit_text(result: LintResult, *, strict: bool) -> None:
    if _RICH:
        console = Console()
        if result.ok(strict=strict):
            console.print(
                f"[bold green]OK[/]: buyer-spec bundle is conformant "
                f"({len(result.warnings)} warning(s))"
            )
        else:
            console.print(
                f"[bold red]FAIL[/]: buyer-spec bundle has "
                f"{len(result.errors)} error(s) and {len(result.warnings)} warning(s) "
                f"in {result.bundle_dir}"
            )
        for line in _file_status_lines(result):
            console.print(line)
        if result.issues:
            table = Table(show_header=True, header_style="bold")
            table.add_column("severity")
            table.add_column("code")
            table.add_column("file")
            table.add_column("location")
            table.add_column("message")
            for issue in result.issues:
                table.add_row(
                    issue.severity,
                    issue.code,
                    issue.file,
                    issue.location,
                    issue.message,
                )
            console.print(table)
        return

    if result.ok(strict=strict):
        print(f"OK: buyer-spec bundle is conformant ({len(result.warnings)} warning(s))")
    else:
        print(
            f"FAIL: buyer-spec bundle has {len(result.errors)} error(s) and "
            f"{len(result.warnings)} warning(s) in {result.bundle_dir}"
        )
    for line in _file_status_lines(result):
        print(line)
    for issue in result.issues:
        loc = f" ({issue.location})" if issue.location else ""
        print(f"  [{issue.severity.upper()}/{issue.code}] {issue.file}{loc}: {issue.message}")


def _emit_json(result: LintResult) -> None:
    print(json.dumps(result.to_dict(), indent=2, sort_keys=True))


# --------------------------------------------------------------------------- CLI


def _build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="lint_buyer_spec",
        description=(
            "Validate a Buyer-Spec v1 bundle (video.mp4 + systeminfo.json + "
            "action_camera.json + gameinfo.xlsx + depth/*.exr) end-to-end."
        ),
    )
    p.add_argument(
        "bundle_dir",
        type=Path,
        help="path to the buyer-spec bundle directory",
    )
    p.add_argument(
        "--json",
        dest="emit_json",
        action="store_true",
        help="emit machine-readable JSON instead of human-readable text",
    )
    p.add_argument(
        "--strict",
        action="store_true",
        help="treat warnings (e.g., quaternion norm slightly off) as failures",
    )
    p.add_argument(
        "--allow-null-engine-fields",
        action="store_true",
        help=(
            "Week 1 mode: allow player_position, player_rotation_*, "
            "camera_Follow Offset, and metric_scale to be null while engine "
            "hooks are not yet landed (see docs/BUYER_SPEC_v1_PLAN.md)"
        ),
    )
    return p


def main(argv: list[str] | None = None) -> int:
    """
    Entry point for lint_buyer_spec CLI.

    Validates a game data bundle directory against the buyer spec schema.
    Checks for required files (video.mp4, action_camera.bin, gameinfo.xlsx)
    and validates their contents (FPS, resolution, intrinsics, etc.).

    Args:
        argv: Command-line arguments. If None, uses sys.argv.
              Supports --bundle-dir, --strict, --emit-json, --allow-null-engine-fields.

    Returns:
        Exit code: 0 if lint passes, 1 if failures found, 2 if bundle_dir missing.
    """
    args = _build_parser().parse_args(argv)
    bundle_dir: Path = args.bundle_dir
    if not bundle_dir.exists():
        msg = f"bundle directory does not exist: {bundle_dir}"
        if args.emit_json:
            print(
                json.dumps(
                    {
                        "ok_strict": False,
                        "ok_lenient": False,
                        "bundle_dir": str(bundle_dir),
                        "errors": [{"code": "BUNDLE_DIR_MISSING", "message": msg}],
                    }
                )
            )
        else:
            print(f"ERROR: {msg}", file=sys.stderr)
        return 2

    result = lint_buyer_spec(
        bundle_dir,
        allow_null_engine_fields=args.allow_null_engine_fields,
    )
    if args.emit_json:
        _emit_json(result)
    else:
        _emit_text(result, strict=args.strict)
    return 0 if result.ok(strict=args.strict) else 1


# --------------------------------------------------------------------------- exr scaffolding (test helper)


def _write_minimal_exr(path: Path, *, width: int = 4, height: int = 4) -> None:
    """Write a tiny single-channel float32 EXR file.

    This is a vendored, dependency-free EXR writer used purely by tests so we
    don't need OpenEXR installed to construct fixtures. The format is the
    minimum-viable scanline EXR: one ``Y`` channel of float32, no compression.

    Reference: OpenEXR file format spec (chapters on header + scanline data).
    """

    # --- header ---
    def attr(name: str, type_name: str, payload: bytes) -> bytes:
        return (
            name.encode("ascii")
            + b"\x00"
            + type_name.encode("ascii")
            + b"\x00"
            + struct.pack("<I", len(payload))
            + payload
        )

    # channels: one channel "Y" of pixel type FLOAT (=2). Channel record:
    # name\0 + i32 pixel_type + u8 pLinear + 3 bytes reserved + i32 xSampling + i32 ySampling
    chan = (
        b"Y\x00"
        + struct.pack("<i", 2)  # FLOAT
        + struct.pack("<B", 0)  # pLinear
        + b"\x00\x00\x00"
        + struct.pack("<i", 1)
        + struct.pack("<i", 1)
    )
    channels_payload = chan + b"\x00"  # trailing null record terminator
    # box2i: xMin yMin xMax yMax (i32 each)
    box = struct.pack("<iiii", 0, 0, width - 1, height - 1)
    line_order = struct.pack("<B", 0)  # increasing y
    # v2f screenWindowCenter, float screenWindowWidth, v2f pixelAspectRatio
    swc = struct.pack("<ff", 0.0, 0.0)
    sww = struct.pack("<f", 1.0)
    par = struct.pack("<f", 1.0)
    # compression: NO_COMPRESSION = 0
    comp = struct.pack("<B", 0)

    header = b""
    header += attr("channels", "chlist", channels_payload)
    header += attr("compression", "compression", comp)
    header += attr("dataWindow", "box2i", box)
    header += attr("displayWindow", "box2i", box)
    header += attr("lineOrder", "lineOrder", line_order)
    header += attr("pixelAspectRatio", "float", par)
    header += attr("screenWindowCenter", "v2f", swc)
    header += attr("screenWindowWidth", "float", sww)
    header += b"\x00"  # header terminator

    # --- scanline offset table + data ---
    # Each scanline block: i32 y, i32 dataSize, dataSize bytes (FLOAT pixels)
    pixel_bytes = width * 4
    # offsets are absolute offsets to each scanline block in the file.
    magic_and_version = struct.pack("<I", 20000630) + struct.pack("<I", 2)  # version 2, no flags
    # offset of first scanline = len(magic+version) + len(header) + 8*height bytes
    base = len(magic_and_version) + len(header) + 8 * height
    offsets = b"".join(struct.pack("<Q", base + y * (8 + pixel_bytes)) for y in range(height))
    blocks = b""
    # All-zero pixel buffer satisfies "single channel float32".
    pixels = b"\x00" * pixel_bytes
    for y in range(height):
        blocks += struct.pack("<i", y) + struct.pack("<i", pixel_bytes) + pixels

    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("wb") as fh:
        fh.write(magic_and_version)
        fh.write(header)
        fh.write(offsets)
        fh.write(blocks)


if __name__ == "__main__":
    raise SystemExit(main())
