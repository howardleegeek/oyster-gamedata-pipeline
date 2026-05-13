"""Tarball — the central buyer-facing class.

Usage:

    tar = Tarball.from_path("clip-00042_v1.tar.gz")
    tar.systeminfo.width                        # → 1920
    list(tar.action_camera)[0].camera_intrinsics.is_pinhole   # → True
    for frame_idx, depth_array in tar.depth:
        ...
    report = tar.validate()
    summary = tar.metadata_summary()

Design notes
------------
* **Lazy parsing** — opening a tarball just inspects the structure. The
  9000-frame action_camera.json (~7 MB JSON) is parsed only when
  ``.action_camera`` is accessed. Depth frames are streamed one at a
  time via the ``.depth`` iterator.
* **Extraction directory** — buyer tarballs are 0.5–1.5 GB so we extract
  on first access into a configurable cache dir (default: a tempdir that
  the SDK cleans up on ``.close()`` / context-manager exit).
* **Already-extracted directories** are also supported:
  ``Tarball.from_path("clip-00042/")`` skips extraction entirely.
* **Optional deps are lazy**: ``cv2`` only loaded if you call
  ``.video.open_cv2()``, ``openpyxl`` only when ``.gameinfo`` is accessed,
  ``OpenEXR``/``numpy`` only when iterating depth.
"""

from __future__ import annotations

import contextlib
import json
import logging
import shutil
import tarfile
import tempfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterator, List, Optional, Tuple, Union

from .errors import (
    DependencyMissingError,
    SchemaValidationError,
    TarballNotFoundError,
    TarballStructureError,
)
from .lint_report import LintReport, run_lint
from .schema import (
    ActionCameraFrame,
    Gameinfo,
    Systeminfo,
    parse_action_camera,
)

logger = logging.getLogger("oyster_gamedata_sdk")

# Required files per BUYER_SPEC_V1 (file 0..4 of PRD p7)
_REQUIRED_FILES = ("video.mp4", "systeminfo.json", "action_camera.json", "gameinfo.xlsx")
_REQUIRED_DIRS = ("depth",)


# ---------------------------------------------------------------------------
# Video accessor
# ---------------------------------------------------------------------------


class Video:
    """Lightweight video accessor.

    The buyer's training pipeline may want either:
      * the raw file path (for ffmpeg / decord / torchcodec ingestion), or
      * a ``cv2.VideoCapture`` if OpenCV is installed.

    Both are exposed without forcing the dep.
    """

    def __init__(self, path: Path):
        self._path = path

    @property
    def path(self) -> Path:
        """Absolute filesystem path to video.mp4."""
        return self._path

    def open_cv2(self) -> Any:
        """Open the video with cv2.VideoCapture. Raises if opencv-python is missing."""
        try:
            import cv2  # type: ignore
        except ImportError as exc:
            raise DependencyMissingError(
                "cv2 is required for Video.open_cv2(); install opencv-python"
            ) from exc
        cap = cv2.VideoCapture(str(self._path))
        if not cap.isOpened():
            raise TarballStructureError(f"cv2 could not open {self._path}")
        return cap

    def __fspath__(self) -> str:
        return str(self._path)

    def __str__(self) -> str:
        return str(self._path)


# ---------------------------------------------------------------------------
# Depth iterator
# ---------------------------------------------------------------------------


@dataclass
class DepthFrame:
    """One depth sample: (frame_index, single-channel float32 ndarray, file_path)."""

    frame_index: int
    array: Any  # numpy.ndarray, lazily-typed to avoid forcing numpy import here
    path: Path


class DepthSequence:
    """Iterator + indexer over depth/*.exr.

    Buyers can:
      * ``for fi, depth in tarball.depth:`` to stream
      * ``tarball.depth[42]`` to grab one
      * ``len(tarball.depth)`` to count
    """

    def __init__(self, depth_dir: Path):
        if not depth_dir.is_dir():
            raise TarballStructureError(f"depth/ directory missing: {depth_dir}")
        self._dir = depth_dir
        self._files: List[Path] = sorted(depth_dir.glob("*.exr"))

    def __len__(self) -> int:
        return len(self._files)

    def __iter__(self) -> Iterator[Tuple[int, Any]]:
        for i, path in enumerate(self._files):
            yield i, self._load(path)

    def __getitem__(self, idx: int) -> Any:
        if idx < 0:
            idx += len(self._files)
        if not 0 <= idx < len(self._files):
            raise IndexError(f"depth frame {idx} out of range (have {len(self._files)})")
        return self._load(self._files[idx])

    def frames(self) -> Iterator[DepthFrame]:
        """Iterator returning :class:`DepthFrame` objects (with path)."""
        for i, path in enumerate(self._files):
            yield DepthFrame(i, self._load(path), path)

    def paths(self) -> List[Path]:
        """File paths only — useful if the buyer wants to read EXRs themselves."""
        return list(self._files)

    @staticmethod
    def _load(path: Path) -> Any:
        try:
            import numpy as np  # type: ignore
        except ImportError as exc:
            raise DependencyMissingError(
                "numpy is required to read depth frames"
            ) from exc
        try:
            import OpenEXR  # type: ignore
        except ImportError as exc:
            raise DependencyMissingError(
                "OpenEXR is required to read depth frames; install OpenEXR"
            ) from exc

        f = OpenEXR.InputFile(str(path))
        header = f.header()
        dw = header["dataWindow"]
        w = dw.max.x - dw.min.x + 1
        h = dw.max.y - dw.min.y + 1
        channels = list(header["channels"].keys())
        # PRD requires single channel "Z" but we accept the first channel as
        # a fallback for older fixtures.
        chan = "Z" if "Z" in channels else channels[0]
        raw = f.channel(chan)
        arr = np.frombuffer(raw, dtype=np.float32).reshape(h, w)
        return arr


# ---------------------------------------------------------------------------
# Metadata summary (cheap, no heavy parsing)
# ---------------------------------------------------------------------------


@dataclass
class MetadataSummary:
    """Quick-look clip statistics — cheap to compute, useful for triage."""

    clip_root: Path
    tarball_path: Optional[Path]
    video_path: Path
    video_size_bytes: int
    systeminfo_resolution: Tuple[int, int]
    systeminfo_game: str
    n_action_frames: int
    n_depth_frames: int
    fps_first: Optional[float]
    route_type_distribution: Dict[int, int]

    def to_dict(self) -> Dict[str, Any]:
        return {
            "clip_root": str(self.clip_root),
            "tarball_path": str(self.tarball_path) if self.tarball_path else None,
            "video": {"path": str(self.video_path), "size_bytes": self.video_size_bytes},
            "systeminfo": {
                "resolution": list(self.systeminfo_resolution),
                "game": self.systeminfo_game,
            },
            "action_camera": {
                "n_frames": self.n_action_frames,
                "fps_first": self.fps_first,
                "route_type_distribution": self.route_type_distribution,
            },
            "depth": {"n_frames": self.n_depth_frames},
        }


# ---------------------------------------------------------------------------
# Tarball
# ---------------------------------------------------------------------------


class Tarball:
    """Buyer-spec v1 tarball / extracted-directory accessor.

    Construct via :meth:`from_path`. Acts as a context manager and frees
    its tempdir on ``__exit__``.
    """

    def __init__(self, root: Path, *, source: Optional[Path] = None, owns_tempdir: bool = False):
        self._root = root.resolve()
        self._source = source
        self._owns_tempdir = owns_tempdir
        self._closed = False

        # Cached parsed payloads (lazy)
        self._systeminfo: Optional[Systeminfo] = None
        self._action_camera: Optional[List[ActionCameraFrame]] = None
        self._gameinfo: Optional[Gameinfo] = None
        self._depth: Optional[DepthSequence] = None

        self._verify_structure()

    # -- constructors --------------------------------------------------------

    @classmethod
    def from_path(
        cls,
        path: Union[str, Path],
        *,
        extract_to: Optional[Union[str, Path]] = None,
    ) -> "Tarball":
        """Open a ``.tar.gz`` file or an already-extracted clip directory.

        Args:
            path: tarball path or directory.
            extract_to: where to extract (defaults to a tempdir we own
                and clean up). Useful when the buyer wants to keep the
                extracted clip on disk.

        Returns:
            :class:`Tarball` instance.
        """
        p = Path(path)
        if not p.exists():
            raise TarballNotFoundError(f"{p} does not exist")

        if p.is_dir():
            return cls(p, source=p, owns_tempdir=False)

        # Tarball — extract.
        owns = extract_to is None
        target = Path(extract_to) if extract_to else Path(tempfile.mkdtemp(prefix="oyster_gamedata_"))
        target.mkdir(parents=True, exist_ok=True)
        try:
            with tarfile.open(p, "r:*") as tf:
                _safe_extract(tf, target)
        except tarfile.TarError as exc:
            if owns:
                shutil.rmtree(target, ignore_errors=True)
            raise TarballStructureError(f"failed to extract {p}: {exc}") from exc

        # Find the clip root inside the extraction. Two layouts seen:
        #   1. tarball/<files at root>            (released sample)
        #   2. tarball/<clip_id>/<files inside>   (vendor submissions per SUBMISSION_FORMAT.md)
        root = _resolve_clip_root(target)
        return cls(root, source=p, owns_tempdir=owns)

    # -- structural ----------------------------------------------------------

    def _verify_structure(self) -> None:
        existing = {entry.name for entry in self._root.iterdir()}
        missing = [f for f in _REQUIRED_FILES if f not in existing] + [
            d for d in _REQUIRED_DIRS if not (self._root / d).is_dir()
        ]
        if missing:
            raise TarballStructureError(
                f"Tarball at {self._root} missing required entries: {missing}"
            )

    @property
    def root(self) -> Path:
        """Absolute path of the extracted clip directory."""
        return self._root

    @property
    def source(self) -> Optional[Path]:
        """Original .tar.gz path (None if we opened a directory directly)."""
        return self._source

    # -- accessors -----------------------------------------------------------

    @property
    def video(self) -> Video:
        return Video(self._root / "video.mp4")

    @property
    def systeminfo(self) -> Systeminfo:
        if self._systeminfo is None:
            path = self._root / "systeminfo.json"
            try:
                payload = json.loads(path.read_text())
            except json.JSONDecodeError as exc:
                raise SchemaValidationError(f"systeminfo.json: {exc}") from exc
            self._systeminfo = Systeminfo.from_dict(payload)
        return self._systeminfo

    @property
    def action_camera(self) -> List[ActionCameraFrame]:
        if self._action_camera is None:
            path = self._root / "action_camera.json"
            try:
                payload = json.loads(path.read_text())
            except json.JSONDecodeError as exc:
                raise SchemaValidationError(f"action_camera.json: {exc}") from exc
            if not isinstance(payload, list):
                raise SchemaValidationError(
                    f"action_camera.json: expected JSON list, got {type(payload).__name__}"
                )
            self._action_camera = parse_action_camera(payload, strict=True)
        return self._action_camera

    @property
    def gameinfo(self) -> Gameinfo:
        if self._gameinfo is None:
            self._gameinfo = self._parse_gameinfo()
        return self._gameinfo

    def _parse_gameinfo(self) -> Gameinfo:
        try:
            import openpyxl  # type: ignore
        except ImportError as exc:
            raise DependencyMissingError(
                "openpyxl is required to read gameinfo.xlsx; install openpyxl"
            ) from exc
        wb = openpyxl.load_workbook(self._root / "gameinfo.xlsx", read_only=True, data_only=True)
        sheet_name = wb.sheetnames[0]
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return Gameinfo(sheet_name=sheet_name, columns=[], rows=[])

        header_row = rows[0]
        columns = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(header_row)]
        parsed_rows: List[Dict[str, Any]] = []
        for raw in rows[1:]:
            # Skip fully-empty rows
            if all(v is None for v in raw):
                continue
            parsed_rows.append({columns[i]: raw[i] for i in range(min(len(columns), len(raw)))})
        return Gameinfo(sheet_name=sheet_name, columns=columns, rows=parsed_rows)

    @property
    def depth(self) -> DepthSequence:
        if self._depth is None:
            self._depth = DepthSequence(self._root / "depth")
        return self._depth

    # -- ops -----------------------------------------------------------------

    def validate(self) -> LintReport:
        """Run the 24-criterion buyer-spec validator on this clip.

        Falls back to a structural-only 5-check validator if the project
        lint script (``bin/lint_v3_prd_grounded.py``) is not reachable.
        """
        return run_lint(self._root)

    def metadata_summary(self) -> MetadataSummary:
        """Quick-look summary — opens systeminfo + counts files. Cheap."""
        video_path = self._root / "video.mp4"
        depth_dir = self._root / "depth"
        n_depth = sum(1 for _ in depth_dir.glob("*.exr"))

        # action_camera summary — but cheap: peek at frame count & route mix
        ac_path = self._root / "action_camera.json"
        try:
            ac_payload = json.loads(ac_path.read_text())
        except (json.JSONDecodeError, OSError):
            ac_payload = []
        n_frames = len(ac_payload) if isinstance(ac_payload, list) else 0
        fps_first: Optional[float] = None
        route_counts: Dict[int, int] = {}
        if isinstance(ac_payload, list) and ac_payload:
            first = ac_payload[0]
            if isinstance(first, dict):
                v = first.get("fps")
                if isinstance(v, (int, float)):
                    fps_first = float(v)
            for entry in ac_payload:
                if not isinstance(entry, dict):
                    continue
                rt = entry.get("route_type")
                if isinstance(rt, int):
                    route_counts[rt] = route_counts.get(rt, 0) + 1

        return MetadataSummary(
            clip_root=self._root,
            tarball_path=self._source,
            video_path=video_path,
            video_size_bytes=video_path.stat().st_size if video_path.is_file() else 0,
            systeminfo_resolution=(self.systeminfo.width, self.systeminfo.height),
            systeminfo_game=self.systeminfo.game_process_name,
            n_action_frames=n_frames,
            n_depth_frames=n_depth,
            fps_first=fps_first,
            route_type_distribution=route_counts,
        )

    # -- lifecycle -----------------------------------------------------------

    def close(self) -> None:
        """Free any owned tempdir."""
        if self._closed:
            return
        self._closed = True
        if self._owns_tempdir and self._root.exists():
            shutil.rmtree(self._root.parent if self._root.parent.name.startswith("oyster_gamedata_")
                          else self._root, ignore_errors=True)

    def __enter__(self) -> "Tarball":
        return self

    def __exit__(self, *exc: Any) -> None:
        self.close()

    def __repr__(self) -> str:
        return f"<Tarball root={self._root!s} source={self._source!s}>"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _safe_extract(tar: tarfile.TarFile, target: Path) -> None:
    """Extract a tarball with traversal-attack protection (CVE-2007-4559)."""
    target = target.resolve()
    members = []
    for member in tar.getmembers():
        member_path = (target / member.name).resolve()
        if not str(member_path).startswith(str(target)):
            raise TarballStructureError(
                f"refusing path-traversal entry in tarball: {member.name}"
            )
        members.append(member)
    # Use data filter when available (Python 3.12+), else explicit allow.
    extract_kwargs: Dict[str, Any] = {}
    if hasattr(tarfile, "data_filter"):
        extract_kwargs["filter"] = "data"
    tar.extractall(target, members=members, **extract_kwargs)


def _resolve_clip_root(extract_dir: Path) -> Path:
    """Return the directory that contains the 5 required files.

    Tarballs are sometimes laid out as ``<clip_id>/video.mp4`` and
    sometimes as ``video.mp4`` at the top level. We detect either.
    """
    if (extract_dir / "video.mp4").is_file():
        return extract_dir
    # Look one level down
    candidates = [p for p in extract_dir.iterdir() if p.is_dir() and (p / "video.mp4").is_file()]
    if len(candidates) == 1:
        return candidates[0]
    if len(candidates) > 1:
        raise TarballStructureError(
            f"multiple clip roots found in tarball: {[str(p) for p in candidates]}"
        )
    raise TarballStructureError(
        f"could not find video.mp4 in {extract_dir} or any immediate subdirectory"
    )
